"""Map each metric to its corresponding data section"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Old Bridge Focused Equity Fund.xlsm', data_only=True)
ws = wb.active

print("=== MAPPING METRICS TO DATA SECTIONS ===\n")

# All 27 metric rows
metrics_info = [
    (37, 'PATM', 146),
    (38, 'QoQ', 216),
    (39, 'YoY', 76),
    (40, '6 year CAGR (Revenue)', 76),
    (41, '(blank)', None),
    (42, 'Current PE', 48),
    (43, '2 year average (PE)', None),
    (44, '5 year average (PE)', None),
    (45, '2 years - Reval / Deval (PE)', None),
    (46, '5 years - Reval / Deval (PE)', None),
    (47, '(blank)', None),
    (48, 'Current PR', None),
    (49, '2 year average (PR)', None),
    (50, '5 year average (PR)', None),
    (51, '2 years - Reval / Deval (PR)', None),
    (52, '5 years - Reval / Deval (PR)', None),
    (53, '(blank)', None),
    (54, '10 quarter- PR- low', None),
    (55, '10 quarter- PR- high', None),
    (56, '(blank)', None),
    (57, 'Alpha over the bond- CAGR', None),
    (58, 'Alpha- Absolute', None),
    (59, 'PE Yield', None),
    (60, 'Growth', None),
    (61, 'Bond Rate', None),
    (62, '(blank)', None),
    (63, '(blank)', None),
]

for row, label, known_col in metrics_info:
    if known_col:
        col = known_col
    else:
        # Find first non-empty value
        col = None
        for c in range(1, min(400, ws.max_column + 1)):
            val = ws.cell(row, c).value
            if val is not None and val != '' and val != label:
                col = c
                break

    if col:
        header = ws.cell(8, col).value
        row6 = ws.cell(6, col).value
        row7 = ws.cell(7, col).value
        value = ws.cell(row, col).value

        print(f"Row {row}: {label}")
        print(f"  Column: {get_column_letter(col)} (index {col})")
        print(f"  Row 8 header: {header}")
        print(f"  Row 6 category: {row6}")
        print(f"  Row 7 subcategory: {row7}")
        print(f"  Sample value: {value}")
        print()
    else:
        print(f"Row {row}: {label}")
        print(f"  No data found")
        print()
