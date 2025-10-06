"""Check sample file metric rows structure"""
from openpyxl import load_workbook

wb = load_workbook('Old Bridge Focused Equity Fund.xlsm', data_only=True)
ws = wb.active

print('Sample File - Metrics Rows (Row 37-63):')
print()

for row in range(37, 64):
    label = ws.cell(row, 1).value
    val_col20 = ws.cell(row, 20).value
    val_col21 = ws.cell(row, 21).value
    val_col22 = ws.cell(row, 22).value
    print(f'Row {row}: "{label}" | Col20={val_col20}, Col21={val_col21}, Col22={val_col22}')
