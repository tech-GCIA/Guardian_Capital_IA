"""
Analyze and compare Excel file formats
"""
from openpyxl import load_workbook
import os

def analyze_excel_structure(filepath, name):
    """Analyze the structure of an Excel file"""
    print(f"\n{'='*80}")
    print(f"ANALYZING: {name}")
    print(f"{'='*80}")

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    print(f"\nWorksheet: {ws.title}")
    print(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns")

    # Check first 10 rows structure
    print(f"\n--- FIRST 10 ROWS ---")
    for row in range(1, min(11, ws.max_row + 1)):
        row_data = [ws.cell(row, col).value for col in range(1, min(25, ws.max_column + 1))]
        # Truncate long values
        display_data = []
        for val in row_data:
            if val is None:
                display_data.append('None')
            elif isinstance(val, str) and len(val) > 30:
                display_data.append(val[:27] + '...')
            else:
                display_data.append(str(val)[:30])
        print(f"Row {row:2d}: {display_data[:10]}")

    # Check for fund name in row 1
    fund_name = ws.cell(1, 1).value
    print(f"\n--- METADATA ---")
    print(f"Fund Name (Row 1, Col A): {fund_name}")

    # Check row 2 for column indicators
    row_2_sample = [ws.cell(2, col).value for col in range(1, 11)]
    print(f"Row 2 (Column indicators): {row_2_sample}")

    # Check row 8 for headers
    row_8_headers = [ws.cell(8, col).value for col in range(1, min(30, ws.max_column + 1))]
    print(f"\n--- ROW 8 HEADERS (first 30) ---")
    for i, header in enumerate(row_8_headers[:30], 1):
        if header:
            print(f"  Col {i}: {header}")

    # Find where stock data starts and ends
    print(f"\n--- DATA ROWS ---")
    stock_data_start = None
    stock_data_end = None
    totals_row = None
    metrics_start = None

    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row, 1).value

        if cell_a and isinstance(cell_a, int) and cell_a == 1:
            stock_data_start = row
            print(f"Stock data starts at row {row} (S.No = 1)")

        if cell_a == 'TOTALS':
            totals_row = row
            stock_data_end = row - 1
            print(f"TOTALS row at row {row}")
            print(f"Stock data ends at row {stock_data_end}")

        # Check for metric rows after TOTALS
        if totals_row and row > totals_row:
            if cell_a and isinstance(cell_a, str) and cell_a not in ['', 'None']:
                if metrics_start is None:
                    metrics_start = row
                    print(f"Metrics rows start at row {row}")

    # Count metrics rows
    if metrics_start:
        metrics_count = ws.max_row - metrics_start + 1
        print(f"Metrics rows count: {metrics_count}")

        # Show first 5 metric labels
        print(f"\n--- FIRST 10 METRIC ROW LABELS ---")
        for row in range(metrics_start, min(metrics_start + 10, ws.max_row + 1)):
            label = ws.cell(row, 1).value
            # Get a few values from this metric row
            values = [ws.cell(row, col).value for col in range(2, 7)]
            print(f"Row {row}: {label} | Values: {values}")

    # Check TOTALS row values
    if totals_row:
        print(f"\n--- TOTALS ROW VALUES ---")
        totals_values = [ws.cell(totals_row, col).value for col in range(1, min(20, ws.max_column + 1))]
        for i, val in enumerate(totals_values[:20], 1):
            if val not in [None, '']:
                print(f"  Col {i}: {val}")

    # Check for separators (empty columns)
    print(f"\n--- SEPARATOR COLUMNS (first 50) ---")
    separators = []
    for col in range(1, min(51, ws.max_column + 1)):
        # Check if column is empty in data rows
        if stock_data_start:
            test_row = stock_data_start
            cell_val = ws.cell(test_row, col).value
            if cell_val is None or cell_val == '':
                separators.append(col)

    print(f"Separator columns: {separators[:20]}")

    print(f"\n{'='*80}\n")

# Analyze all files
files = [
    ("Old Bridge Focused Equity Fund.xlsm", "SAMPLE FILE (Old Bridge)"),
    ("Aditya Birla SL Balanced Advantage Fund(G) (INF084M01AB8)_Portfolio_Analysis_20251006_204221.xlsx", "GENERATED - Aditya Birla"),
    ("360 ONE ELSS Tax Saver Nifty 50 Index Fund-Reg(G) (INF579M01AL6)_Portfolio_Analysis_20251006_204027.xlsx", "GENERATED - 360 ONE ELSS"),
    ("360 ONE Flexicap Fund-Reg(G) (INF579M01AP7)_Portfolio_Analysis_20251006_204122.xlsx", "GENERATED - 360 ONE Flexicap"),
]

for filepath, name in files:
    if os.path.exists(filepath):
        try:
            analyze_excel_structure(filepath, name)
        except Exception as e:
            print(f"ERROR analyzing {name}: {e}")
    else:
        print(f"FILE NOT FOUND: {filepath}")
