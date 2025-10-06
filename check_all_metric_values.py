"""Check if all metrics actually have values somewhere"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Old Bridge Focused Equity Fund.xlsm', data_only=True)
ws = wb.active

print("=== CHECKING ALL 27 METRIC ROWS FOR DATA VALUES ===\n")

metric_rows = range(37, 64)  # Rows 37-63

for row in metric_rows:
    label = ws.cell(row, 1).value

    # Search for any non-empty value in this row
    values_found = []
    for col in range(2, min(400, ws.max_column + 1)):
        val = ws.cell(row, col).value
        if val is not None and val != '':
            values_found.append((col, val))

    if values_found:
        first_col, first_val = values_found[0]
        print(f"Row {row}: '{label}'")
        print(f"  First value at Col {get_column_letter(first_col)} ({first_col}): {first_val}")
        print(f"  Total values found: {len(values_found)}")
    else:
        print(f"Row {row}: '{label}'")
        print(f"  NO VALUES FOUND (empty row)")
    print()

# Also check if there might be formulas (not captured by data_only=True)
print("\n=== NOTE ===")
print("If a row shows 'NO VALUES FOUND' but should have data,")
print("it might contain Excel formulas that aren't captured by data_only=True")
print("Open the file manually to check for formulas.")
