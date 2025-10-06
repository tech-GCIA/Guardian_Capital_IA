"""Analyze where metric values are in the sample file"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Old Bridge Focused Equity Fund.xlsm', data_only=True)
ws = wb.active

print("=== ANALYZING METRIC ROW STRUCTURE ===\n")

# Check rows 37-45 (first few metric rows)
metric_rows = [
    (37, 'PATM', 'EP'),
    (38, 'QoQ', 'HH'),
    (39, 'YoY', None),
    (40, '6 year CAGR', None),
    (42, 'Current PE', None),
]

for row_num, label, expected_col in metric_rows:
    print(f"\nRow {row_num}: {label}")
    print(f"  Expected column: {expected_col if expected_col else 'Unknown'}")

    # Find first non-None value
    first_value_col = None
    for col in range(1, min(300, ws.max_column + 1)):
        val = ws.cell(row_num, col).value
        if val is not None and val != '' and val != label:
            first_value_col = col
            print(f"  First value at Col {get_column_letter(col)} (index {col}): {val}")
            break

    if first_value_col:
        # Show next 5 values
        print(f"  Next 5 values:")
        for i in range(5):
            col = first_value_col + i
            val = ws.cell(row_num, col).value
            print(f"    Col {get_column_letter(col)} ({col}): {val}")
    else:
        print("  No values found")

# Check column EP (column index for 'EP')
print("\n=== CHECKING SPECIFIC COLUMNS ===")

# EP column
ep_col = None
for col in range(1, ws.max_column + 1):
    if get_column_letter(col) == 'EP':
        ep_col = col
        break

if ep_col:
    print(f"\nColumn EP (index {ep_col}):")
    print(f"  Row 8 (header): {ws.cell(8, ep_col).value}")
    print(f"  Row 9 (first stock): {ws.cell(9, ep_col).value}")
    print(f"  Row 37 (PATM): {ws.cell(37, ep_col).value}")

# HH column
hh_col = None
for col in range(1, ws.max_column + 1):
    if get_column_letter(col) == 'HH':
        hh_col = col
        break

if hh_col:
    print(f"\nColumn HH (index {hh_col}):")
    print(f"  Row 8 (header): {ws.cell(8, hh_col).value}")
    print(f"  Row 9 (first stock): {ws.cell(9, hh_col).value}")
    print(f"  Row 38 (QoQ): {ws.cell(38, hh_col).value}")

# Check Row 6-7 for multilevel headers
print("\n=== MULTILEVEL HEADERS (Rows 6-7) ===")

print("\nRow 6 (first 30 columns):")
for col in range(1, 31):
    val = ws.cell(6, col).value
    if val:
        print(f"  Col {get_column_letter(col)} ({col}): {val}")

print("\nRow 7 (first 30 columns):")
for col in range(1, 31):
    val = ws.cell(7, col).value
    if val:
        print(f"  Col {get_column_letter(col)} ({col}): {val}")

# Check where "EP" and "HH" columns are in relation to periods
print("\n=== PERIOD COLUMNS AROUND EP ===")
if ep_col:
    for col in range(max(1, ep_col - 3), min(ws.max_column + 1, ep_col + 4)):
        header = ws.cell(8, col).value
        print(f"  Col {get_column_letter(col)} ({col}): {header}")
