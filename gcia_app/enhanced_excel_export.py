"""
Enhanced Excel Export with Portfolio Analysis Metrics
=====================================================

This module provides the enhanced Excel export functionality that includes
calculated portfolio analysis metrics in the Portfolio Analysis format.
Now aligned with Stock Base Sheet export structure for consistency.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils import timezone
from io import BytesIO
from .metrics_calculator import DynamicHeaderGenerator
from .models import FundMetricsLog, AMCFundScheme, FundHolding, PortfolioMetricsLog
from .dynamic_admin_export import BlockBasedExportGenerator
from datetime import datetime
import logging

# Import calculation functions for portfolio metrics
from .excel_calc_functions import (
    calculate_patm_from_totals, calculate_qoq_from_totals,
    calculate_yoy_from_totals, calculate_6yr_cagr_from_totals,
    calculate_pe_pr_from_totals, calculate_pe_pr_averages_from_totals,
    calculate_reval_deval_from_totals, calculate_pr_10q_extremes_from_totals,
    calculate_pe_yield_from_totals, calculate_growth_from_totals, get_bond_rate,
    build_section_column_mapping, create_metric_row
)

logger = logging.getLogger(__name__)


class FundPortfolioExportGenerator(BlockBasedExportGenerator):
    """
    Extends BlockBasedExportGenerator to add fund-specific columns and portfolio metrics.
    Maintains same block-based structure as Stock Base Sheet export for consistency.
    """

    def __init__(self, scheme):
        """Initialize with specific fund scheme"""
        super().__init__()
        self.scheme = scheme
        self.section_start_columns = {}  # Track where each data section starts for metric rows

    def _define_block_structure(self, periods):
        """
        Override to add fund-specific data block after basic info.
        Maintains same structure as parent but inserts Fund Data block.
        """
        blocks = [
            {
                'name': 'basic_info',
                'type': 'fixed',
                'size': 18,  # Expanded: removed S.No, added Market Cap, moved fund data here, added Share Price & Free Float
                'columns': [
                    'Company Name', 'Accord Code', 'Sector', 'Cap',
                    'Market Cap', 'Weights', 'Factor', 'Value', 'No.of shares',
                    'Share Price', 'Free Float',
                    '6 Year CAGR', 'TTM', '6 Year CAGR', 'TTM',
                    'Current', '2 Yr Avg', 'Reval/deval'
                ],
                'row6_label': None,  # First 11 columns have no category
                'row7_label': None
            },
            {'name': 'sep_0', 'type': 'separator', 'size': 1},
            {
                'name': 'market_cap',
                'type': 'dynamic',
                'data_type': 'market_cap',
                'periods': periods['market_cap_dates'],
                'row6_label': 'Market Cap (in crores)',
                'row7_label': 'Market Cap (in crores)'
            },
            {'name': 'sep_2', 'type': 'separator', 'size': 1},
            {
                'name': 'market_cap_ff',
                'type': 'dynamic',
                'data_type': 'market_cap_free_float',
                'periods': periods['market_cap_dates'],
                'row6_label': 'Market Cap- Free Float (in crores)',
                'row7_label': 'Market Cap- Free Float (in crores)'
            },
            {'name': 'sep_3', 'type': 'separator', 'size': 1},
            {
                'name': 'ttm_revenue',
                'type': 'dynamic',
                'data_type': 'ttm_revenue',
                'periods': periods['ttm_periods'],
                'row6_label': 'TTM Revenue',
                'row7_label': 'TTM Revenue'
            },
            {'name': 'sep_4', 'type': 'separator', 'size': 1},
            {
                'name': 'ttm_revenue_ff',
                'type': 'dynamic',
                'data_type': 'ttm_revenue_free_float',
                'periods': periods['ttm_periods'],
                'row6_label': 'TTM Revenue- Free Float',
                'row7_label': 'TTM Revenue- Free Float'
            },
            {'name': 'sep_5', 'type': 'separator', 'size': 1},
            {
                'name': 'ttm_pat',
                'type': 'dynamic',
                'data_type': 'ttm_pat',
                'periods': periods['ttm_periods'],
                'row6_label': 'TTM PAT',
                'row7_label': 'TTM PAT'
            },
            {'name': 'sep_6', 'type': 'separator', 'size': 1},
            {
                'name': 'ttm_pat_ff',
                'type': 'dynamic',
                'data_type': 'ttm_pat_free_float',
                'periods': periods['ttm_periods'],
                'row6_label': 'TTM PAT- Free Float',
                'row7_label': 'TTM PAT- Free Float'
            },
            {'name': 'sep_7', 'type': 'separator', 'size': 1},
            {
                'name': 'quarterly_revenue',
                'type': 'dynamic',
                'data_type': 'quarterly_revenue',
                'periods': periods['quarterly_periods'],
                'row6_label': 'Quarterly- Revenue',
                'row7_label': 'Quarterly- Revenue'
            },
            {'name': 'sep_8', 'type': 'separator', 'size': 1},
            {
                'name': 'quarterly_revenue_ff',
                'type': 'dynamic',
                'data_type': 'quarterly_revenue_free_float',
                'periods': periods['quarterly_periods'],
                'row6_label': 'Quarterly- Revenue- Free Float',
                'row7_label': 'Quarterly- Revenue- Free Float'
            },
            {'name': 'sep_9', 'type': 'separator', 'size': 1},
            {
                'name': 'quarterly_pat',
                'type': 'dynamic',
                'data_type': 'quarterly_pat',
                'periods': periods['quarterly_periods'],
                'row6_label': 'Quarterly- PAT',
                'row7_label': 'Quarterly- PAT'
            },
            {'name': 'sep_10', 'type': 'separator', 'size': 1},
            {
                'name': 'quarterly_pat_ff',
                'type': 'dynamic',
                'data_type': 'quarterly_pat_free_float',
                'periods': periods['quarterly_periods'],
                'row6_label': 'Quarterly- PAT- Free Float',
                'row7_label': 'Quarterly- PAT- Free Float'
            },
            {'name': 'sep_11', 'type': 'separator', 'size': 1},
            {
                'name': 'roce',
                'type': 'dynamic',
                'data_type': 'roce',
                'periods': periods['annual_years'],
                'row6_label': 'ROCE (%)',
                'row7_label': 'ROCE (%)'
            },
            {'name': 'sep_12', 'type': 'separator', 'size': 1},
            {
                'name': 'roe',
                'type': 'dynamic',
                'data_type': 'roe',
                'periods': periods['annual_years'],
                'row6_label': 'ROE (%)',
                'row7_label': 'ROE (%)'
            },
            {'name': 'sep_13', 'type': 'separator', 'size': 1},
            {
                'name': 'retention',
                'type': 'dynamic',
                'data_type': 'retention',
                'periods': periods['annual_years'],
                'row6_label': 'Retention (%)',
                'row7_label': 'Retention (%)'
            },
            {'name': 'sep_14', 'type': 'separator', 'size': 1},
            {
                'name': 'share_price',
                'type': 'dynamic',
                'data_type': 'share_price',
                'periods': periods['share_price_dates'],
                'row6_label': 'Share Price',
                'row7_label': 'Share Price'
            },
            {'name': 'sep_15', 'type': 'separator', 'size': 1},
            {
                'name': 'pr_ratio',
                'type': 'dynamic',
                'data_type': 'pr_ratio',
                'periods': periods['pr_dates'],
                'row6_label': 'PR',
                'row7_label': 'PR'
            },
            {'name': 'sep_16', 'type': 'separator', 'size': 1},
            {
                'name': 'pe_ratio',
                'type': 'dynamic',
                'data_type': 'pe_ratio',
                'periods': periods['pe_dates'],
                'row6_label': 'PE',
                'row7_label': 'PE'
            },
            {'name': 'sep_17', 'type': 'separator', 'size': 1},
            {
                'name': 'identifiers',
                'type': 'fixed',
                'size': 3,
                'columns': ['BSE Code', 'NSE Code', 'ISIN']
            }
        ]

        return blocks

    def collect_all_periods_for_fund(self):
        """
        Collect all available periods from holdings' stocks in this fund.
        Similar to parent's collect_all_periods_from_database but filtered to fund stocks.
        """
        from .models import StockMarketCap, StockTTMData, StockQuarterlyData, StockAnnualRatios, StockPrice

        # Get all stocks in this fund
        holdings = FundHolding.objects.filter(scheme=self.scheme).select_related('stock')
        stock_ids = [h.stock.stock_id for h in holdings]

        # Collect periods from these stocks only
        market_cap_dates = set(StockMarketCap.objects.filter(
            stock__stock_id__in=stock_ids
        ).values_list('date', flat=True).distinct())

        ttm_periods = set(StockTTMData.objects.filter(
            stock__stock_id__in=stock_ids
        ).values_list('period', flat=True).distinct())

        quarterly_periods = set(StockQuarterlyData.objects.filter(
            stock__stock_id__in=stock_ids
        ).values_list('period', flat=True).distinct())

        annual_years = set(StockAnnualRatios.objects.filter(
            stock__stock_id__in=stock_ids
        ).values_list('financial_year', flat=True).distinct())

        share_price_dates = set(StockPrice.objects.filter(
            stock__stock_id__in=stock_ids
        ).values_list('price_date', flat=True).distinct())

        # PR and PE dates from StockPrice
        pr_dates = set(StockPrice.objects.filter(
            stock__stock_id__in=stock_ids,
            pr_ratio__isnull=False
        ).values_list('price_date', flat=True).distinct())

        pe_dates = set(StockPrice.objects.filter(
            stock__stock_id__in=stock_ids,
            pe_ratio__isnull=False
        ).values_list('price_date', flat=True).distinct())

        # Sort all periods
        return {
            'market_cap_dates': sorted(market_cap_dates, reverse=True),
            'ttm_periods': sorted(ttm_periods, reverse=True),
            'quarterly_periods': sorted(quarterly_periods, reverse=True),
            'annual_years': sorted(annual_years, reverse=True),
            'share_price_dates': sorted(share_price_dates, reverse=True),
            'pr_dates': sorted(pr_dates, reverse=True),
            'pe_dates': sorted(pe_dates, reverse=True)
        }

    def collect_periods_for_filtered_holdings(self, filtered_holdings):
        """
        Collect all available periods from filtered holdings' stocks.
        Similar to collect_all_periods_for_fund but uses pre-filtered holdings QuerySet.

        Args:
            filtered_holdings: QuerySet of FundHolding objects (already filtered)

        Returns:
            Dict with sorted period lists for each data type
        """
        from .models import StockMarketCap, StockTTMData, StockQuarterlyData, StockAnnualRatios, StockPrice

        # Get stock IDs from filtered holdings
        stock_ids = [h.stock.stock_id for h in filtered_holdings]

        if not stock_ids:
            # Return empty periods if no stocks
            logger.warning("No stocks in filtered holdings - returning empty periods")
            return {
                'market_cap_dates': [],
                'ttm_periods': [],
                'quarterly_periods': [],
                'annual_years': [],
                'share_price_dates': [],
                'pr_dates': [],
                'pe_dates': []
            }

        # Collect periods from filtered stocks only
        market_cap_dates = set(StockMarketCap.objects.filter(
            stock__stock_id__in=stock_ids
        ).values_list('date', flat=True).distinct())

        ttm_periods = set(StockTTMData.objects.filter(
            stock__stock_id__in=stock_ids
        ).values_list('period', flat=True).distinct())

        quarterly_periods = set(StockQuarterlyData.objects.filter(
            stock__stock_id__in=stock_ids
        ).values_list('period', flat=True).distinct())

        annual_years = set(StockAnnualRatios.objects.filter(
            stock__stock_id__in=stock_ids
        ).values_list('financial_year', flat=True).distinct())

        share_price_dates = set(StockPrice.objects.filter(
            stock__stock_id__in=stock_ids
        ).values_list('price_date', flat=True).distinct())

        # PR and PE dates from StockPrice
        pr_dates = set(StockPrice.objects.filter(
            stock__stock_id__in=stock_ids,
            pr_ratio__isnull=False
        ).values_list('price_date', flat=True).distinct())

        pe_dates = set(StockPrice.objects.filter(
            stock__stock_id__in=stock_ids,
            pe_ratio__isnull=False
        ).values_list('price_date', flat=True).distinct())

        logger.info(f"Collected periods from {len(stock_ids)} filtered stocks")

        # Sort all periods
        return {
            'market_cap_dates': sorted(market_cap_dates, reverse=True),
            'ttm_periods': sorted(ttm_periods, reverse=True),
            'quarterly_periods': sorted(quarterly_periods, reverse=True),
            'annual_years': sorted(annual_years, reverse=True),
            'share_price_dates': sorted(share_price_dates, reverse=True),
            'pr_dates': sorted(pr_dates, reverse=True),
            'pe_dates': sorted(pe_dates, reverse=True)
        }

    def _generate_import_style_headers(self, blocks, total_columns):
        """
        Override parent to correctly position fundamental headers for Fund Analysis.
        Fund has 18 columns in basic_info, so fundamentals are at columns 11-17 (not 6-12).
        """
        # Call parent to get base headers
        headers = super()._generate_import_style_headers(blocks, total_columns)

        # Clear parent's incorrect fundamental headers at columns 6-10
        # Parent class hardcodes "Stock Fundamentals" at 6-12 for Stock Base Sheet
        # But Fund Analysis has different structure: fundamentals are at 11-17
        for col_idx in range(6, 11):
            if col_idx < total_columns:
                headers['row_6'][col_idx] = ''  # Clear incorrect section header
                headers['row_7'][col_idx] = ''  # Clear incorrect subcategory

        # Override the hardcoded fundamental headers from parent
        # Fund Analysis: Columns 0-10 are Company, Accord, Sector, Cap, Market Cap,
        #                Weights, Factor, Value, Shares, Price, Free Float (NO fundamentals header)
        # Fund Analysis: Columns 11-17 are Fundamentals (Revenue/PAT CAGR/TTM, PR ratios)
        fundamental_start_col = 11

        # Row 6: Section header spanning fundamental columns (11-17)
        fundamentals_section = "Stock wise Fundamentals and Valuations"
        for col_idx in range(fundamental_start_col, fundamental_start_col + 7):
            if col_idx < total_columns:
                headers['row_6'][col_idx] = fundamentals_section

        # Row 7: Subcategory labels
        if total_columns > 11:
            headers['row_7'][11] = "Revenue"   # 6 Year CAGR
        if total_columns > 12:
            headers['row_7'][12] = "Revenue"   # TTM
        if total_columns > 13:
            headers['row_7'][13] = "PAT"       # 6 Year CAGR
        if total_columns > 14:
            headers['row_7'][14] = "PAT"       # TTM
        if total_columns > 15:
            headers['row_7'][15] = "PR"        # Current
        if total_columns > 16:
            headers['row_7'][16] = "PR"        # 2 Yr Avg
        if total_columns > 17:
            headers['row_7'][17] = "PR"        # Reval/Deval

        return headers

    def populate_fund_stock_row(self, holding, blocks, total_columns):
        """
        Populate a single stock row with fund-specific data.
        Extends parent's populate_stock_row_header_driven with fund data.
        """
        stock = holding.stock
        row_data = [''] * total_columns
        current_col = 0

        # Calculate total portfolio value for factor calculation
        holdings = FundHolding.objects.filter(scheme=self.scheme)
        total_portfolio_value = sum(h.market_value or 0 for h in holdings if h.market_value)

        for block in blocks:
            if block['type'] == 'separator':
                # Empty column
                row_data[current_col] = ''
                current_col += 1

            elif block['type'] == 'fixed':
                if block['name'] == 'basic_info':
                    # 18 columns: Company info + Fund Data + Share Price + Free Float + Fundamentals
                    from .models import StockMarketCap, StockPrice, FundMetricsLog

                    # Col 1: Company Name (NO S.No!)
                    row_data[current_col] = stock.company_name

                    # Col 2: Accord Code
                    row_data[current_col + 1] = stock.accord_code or ''

                    # Col 3: Sector
                    row_data[current_col + 2] = stock.sector or 'Unknown'

                    # Col 4: Cap
                    row_data[current_col + 3] = stock.cap or 'Unknown'

                    # Col 5: Market Cap (latest)
                    market_cap_data = StockMarketCap.objects.filter(stock=stock).order_by('-date').first()
                    row_data[current_col + 4] = market_cap_data.market_cap if market_cap_data else None

                    # Col 6-9: Fund Data
                    weights = (holding.holding_percentage / 100) if holding.holding_percentage else 0
                    factor = (holding.market_value / total_portfolio_value) if holding.market_value and total_portfolio_value else 0
                    row_data[current_col + 5] = weights
                    row_data[current_col + 6] = factor
                    row_data[current_col + 7] = holding.market_value or 0
                    row_data[current_col + 8] = holding.number_of_shares or 0

                    # Col 10: Share Price
                    price_data = StockPrice.objects.filter(stock=stock).order_by('-price_date').first()
                    row_data[current_col + 9] = price_data.share_price if price_data else None

                    # Col 11: Free Float
                    row_data[current_col + 10] = stock.free_float if stock.free_float else 0

                    # Col 12-18: Fundamentals (get from latest metrics)
                    latest_metrics = FundMetricsLog.objects.filter(
                        scheme=self.scheme,
                        stock=stock
                    ).order_by('-period_date').first()

                    if latest_metrics:
                        row_data[current_col + 11] = latest_metrics.revenue_6yr_cagr
                        row_data[current_col + 12] = None  # TTM Revenue
                        row_data[current_col + 13] = latest_metrics.pat_6yr_cagr
                        row_data[current_col + 14] = None  # TTM PAT
                        row_data[current_col + 15] = latest_metrics.current_pr
                        row_data[current_col + 16] = latest_metrics.pr_2yr_avg
                        row_data[current_col + 17] = latest_metrics.pr_2yr_reval_deval

                    current_col += 18

                elif block['name'] == 'identifiers':
                    # Stock identifiers (3 columns)
                    row_data[current_col] = stock.bse_code or ''
                    row_data[current_col + 1] = stock.nse_symbol or ''
                    row_data[current_col + 2] = stock.isin or ''
                    current_col += 3

            elif block['type'] == 'dynamic':
                # Track section start column for metric rows
                data_type = block['data_type']
                if data_type not in self.section_start_columns:
                    self.section_start_columns[data_type] = current_col

                # Populate dynamic data columns using parent's logic
                periods = block['periods']

                for period in periods:
                    value = self._get_stock_data_value(stock, data_type, period)
                    row_data[current_col] = value if value is not None else ''
                    current_col += 1

        return row_data

    def _get_stock_data_value(self, stock, data_type, period):
        """Get data value for stock at specific period"""
        from .models import StockMarketCap, StockTTMData, StockQuarterlyData, StockAnnualRatios, StockPrice

        try:
            if data_type == 'market_cap':
                obj = StockMarketCap.objects.filter(stock=stock, date=period).first()
                return obj.market_cap if obj else None
            elif data_type == 'market_cap_free_float':
                obj = StockMarketCap.objects.filter(stock=stock, date=period).first()
                return obj.market_cap_free_float if obj else None
            elif data_type == 'ttm_revenue':
                obj = StockTTMData.objects.filter(stock=stock, period=period).first()
                return obj.ttm_revenue if obj else None
            elif data_type == 'ttm_revenue_free_float':
                obj = StockTTMData.objects.filter(stock=stock, period=period).first()
                return obj.ttm_revenue_free_float if obj else None
            elif data_type == 'ttm_pat':
                obj = StockTTMData.objects.filter(stock=stock, period=period).first()
                return obj.ttm_pat if obj else None
            elif data_type == 'ttm_pat_free_float':
                obj = StockTTMData.objects.filter(stock=stock, period=period).first()
                return obj.ttm_pat_free_float if obj else None
            elif data_type == 'quarterly_revenue':
                obj = StockQuarterlyData.objects.filter(stock=stock, period=period).first()
                return obj.quarterly_revenue if obj else None
            elif data_type == 'quarterly_revenue_free_float':
                obj = StockQuarterlyData.objects.filter(stock=stock, period=period).first()
                return obj.quarterly_revenue_free_float if obj else None
            elif data_type == 'quarterly_pat':
                obj = StockQuarterlyData.objects.filter(stock=stock, period=period).first()
                return obj.quarterly_pat if obj else None
            elif data_type == 'quarterly_pat_free_float':
                obj = StockQuarterlyData.objects.filter(stock=stock, period=period).first()
                return obj.quarterly_pat_free_float if obj else None
            elif data_type == 'roce':
                obj = StockAnnualRatios.objects.filter(stock=stock, financial_year=period).first()
                return obj.roce_percentage if obj else None
            elif data_type == 'roe':
                obj = StockAnnualRatios.objects.filter(stock=stock, financial_year=period).first()
                return obj.roe_percentage if obj else None
            elif data_type == 'retention':
                obj = StockAnnualRatios.objects.filter(stock=stock, financial_year=period).first()
                return obj.retention_percentage if obj else None
            elif data_type == 'share_price':
                obj = StockPrice.objects.filter(stock=stock, price_date=period).first()
                return obj.share_price if obj else None
            elif data_type == 'pr_ratio':
                obj = StockPrice.objects.filter(stock=stock, price_date=period).first()
                return obj.pr_ratio if obj else None
            elif data_type == 'pe_ratio':
                obj = StockPrice.objects.filter(stock=stock, price_date=period).first()
                return obj.pe_ratio if obj else None
        except Exception as e:
            logger.error(f"Error getting {data_type} for {stock.company_name} at {period}: {e}")
            return None

        return None


def generate_enhanced_portfolio_analysis_excel(scheme):
    """
    Generate enhanced Portfolio Analysis Excel file with calculated metrics

    Args:
        scheme: AMCFundScheme instance

    Returns:
        BytesIO: Excel file content
    """

    logger.info(f"Generating Fund Portfolio Analysis Excel with block-based structure for {scheme.name}")

    # Get fund holdings
    holdings = FundHolding.objects.filter(scheme=scheme).select_related('stock').order_by('-holding_percentage')

    if not holdings.exists():
        raise ValueError(f"No holdings data found for {scheme.name}")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Analysis"

    # Initialize FundPortfolioExportGenerator
    generator = FundPortfolioExportGenerator(scheme)

    # Collect periods for this fund's stocks
    periods = generator.collect_all_periods_for_fund()

    logger.info(f"Periods collected:")
    logger.info(f"  - Market cap dates: {len(periods['market_cap_dates'])}")
    logger.info(f"  - TTM periods: {len(periods['ttm_periods'])}")
    logger.info(f"  - Quarterly periods: {len(periods['quarterly_periods'])}")
    logger.info(f"  - Annual years: {len(periods['annual_years'])}")
    logger.info(f"  - Share price dates: {len(periods['share_price_dates'])}")
    logger.info(f"  - PR dates: {len(periods['pr_dates'])}")
    logger.info(f"  - PE dates: {len(periods['pe_dates'])}")

    # Define block structure
    blocks = generator._define_block_structure(periods)

    # Calculate total columns
    total_columns = generator._calculate_total_columns(blocks)

    # Generate 8-row headers
    headers = generator._generate_import_style_headers(blocks, total_columns)

    logger.info(f"Generated {total_columns} columns with block-based structure")

    # Add Row 1: Fund name
    row_1 = [''] * total_columns
    row_1[0] = scheme.name
    ws.append(row_1)

    # Add Row 2: Column indicators
    ws.append(headers['row_2'])

    # Add Row 3: Portfolio as on: [date]
    latest_date = max(periods['market_cap_dates']) if periods['market_cap_dates'] else None
    row_3_text = f"Portfolio as on: {latest_date.strftime('%d %B %Y')}" if latest_date else "Portfolio as on: N/A"
    row_3 = [row_3_text] + [''] * (total_columns - 1)
    ws.append(row_3)

    # Add Rows 4-8: Remaining headers (blank rows + category headers + detail headers)
    for row_num in range(3, 9):
        header_row = headers[f'row_{row_num}']
        ws.append(header_row)

    # Add stock data rows (starting from row 9)
    logger.info(f"Populating {len(holdings)} stock rows using block-based structure")

    # Collect stock row data for TOTALS calculation
    stock_rows_data = []
    for idx, holding in enumerate(holdings, start=1):
        row_data = generator.populate_fund_stock_row(holding, blocks, total_columns)
        # NOTE: No S.No column - removed as per user requirement
        stock_rows_data.append(row_data)
        ws.append(row_data)

    # Add 5 blank rows gap before TOTALS (as per sample file)
    for _ in range(5):
        blank_row = [''] * total_columns
        ws.append(blank_row)

    # Calculate TOTALS row by summing each column across all stock rows
    totals_row = ['TOTALS'] + ['' for _ in range(total_columns - 1)]

    logger.info("Calculating TOTALS row by summing all numeric columns")

    # Sum columns 4 onwards (skip Company Name, Accord Code, Sector, Cap which are non-numeric)
    for col_idx in range(4, total_columns):
        column_sum = 0
        has_numeric_data = False

        for stock_row in stock_rows_data:
            value = stock_row[col_idx]
            # Try to add numeric values
            if value is not None and value != '':
                try:
                    column_sum += float(value)
                    has_numeric_data = True
                except (ValueError, TypeError):
                    pass  # Skip non-numeric values (e.g., text)

        # Set total if column had any numeric data
        if has_numeric_data:
            totals_row[col_idx] = column_sum

    ws.append(totals_row)
    totals_row_index = ws.max_row

    # Add 27 portfolio metric rows at bottom (22 with data + 5 blank)
    logger.info("Adding portfolio metric rows")
    add_portfolio_metric_rows(ws, scheme, generator.section_start_columns, periods, total_columns, totals_row_index)

    logger.info("Block-based export completed with portfolio metrics")

    # Apply formatting
    apply_portfolio_analysis_formatting(ws, total_columns)

    logger.info("Enhanced Excel generation completed successfully")

    # Save to BytesIO
    excel_content = BytesIO()
    wb.save(excel_content)
    excel_content.seek(0)

    return excel_content


def generate_recalculated_analysis_excel(scheme, filtered_holdings):
    """
    Generate recalculated Portfolio Analysis Excel file using filtered holdings.
    This is used for the "Recalculated Analysis" sheet in dual-sheet export.

    Args:
        scheme: AMCFundScheme instance
        filtered_holdings: QuerySet of FundHolding objects (pre-filtered by exclusion criteria)

    Returns:
        BytesIO: Excel file content with recalculated metrics
    """

    logger.info(f"Generating Recalculated Portfolio Analysis Excel for {scheme.name} with {filtered_holdings.count()} filtered holdings")

    if not filtered_holdings.exists():
        raise ValueError(f"No holdings data found after filtering for {scheme.name}")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Recalculated Analysis"

    # Initialize FundPortfolioExportGenerator
    generator = FundPortfolioExportGenerator(scheme)

    # Collect periods ONLY from filtered stocks
    periods = generator.collect_periods_for_filtered_holdings(filtered_holdings)

    logger.info(f"Periods collected from filtered stocks:")
    logger.info(f"  - Market cap dates: {len(periods['market_cap_dates'])}")
    logger.info(f"  - TTM periods: {len(periods['ttm_periods'])}")
    logger.info(f"  - Quarterly periods: {len(periods['quarterly_periods'])}")
    logger.info(f"  - Annual years: {len(periods['annual_years'])}")
    logger.info(f"  - Share price dates: {len(periods['share_price_dates'])}")
    logger.info(f"  - PR dates: {len(periods['pr_dates'])}")
    logger.info(f"  - PE dates: {len(periods['pe_dates'])}")

    # Define block structure
    blocks = generator._define_block_structure(periods)

    # Calculate total columns
    total_columns = generator._calculate_total_columns(blocks)

    # Generate 8-row headers
    headers = generator._generate_import_style_headers(blocks, total_columns)

    logger.info(f"Generated {total_columns} columns with block-based structure")

    # Add Row 1: Fund name
    row_1 = [''] * total_columns
    row_1[0] = scheme.name
    ws.append(row_1)

    # Add Row 2: Column indicators
    ws.append(headers['row_2'])

    # Add Row 3: Portfolio as on: [date]
    latest_date = max(periods['market_cap_dates']) if periods['market_cap_dates'] else None
    row_3_text = f"Portfolio as on: {latest_date.strftime('%d %B %Y')}" if latest_date else "Portfolio as on: N/A"
    row_3 = [row_3_text] + [''] * (total_columns - 1)
    ws.append(row_3)

    # Add Rows 4-8: Remaining headers (blank rows + category headers + detail headers)
    for row_num in range(3, 9):
        header_row = headers[f'row_{row_num}']
        ws.append(header_row)

    # Add stock data rows (starting from row 9) - using FILTERED holdings
    logger.info(f"Populating {filtered_holdings.count()} filtered stock rows using block-based structure")

    # Collect stock row data for TOTALS calculation
    stock_rows_data = []
    for idx, holding in enumerate(filtered_holdings, start=1):
        row_data = generator.populate_fund_stock_row(holding, blocks, total_columns)
        stock_rows_data.append(row_data)
        ws.append(row_data)

    # Add 5 blank rows gap before TOTALS (as per sample file)
    for _ in range(5):
        blank_row = [''] * total_columns
        ws.append(blank_row)

    # Calculate TOTALS row by summing each column across all stock rows
    totals_row = ['TOTALS'] + ['' for _ in range(total_columns - 1)]

    logger.info("Calculating TOTALS row by summing all numeric columns (from filtered stocks)")

    # Sum columns 4 onwards (skip Company Name, Accord Code, Sector, Cap which are non-numeric)
    for col_idx in range(4, total_columns):
        column_sum = 0
        has_numeric_data = False

        for stock_row in stock_rows_data:
            value = stock_row[col_idx]
            # Try to add numeric values
            if value is not None and value != '':
                try:
                    column_sum += float(value)
                    has_numeric_data = True
                except (ValueError, TypeError):
                    pass  # Skip non-numeric values (e.g., text)

        # Set total if column had any numeric data
        if has_numeric_data:
            totals_row[col_idx] = column_sum

    ws.append(totals_row)
    totals_row_index = ws.max_row

    # Add 27 portfolio metric rows at bottom (22 with data + 5 blank)
    # NOTE: Portfolio metrics are recalculated from TOTALS row, not from database
    logger.info("Adding portfolio metric rows (recalculated from filtered TOTALS)")
    add_portfolio_metric_rows(ws, scheme, generator.section_start_columns, periods, total_columns, totals_row_index)

    logger.info("Recalculated Excel generation completed")

    # Apply formatting
    apply_portfolio_analysis_formatting(ws, total_columns)

    logger.info("Recalculated Excel formatting applied successfully")

    # Save to BytesIO
    excel_content = BytesIO()
    wb.save(excel_content)
    excel_content.seek(0)

    return excel_content


def apply_portfolio_analysis_formatting(ws, total_columns):
    """
    Apply professional formatting to the Portfolio Analysis worksheet
    """

    # Header formatting
    header_font = Font(name='Calibri', size=11, bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

    # Apply formatting to header rows (1-8)
    for row_num in range(1, 9):
        for col_num in range(1, min(total_columns + 1, 100)):  # Limit formatting for performance
            cell = ws.cell(row=row_num, column=col_num)
            if row_num <= 3:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill

    # Data formatting
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='right', vertical='center')

    # Metric row formatting (bold font, light yellow background)
    metric_font = Font(name='Calibri', size=10, bold=True)
    metric_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    # Calculate metric row range
    # Structure: 8 header rows + N stock rows + 5 blank rows + TOTALS row + 27 metric rows
    # Metric rows: 22 with data + 5 blank = 27 total
    total_rows = ws.max_row
    metric_start_row = total_rows - 26  # Last 27 rows are metrics (22 data + 5 blank)
    totals_row = metric_start_row - 1   # TOTALS row is before metrics

    logger.info(f"Applying formatting - Total rows: {total_rows}, Metric rows: {metric_start_row}-{total_rows}, Totals row: {totals_row}")

    # Apply data formatting to stock rows and identify special rows
    for row_num in range(9, total_rows + 1):
        for col_num in range(1, min(total_columns + 1, 100)):  # Limit columns for performance
            cell = ws.cell(row=row_num, column=col_num)

            # Check if this is a totals row
            if row_num == totals_row:
                cell.font = Font(name='Calibri', size=10, bold=True)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                if col_num > 4:
                    cell.alignment = data_alignment

            # Check if this is a metric row
            elif row_num >= metric_start_row:
                cell.font = metric_font
                if col_num == 1:  # Metric label column
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Metric value columns
                    # No color fill - user requested no colors
                    cell.alignment = data_alignment

            # Regular stock data rows
            else:
                cell.font = data_font
                if col_num > 4:  # Numeric columns
                    cell.alignment = data_alignment

    # Set column widths
    ws.column_dimensions['A'].width = 25  # Company Name / Metric Label
    ws.column_dimensions['B'].width = 15  # Accord Code
    ws.column_dimensions['C'].width = 20  # Sector
    ws.column_dimensions['D'].width = 10  # Cap

    logger.info("Excel formatting applied successfully with metric row highlighting")


def add_portfolio_metric_rows(ws, scheme, section_start_columns, periods, total_columns, totals_row_index):
    """
    Calculate and add 27 portfolio metric rows from TOTALS row data

    Args:
        ws: Worksheet object
        scheme: AMCFundScheme instance
        section_start_columns: Dict mapping data_type to column index
        periods: Dict with period lists (ttm_periods, quarterly_periods, etc.)
        total_columns: Total number of columns in sheet
        totals_row_index: Row index of TOTALS row in worksheet
    """
    logger.info(f"Calculating portfolio metrics from TOTALS row (index: {totals_row_index}) for {scheme.name}")

    # Step 1: Read TOTALS row data from worksheet
    totals_row_data = []
    for col_idx in range(total_columns):
        cell_value = ws.cell(row=totals_row_index, column=col_idx + 1).value
        totals_row_data.append(cell_value if cell_value is not None else 0)

    logger.info(f"Read TOTALS row with {len(totals_row_data)} columns")

    # Step 2: Build section column mapping
    section_cols = build_section_column_mapping(section_start_columns, periods)
    logger.info(f"Built section column mapping for {len(section_cols)} sections")

    # Step 3: Calculate all metrics from TOTALS row
    patm_metrics = calculate_patm_from_totals(totals_row_data, section_cols)
    qoq_metrics = calculate_qoq_from_totals(totals_row_data, section_cols)
    yoy_metrics = calculate_yoy_from_totals(totals_row_data, section_cols)
    cagr_metrics = calculate_6yr_cagr_from_totals(totals_row_data, section_cols)
    pe_pr_metrics = calculate_pe_pr_from_totals(totals_row_data, section_cols)
    pe_pr_avgs = calculate_pe_pr_averages_from_totals(totals_row_data, section_cols)
    reval_deval = calculate_reval_deval_from_totals({**pe_pr_metrics, **pe_pr_avgs})
    pr_extremes = calculate_pr_10q_extremes_from_totals(totals_row_data, section_cols)
    pe_yield = calculate_pe_yield_from_totals(pe_pr_metrics)
    growth = calculate_growth_from_totals(cagr_metrics)
    bond_rate = get_bond_rate()

    logger.info("Calculated all portfolio metrics from TOTALS")

    # Step 4: Define 27 metric rows
    metric_rows = [
        {
            'label': 'PATM',
            'data': patm_metrics,
            'sections': ['ttm_pat', 'ttm_pat_free_float', 'quarterly_pat', 'quarterly_pat_free_float']
        },
        {
            'label': 'QoQ',
            'data': qoq_metrics,
            'sections': ['quarterly_revenue', 'quarterly_revenue_free_float',
                        'quarterly_pat', 'quarterly_pat_free_float']
        },
        {
            'label': 'YoY',
            'data': yoy_metrics,
            'sections': ['ttm_revenue', 'ttm_revenue_free_float', 'quarterly_revenue',
                        'quarterly_revenue_free_float', 'ttm_pat', 'ttm_pat_free_float',
                        'quarterly_pat', 'quarterly_pat_free_float']
        },
        {
            'label': '6 year CAGR',
            'data': cagr_metrics,
            'sections': ['ttm_revenue', 'ttm_revenue_free_float', 'ttm_pat', 'ttm_pat_free_float']
        },
        None,  # Blank
        {
            'label': 'Current PE',
            'data': {'value': pe_pr_metrics.get('current_pe')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': '2 year average',
            'data': {'value': pe_pr_avgs.get('pe_2yr_avg')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': '5 year average',
            'data': {'value': pe_pr_avgs.get('pe_5yr_avg')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': '2 years - Reval / Deval',
            'data': {'value': reval_deval.get('pe_2yr_reval_deval')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': '5 years - Reval / Deval',
            'data': {'value': reval_deval.get('pe_5yr_reval_deval')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        None,  # Blank
        {
            'label': 'Current PR',
            'data': {'value': pe_pr_metrics.get('current_pr')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': '2 year average',
            'data': {'value': pe_pr_avgs.get('pr_2yr_avg')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': '5 year average',
            'data': {'value': pe_pr_avgs.get('pr_5yr_avg')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': '2 years - Reval / Deval',
            'data': {'value': reval_deval.get('pr_2yr_reval_deval')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': '5 years - Reval / Deval',
            'data': {'value': reval_deval.get('pr_5yr_reval_deval')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        None,  # Blank
        {
            'label': '10 quarter- PR- low',
            'data': {'value': pr_extremes.get('pr_10q_low')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': '10 quarter- PR- high',
            'data': {'value': pr_extremes.get('pr_10q_high')},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        None,  # Blank
        {
            'label': 'Alpha over the bond- CAGR',
            'data': {'value': 0.0},  # TODO: Implement if needed
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': 'Alpha- Absolute',
            'data': {'value': 0.0},  # TODO: Implement if needed
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': 'PE Yield',
            'data': {'value': pe_yield},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': 'Growth',
            'data': {'value': growth},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        {
            'label': 'Bond Rate',
            'data': {'value': bond_rate},
            'sections': ['market_cap_free_float'],
            'single_value': True
        },
        None,  # Blank
        None,  # Blank
    ]

    # Step 5: Write rows to worksheet
    for row_def in metric_rows:
        metric_row = create_metric_row(row_def, section_cols, total_columns)
        ws.append(metric_row)

    logger.info(f"Added 27 portfolio metric rows calculated from TOTALS")


# ============================================================================
# SUMMARY SHEET CREATION FUNCTIONS
# ============================================================================

def extract_summary_data_from_worksheet(ws, scheme, stock_rows_start=9):
    """
    Extract key metrics from an analysis worksheet for the Summary sheet.

    Args:
        ws: Worksheet object (Default Analysis or Recalculated Analysis)
        scheme: AMCFundScheme instance
        stock_rows_start: Row number where stock data starts (default: 9)

    Returns:
        dict: Summary data including all metrics needed for Summary sheet
    """
    logger.info(f"Extracting summary data from worksheet: {ws.title}")

    summary_data = {
        'fund_name': scheme.name,
        'num_securities': 0,
        'stocks_80_pct': 0,
        'pct_companies_considered': 0,
        'current_pe': None,
        'pe_2yr_avg': None,
        'pe_5yr_avg': None,
        'pe_reval_2yr': None,
        'pe_reval_5yr': None,
        'current_pr': None,
        'pr_2yr_avg': None,
        'pr_5yr_avg': None,
        'pr_reval_2yr': None,
        'pr_reval_5yr': None,
        'revenue_yoy': None,
        'revenue_6yr_cagr': None,
        'pat_yoy': None,
        'pat_6yr_cagr': None,
        'roe_current': None,
        'roce_current': None,
        'retention_current': None,
        'sector_weights': {},
        'market_cap_breakdown': {},
        'top_10_weight': 0,
    }

    # Find TOTALS row and metric rows
    totals_row_idx = None
    metric_rows_start = None

    for row_idx in range(stock_rows_start, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value == 'TOTALS':
            totals_row_idx = row_idx
            metric_rows_start = row_idx + 1
            break

    if not totals_row_idx:
        logger.warning("Could not find TOTALS row in worksheet")
        return summary_data

    # Count securities (rows between header and blank rows before TOTALS)
    stock_count = 0
    cumulative_weight = 0
    stocks_80_pct = 0
    top_10_weight = 0
    sector_weights = {}
    cap_breakdown = {'Large Cap': 0, 'Mid Cap': 0, 'Small Cap': 0, 'Next 250': 0, 'Micro Cap': 0}

    for row_idx in range(stock_rows_start, totals_row_idx):
        company_name = ws.cell(row=row_idx, column=1).value
        if company_name and company_name.strip() and company_name != '':
            stock_count += 1

            # Get weight (column 6 - Weights)
            weight = ws.cell(row=row_idx, column=6).value
            weight_val = 0  # Initialize to 0 for each row
            if weight:
                try:
                    weight_val = float(weight)

                    # Top 10 weight
                    if stock_count <= 10:
                        top_10_weight += weight_val

                    # 80% of fund calculation
                    cumulative_weight += weight_val
                    if cumulative_weight <= 0.80:
                        stocks_80_pct = stock_count
                    elif stocks_80_pct == 0:
                        stocks_80_pct = stock_count
                except (ValueError, TypeError):
                    weight_val = 0  # Reset to 0 if conversion fails

            # Get sector (column 3)
            sector = ws.cell(row=row_idx, column=3).value
            if sector:
                sector_weights[sector] = sector_weights.get(sector, 0) + weight_val

            # Get cap (column 4)
            cap = ws.cell(row=row_idx, column=4).value
            if cap and cap in cap_breakdown:
                cap_breakdown[cap] = cap_breakdown.get(cap, 0) + weight_val

    summary_data['num_securities'] = stock_count
    summary_data['stocks_80_pct'] = stocks_80_pct
    summary_data['pct_companies_considered'] = 1.0 if stock_count > 0 else 0
    summary_data['top_10_weight'] = top_10_weight
    summary_data['sector_weights'] = sector_weights
    summary_data['market_cap_breakdown'] = cap_breakdown

    # Helper function to get first numeric value from a metric row
    def get_first_metric_value(row_idx):
        """Get the first non-empty numeric value from a metric row"""
        if row_idx is None:
            return None
        for col_idx in range(2, min(ws.max_column + 1, 50)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and val != '':
                try:
                    return float(val)
                except (ValueError, TypeError):
                    continue
        return None

    # Find anchor labels to identify metric sections by position
    # This avoids the duplicate label problem (e.g., "2 year average" appears for both PE and PR)
    pe_section_start = None
    pr_section_start = None
    yoy_row = None
    cagr_row = None
    pr_10q_low_row = None
    pr_10q_high_row = None

    for row_idx in range(metric_rows_start, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if label:
            label_str = label.strip()
            if label_str == 'Current PE' and pe_section_start is None:
                pe_section_start = row_idx
            elif label_str == 'Current PR' and pr_section_start is None:
                pr_section_start = row_idx
            elif label_str == 'YoY' and yoy_row is None:
                yoy_row = row_idx
            elif label_str == '6 year CAGR' and cagr_row is None:
                cagr_row = row_idx
            elif label_str == '10 quarter- PR- low' and pr_10q_low_row is None:
                pr_10q_low_row = row_idx
            elif label_str == '10 quarter- PR- high' and pr_10q_high_row is None:
                pr_10q_high_row = row_idx

    # Extract PE metrics using position relative to "Current PE"
    # Structure: Current PE, 2yr avg, 5yr avg, 2yr Reval, 5yr Reval, [blank]
    if pe_section_start:
        summary_data['current_pe'] = get_first_metric_value(pe_section_start)
        summary_data['pe_2yr_avg'] = get_first_metric_value(pe_section_start + 1)
        summary_data['pe_5yr_avg'] = get_first_metric_value(pe_section_start + 2)
        summary_data['pe_reval_2yr'] = get_first_metric_value(pe_section_start + 3)
        summary_data['pe_reval_5yr'] = get_first_metric_value(pe_section_start + 4)

    # Extract PR metrics using position relative to "Current PR"
    # Structure: Current PR, 2yr avg, 5yr avg, 2yr Reval, 5yr Reval, [blank]
    if pr_section_start:
        summary_data['current_pr'] = get_first_metric_value(pr_section_start)
        summary_data['pr_2yr_avg'] = get_first_metric_value(pr_section_start + 1)
        summary_data['pr_5yr_avg'] = get_first_metric_value(pr_section_start + 2)
        summary_data['pr_reval_2yr'] = get_first_metric_value(pr_section_start + 3)
        summary_data['pr_reval_5yr'] = get_first_metric_value(pr_section_start + 4)

    # Extract growth metrics
    if yoy_row:
        summary_data['revenue_yoy'] = get_first_metric_value(yoy_row)
        summary_data['pat_yoy'] = get_first_metric_value(yoy_row)
    if cagr_row:
        summary_data['revenue_6yr_cagr'] = get_first_metric_value(cagr_row)
        summary_data['pat_6yr_cagr'] = get_first_metric_value(cagr_row)

    # Extract ROE, ROCE, Retention by searching multiple header rows (6, 7, 8)
    # These values are in the data columns, so we search for the column header
    for header_row in [6, 7, 8]:
        for col_idx in range(1, min(ws.max_column + 1, 500)):
            header_val = ws.cell(row=header_row, column=col_idx).value
            if header_val:
                header_str = str(header_val).lower()
                if 'roe' in header_str and summary_data['roe_current'] is None:
                    val = ws.cell(row=totals_row_idx, column=col_idx).value
                    if val:
                        try:
                            summary_data['roe_current'] = float(val)
                        except (ValueError, TypeError):
                            pass
                elif 'roce' in header_str and summary_data['roce_current'] is None:
                    val = ws.cell(row=totals_row_idx, column=col_idx).value
                    if val:
                        try:
                            summary_data['roce_current'] = float(val)
                        except (ValueError, TypeError):
                            pass
                elif 'retention' in header_str and summary_data['retention_current'] is None:
                    val = ws.cell(row=totals_row_idx, column=col_idx).value
                    if val:
                        try:
                            summary_data['retention_current'] = float(val)
                        except (ValueError, TypeError):
                            pass

    logger.info(f"Extracted summary data: {stock_count} securities, {len(sector_weights)} sectors")
    logger.info(f"  PE: current={summary_data['current_pe']}, 2yr={summary_data['pe_2yr_avg']}")
    logger.info(f"  PR: current={summary_data['current_pr']}, 2yr={summary_data['pr_2yr_avg']}")
    return summary_data


def create_summary_sheet(workbook, scheme, summary_data):
    """
    Create a Summary sheet in the workbook with all sections.

    Args:
        workbook: openpyxl Workbook object
        scheme: AMCFundScheme instance
        summary_data: Dict containing extracted metrics

    Returns:
        Worksheet: The created Summary sheet
    """
    logger.info(f"Creating Summary sheet for {scheme.name}")

    # Create sheet at position 0 (first sheet)
    ws = workbook.create_sheet("Summary", 0)

    # Section 1: Title Header (Row 2)
    ws['B2'] = f"{summary_data['fund_name']} - Summary"

    # Section 2: Basic Details (Rows 4-10)
    ws['B4'] = "# Basic Details"
    ws['B6'] = "Particulars"
    ws['C6'] = "Value"
    ws['B7'] = "No. of securities"
    ws['C7'] = summary_data['num_securities']
    ws['B8'] = "80% of the Fund"
    ws['C8'] = summary_data['stocks_80_pct']
    ws['B9'] = "Cash"
    ws['C9'] = ""  # Leave blank as not available
    ws['B10'] = "% of companies considered"
    # pct_companies_considered is already 1.0 (100%), display as percentage
    ws['C10'] = summary_data['pct_companies_considered']

    # Section 3: Pricing & Concentration (Rows 12-18)
    ws['B12'] = "# Pricing & Concentration"

    # Headers (rows 14-15 with merged cells)
    ws['B14'] = "Funds (Peers)"
    ws['C14'] = "Expense Ratio"
    ws['E14'] = "No. of Stocks"
    ws['F14'] = "Top 10 Stocks"
    ws['G14'] = "PE Ratio"
    ws['H14'] = "Revenue Growth (6 yrs)"
    ws['I14'] = "AUM"
    ws['J14'] = "Cash (%)"

    ws['C15'] = "Direct"
    ws['D15'] = "Regular"

    # Current fund data (row 16)
    ws['B16'] = summary_data['fund_name']
    ws['C16'] = ""  # Expense ratio direct - leave blank
    ws['D16'] = ""  # Expense ratio regular - leave blank
    ws['E16'] = summary_data['num_securities']
    # top_10_weight is in decimal form (e.g., 0.45 for 45%)
    ws['F16'] = summary_data['top_10_weight']
    # PE ratio - display as number (not percentage)
    ws['G16'] = summary_data['current_pe']
    # revenue_6yr_cagr is in decimal form (e.g., 0.12 for 12%)
    ws['H16'] = summary_data['revenue_6yr_cagr']
    ws['I16'] = ""  # AUM - leave blank
    ws['J16'] = ""  # Cash - leave blank

    # Section 4: Fundamentals (Rows 20-27)
    ws['B20'] = "# Fundamentals"
    ws['B22'] = "Particulars"
    ws['C22'] = "Current Year"
    ws['D22'] = "6 Year CAGR"
    ws['E22'] = ""

    ws['B23'] = "Revenue Growth"
    # YoY and CAGR are in decimal form from calculation functions
    ws['C23'] = summary_data['revenue_yoy']
    ws['D23'] = summary_data['revenue_6yr_cagr']

    ws['B24'] = "PAT Growth"
    ws['C24'] = summary_data['pat_yoy']
    ws['D24'] = summary_data['pat_6yr_cagr']

    ws['B25'] = "ROE"
    # ROE/ROCE/Retention are read from TOTALS row - check if they need conversion
    roe_val = summary_data['roe_current']
    ws['C25'] = roe_val
    ws['D25'] = roe_val  # Using same as current for avg
    ws['E25'] = "Avg"

    ws['B26'] = "ROCE"
    roce_val = summary_data['roce_current']
    ws['C26'] = roce_val
    ws['D26'] = roce_val  # Using same as current for avg
    ws['E26'] = "Avg"

    ws['B27'] = "Retention Rate"
    retention_val = summary_data['retention_current']
    ws['C27'] = retention_val
    ws['D27'] = retention_val  # Using same as current for avg
    ws['E27'] = "Avg"

    # Section 5: Valuation (Rows 29-37)
    ws['B29'] = "# Valuation"
    ws['B31'] = "Particulars"
    ws['C31'] = "Re/Devaluation"
    ws['D31'] = "Current"
    ws['E31'] = "2 Yr Avg"

    # PE row - PE values are ratios, not percentages
    ws['B32'] = "PE"
    current_pe = summary_data['current_pe']
    pe_2yr = summary_data['pe_2yr_avg']
    if current_pe and pe_2yr and current_pe != 0:
        # Reval/Deval is (avg - current) / current, returns decimal
        ws['C32'] = (pe_2yr - current_pe) / current_pe
    else:
        ws['C32'] = ""
    ws['D32'] = current_pe
    ws['E32'] = pe_2yr

    # PR row - PR values are ratios, not percentages
    ws['B33'] = "PR"
    current_pr = summary_data['current_pr']
    pr_2yr = summary_data['pr_2yr_avg']
    if current_pr and pr_2yr and current_pr != 0:
        ws['C33'] = (pr_2yr - current_pr) / current_pr
    else:
        ws['C33'] = ""
    ws['D33'] = current_pr
    ws['E33'] = pr_2yr

    # Expected Return calculations
    ws['B35'] = "Expected Return (PR & Revenue)"
    rev_cagr = summary_data['revenue_6yr_cagr']
    pr_reval = ws['C33'].value if ws['C33'].value else 0
    if rev_cagr is not None:
        try:
            # rev_cagr is decimal (0.12), pr_reval is decimal (0.05)
            # Result should be decimal for percentage display
            ws['C35'] = rev_cagr + (pr_reval / 5 if pr_reval else 0)
        except:
            ws['C35'] = ""

    ws['B37'] = "Expected Return (PE & PAT)"
    pat_cagr = summary_data['pat_6yr_cagr']
    pe_reval = ws['C32'].value if ws['C32'].value else 0
    if pat_cagr is not None:
        try:
            ws['C37'] = pat_cagr + (pe_reval / 5 if pe_reval else 0)
        except:
            ws['C37'] = ""

    # Section 6: Portfolio Analysis - DYNAMIC ROW POSITIONING
    ws['B39'] = "# Portfolio Analysis"

    # Sector Weights (left side)
    ws['B41'] = "Sector"
    ws['C41'] = "Weight"

    sector_row_idx = 42
    for sector, weight in sorted(summary_data['sector_weights'].items(), key=lambda x: -x[1]):
        ws.cell(row=sector_row_idx, column=2, value=sector)
        ws.cell(row=sector_row_idx, column=3, value=weight)
        sector_row_idx += 1

    # Total sector weights
    ws.cell(row=sector_row_idx, column=2, value="Total")
    ws.cell(row=sector_row_idx, column=3, value=sum(summary_data['sector_weights'].values()))
    sector_end_row = sector_row_idx

    # Market Cap Breakdown (right side) - starts at same row as sectors
    ws['E41'] = "Market Cap"
    ws['F41'] = "Weight"

    cap_order = ['Large Cap', 'Mid Cap', 'Small Cap', 'Next 250', 'Micro Cap']
    cap_row_idx = 42
    for cap_type in cap_order:
        ws.cell(row=cap_row_idx, column=5, value=cap_type)
        ws.cell(row=cap_row_idx, column=6, value=summary_data['market_cap_breakdown'].get(cap_type, 0))
        cap_row_idx += 1

    # Total market cap
    ws.cell(row=cap_row_idx, column=5, value="Total")
    ws.cell(row=cap_row_idx, column=6, value=sum(summary_data['market_cap_breakdown'].values()))
    cap_end_row = cap_row_idx

    # Determine where Portfolio Analysis section ends
    portfolio_analysis_end_row = max(sector_end_row, cap_end_row)

    # Section 7: Fund Performance - DYNAMIC positioning after Portfolio Analysis
    # Add 2 blank rows after Portfolio Analysis section
    fund_perf_start_row = portfolio_analysis_end_row + 3

    ws.cell(row=fund_perf_start_row, column=2, value=f"# Fund Performance (as of {datetime.now().strftime('%d %B %Y')})")

    ws.cell(row=fund_perf_start_row + 2, column=2, value="Funds")
    ws.cell(row=fund_perf_start_row + 2, column=3, value="1M")
    ws.cell(row=fund_perf_start_row + 2, column=4, value="3M")
    ws.cell(row=fund_perf_start_row + 2, column=5, value="6M")
    ws.cell(row=fund_perf_start_row + 2, column=6, value="1Y")
    ws.cell(row=fund_perf_start_row + 2, column=7, value="3Y")
    ws.cell(row=fund_perf_start_row + 2, column=8, value="5Y")

    # Current fund row (performance data if available)
    ws.cell(row=fund_perf_start_row + 3, column=2, value=summary_data['fund_name'])
    # Leave performance columns blank as data not available
    ws.cell(row=fund_perf_start_row + 3, column=3, value="")
    ws.cell(row=fund_perf_start_row + 3, column=4, value="")
    ws.cell(row=fund_perf_start_row + 3, column=5, value="")
    ws.cell(row=fund_perf_start_row + 3, column=6, value="")
    ws.cell(row=fund_perf_start_row + 3, column=7, value="")
    ws.cell(row=fund_perf_start_row + 3, column=8, value="")

    # Store dynamic row info for formatting
    ws._summary_layout = {
        'portfolio_analysis_end_row': portfolio_analysis_end_row,
        'fund_perf_start_row': fund_perf_start_row,
        'sector_end_row': sector_end_row,
        'cap_end_row': cap_end_row
    }

    # Apply formatting
    format_summary_sheet(ws)

    logger.info("Summary sheet created successfully")
    return ws


def format_summary_sheet(ws):
    """
    Apply professional formatting to the Summary sheet.

    Args:
        ws: Worksheet object
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Get dynamic layout info if available
    layout = getattr(ws, '_summary_layout', {})
    portfolio_analysis_end_row = layout.get('portfolio_analysis_end_row', 60)
    fund_perf_start_row = layout.get('fund_perf_start_row', 61)
    sector_end_row = layout.get('sector_end_row', 55)
    cap_end_row = layout.get('cap_end_row', 48)

    # Define styles
    title_font = Font(name='Calibri', size=14, bold=True)
    section_header_font = Font(name='Calibri', size=11, bold=True)
    table_header_font = Font(name='Calibri', size=10, bold=True)
    data_font = Font(name='Calibri', size=10)

    title_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Helper function to apply borders to a table range
    def apply_table_borders(start_row, end_row, start_col, end_col):
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = thin_border

    # Format title (B2)
    ws['B2'].font = title_font
    ws['B2'].fill = title_fill
    ws['B2'].alignment = center_align

    # Format section headers (# Basic Details, etc.) - use dynamic fund_perf_start_row
    section_header_rows = [4, 12, 20, 29, 39, fund_perf_start_row]
    for row in section_header_rows:
        cell = ws.cell(row=row, column=2)
        cell.font = section_header_font

    # Format table headers - use dynamic fund_perf_start_row + 2 for performance header
    table_header_rows = [6, 14, 15, 22, 31, 41, fund_perf_start_row + 2]
    for row in table_header_rows:
        for col in range(2, 11):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.font = table_header_font
                cell.fill = header_fill
                cell.alignment = center_align

    # Set column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12

    # Apply specific number formats based on cell purpose
    # NOTE: Values from analysis sheet are in DECIMAL form (0.15 = 15%)
    # Applying '0.00%' format will correctly display them as percentages

    # Percentage cells - values in DECIMAL form (e.g., 0.15 for 15%)
    # These include: pct_companies_considered, YoY, CAGR, reval/deval, expected returns
    percentage_cells_decimal = ['C10', 'C23', 'C24', 'D23', 'D24', 'C32', 'C33', 'C35', 'C37',
                                'F16', 'H16']
    for cell_ref in percentage_cells_decimal:
        cell = ws[cell_ref]
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '0.00%'

    # ROE/ROCE/Retention - these are typically stored as percentages in DB (e.g., 15.5 for 15.5%)
    # Display them as plain numbers with 2 decimal places, NOT as percentage format
    ratio_style_cells = ['C25', 'D25', 'C26', 'D26', 'C27', 'D27']
    for cell_ref in ratio_style_cells:
        cell = ws[cell_ref]
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '0.00'

    # Ratio cells (PE/PR values that should show as X.XX - these are ratios, not percentages)
    ratio_cells = ['D32', 'E32', 'D33', 'E33', 'G16']
    for cell_ref in ratio_cells:
        cell = ws[cell_ref]
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '0.00'

    # Integer cells (security counts)
    integer_cells = ['C7', 'C8', 'E16']
    for cell_ref in integer_cells:
        cell = ws[cell_ref]
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '0'

    # Format sector weights (column C) and market cap (column F) - DYNAMIC rows
    # Weights are in decimal form (0.055 for 5.5%)
    for row in range(42, sector_end_row + 1):
        cell = ws.cell(row=row, column=3)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '0.00%'

    for row in range(42, cap_end_row + 1):
        cell = ws.cell(row=row, column=6)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.number_format = '0.00%'

    # Apply borders to data tables - use dynamic rows
    apply_table_borders(6, 10, 2, 3)     # Basic Details
    apply_table_borders(14, 16, 2, 10)   # Pricing & Concentration
    apply_table_borders(22, 27, 2, 5)    # Fundamentals
    apply_table_borders(31, 37, 2, 5)    # Valuation
    apply_table_borders(41, sector_end_row, 2, 3)    # Sector Weights (dynamic)
    apply_table_borders(41, cap_end_row, 5, 6)       # Market Cap Breakdown (dynamic)
    apply_table_borders(fund_perf_start_row + 2, fund_perf_start_row + 4, 2, 8)    # Fund Performance (dynamic)

    # Merge cells for title
    ws.merge_cells('B2:E2')

    # Merge cells for Pricing & Concentration headers
    ws.merge_cells('B14:B15')
    ws.merge_cells('C14:D14')
    ws.merge_cells('E14:E15')
    ws.merge_cells('F14:F15')
    ws.merge_cells('G14:G15')
    ws.merge_cells('H14:H15')
    ws.merge_cells('I14:I15')
    ws.merge_cells('J14:J15')

    logger.info("Summary sheet formatting applied")