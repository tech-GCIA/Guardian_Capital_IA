# views.py
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.utils.timezone import now
from django.utils import timezone
from datetime import timedelta
from gcia_app.forms import CustomerCreationForm
from django.contrib import messages
import openpyxl
from django.http import HttpResponse
from gcia_app.forms import ExcelUploadForm, MasterDataExcelUploadForm
from gcia_app.models import AMCFundScheme, AMCFundSchemeNavLog, Stock, StockMarketCap, StockTTMData, StockQuarterlyData, StockAnnualRatios, StockPrice, FundHolding, FundMetricsLog, FileStructureMetadata
import pandas as pd
import os
import datetime
from django.db import transaction
from openpyxl import load_workbook
from io import BytesIO
import traceback
from gcia_app.utils import update_avg_category_returns, calculate_scheme_age
from gcia_app.portfolio_analysis_ppt import create_fund_presentation
import re
from difflib import SequenceMatcher
from django.db import transaction
from gcia_app.index_scrapper_from_screener import get_bse500_pe_ratio
from gcia_app.dynamic_stock_analyzer import DynamicStockSheetAnalyzer

def clean_all_stock_data():
    """
    Clean all existing stock-related data from the database.
    This includes all Stock records and their related time-series data,
    as well as Fund Holdings and Metrics that reference stocks.

    Returns:
        dict: Statistics about the cleanup operation
    """
    stats = {
        'stocks_deleted': 0,
        'market_cap_records_deleted': 0,
        'ttm_records_deleted': 0,
        'quarterly_records_deleted': 0,
        'annual_ratios_deleted': 0,
        'price_records_deleted': 0,
        'fund_holdings_deleted': 0,
        'fund_metrics_deleted': 0,
    }

    try:
        with transaction.atomic():
            # Delete Fund Metrics Log records first (they reference stocks)
            fund_metrics_count = FundMetricsLog.objects.count()
            FundMetricsLog.objects.all().delete()
            stats['fund_metrics_deleted'] = fund_metrics_count

            # Delete Fund Holdings (they reference stocks)
            fund_holdings_count = FundHolding.objects.count()
            FundHolding.objects.all().delete()
            stats['fund_holdings_deleted'] = fund_holdings_count

            # Delete all stock time-series data
            market_cap_count = StockMarketCap.objects.count()
            StockMarketCap.objects.all().delete()
            stats['market_cap_records_deleted'] = market_cap_count

            ttm_count = StockTTMData.objects.count()
            StockTTMData.objects.all().delete()
            stats['ttm_records_deleted'] = ttm_count

            quarterly_count = StockQuarterlyData.objects.count()
            StockQuarterlyData.objects.all().delete()
            stats['quarterly_records_deleted'] = quarterly_count

            annual_ratios_count = StockAnnualRatios.objects.count()
            StockAnnualRatios.objects.all().delete()
            stats['annual_ratios_deleted'] = annual_ratios_count

            price_count = StockPrice.objects.count()
            StockPrice.objects.all().delete()
            stats['price_records_deleted'] = price_count

            # Finally delete all Stock records
            stocks_count = Stock.objects.count()
            Stock.objects.all().delete()
            stats['stocks_deleted'] = stocks_count

            print(f"Successfully cleaned all stock data:")
            print(f"   - Stocks: {stats['stocks_deleted']}")
            print(f"   - Market Cap records: {stats['market_cap_records_deleted']}")
            print(f"   - TTM records: {stats['ttm_records_deleted']}")
            print(f"   - Quarterly records: {stats['quarterly_records_deleted']}")
            print(f"   - Annual ratios: {stats['annual_ratios_deleted']}")
            print(f"   - Price records: {stats['price_records_deleted']}")
            print(f"   - Fund holdings: {stats['fund_holdings_deleted']}")
            print(f"   - Fund metrics: {stats['fund_metrics_deleted']}")

    except Exception as e:
        print(f"Error during stock data cleanup: {str(e)}")
        print(traceback.format_exc())
        raise

    return stats

def signup_view(request):
    if request.method == 'POST':
        form = CustomerCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = CustomerCreationForm()
    return render(request, 'gcia_app/signup.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            request.session['last_activity'] = now().isoformat()
            return redirect('home')
    return render(request, 'gcia_app/login.html')

@login_required
def home_view(request):
    last_activity = request.session.get('last_activity')
    if last_activity and now() - timedelta(minutes=30) > now().fromisoformat(last_activity):
        logout(request)
        return redirect('login')
    request.session['last_activity'] = now().isoformat()
    return render(request, 'gcia_app/home.html')

def logout_view(request):
    logout(request)
    return redirect('login')

def calculate_returns_for_index(scheme, index_data):
    """
    Calculate returns for various periods and update the scheme
    
    Args:
        scheme: The AMCFundScheme object to update
        index_data: DataFrame containing the index data with Date and Close Price columns
    """
    # Ensure data is sorted by date
    index_data = index_data.sort_values("Date")
    
    # Get latest date and NAV
    latest_date = index_data.iloc[-1]["Date"]
    latest_nav = float(index_data.iloc[-1]["Close Price"])
    
    # Function to calculate returns for a specific period
    def get_return_for_period(days):
        try:
            # Get data from the past
            past_date = latest_date - datetime.timedelta(days=days)
            past_data = index_data[index_data["Date"] <= past_date]
            
            if len(past_data) == 0:
                return None  # No data available for this period
            
            # Get the closest past NAV
            past_nav = float(past_data.iloc[-1]["Close Price"])
            
            # Calculate return as percentage
            if past_nav > 0:
                return ((latest_nav - past_nav) / past_nav) * 100
            return None
        except Exception:
            return None
    
    # Calculate returns for various periods
    scheme.returns_1_day = get_return_for_period(1)
    scheme.returns_7_day = get_return_for_period(7)
    scheme.returns_15_day = get_return_for_period(15)
    scheme.returns_1_mth = get_return_for_period(30)
    scheme.returns_3_mth = get_return_for_period(90)
    scheme.returns_6_mth = get_return_for_period(180)
    scheme.returns_1_yr = get_return_for_period(365)
    scheme.returns_2_yr = get_return_for_period(730)
    scheme.returns_3_yr = get_return_for_period(1095)
    scheme.returns_5_yr = get_return_for_period(1825)
    scheme.returns_7_yr = get_return_for_period(2555)
    scheme.returns_10_yr = get_return_for_period(3650)
    scheme.returns_15_yr = get_return_for_period(5475)
    scheme.returns_20_yr = get_return_for_period(7300)
    scheme.returns_25_yr = get_return_for_period(9125)
    
    # Calculate returns since inception (first available data point)
    try:
        first_data = index_data.iloc[0]
        first_nav = float(first_data["Close Price"])
        first_date = first_data["Date"]
        
        # Calculate annualized returns from inception
        if first_nav > 0:
            # Calculate days since inception
            days_since_inception = (latest_date - first_date).days
            if days_since_inception > 0:
                # Calculate simple return
                total_return = ((latest_nav - first_nav) / first_nav) * 100
                # Annualize the return
                years = days_since_inception / 365.0
                if years > 0:
                    annualized_return = ((1 + (total_return / 100)) ** (1 / years) - 1) * 100
                    scheme.returns_from_launch = annualized_return
                else:
                    scheme.returns_from_launch = total_return  # If less than a year, use simple return
    except Exception:
        scheme.returns_from_launch = None

def process_index_nav_file(excel_file):
    """
    Process INDEX NAV file to update daily prices and calculate returns
    
    Args:
        excel_file: The uploaded Excel file containing index NAV data
        
    Returns:
        dict: Statistics about the processing
    """
    # Read the Excel file, skipping the first 2 rows as header starts at row 3
    try:
        index_nav_df = pd.read_excel(excel_file, skiprows=2)
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}")
    
    # Verify the expected columns exist
    expected_columns = ["Index Name", "Date", "Close Price"]
    for col in expected_columns:
        if col not in index_nav_df.columns:
            raise ValueError(f"Expected column '{col}' not found in the INDEX NAV file")
    
    # Drop any rows where Index Name or Date or Close Price is NaN
    index_nav_df = index_nav_df.dropna(subset=["Index Name", "Date", "Close Price"])
    
    # Ensure Date is in datetime format
    index_nav_df["Date"] = pd.to_datetime(index_nav_df["Date"]).dt.date
    
    # Get unique index names
    unique_indices = index_nav_df["Index Name"].unique()
    
    # Statistics for tracking the update process
    stats = {
        'total_indices': len(unique_indices),
        'indices_updated': 0,
        'indices_created': 0,
        'nav_entries_added': 0,
        'return_calculations_updated': 0
    }
    
    # Process each index
    with transaction.atomic():  # Use transaction to ensure data integrity
        for index_name in unique_indices:
            # Get rows for this index
            index_data = index_nav_df[index_nav_df["Index Name"] == index_name].sort_values("Date")
            
            if len(index_data) == 0:
                continue  # Skip if no data
            
            # Try to find this index in AMCFundScheme
            scheme = AMCFundScheme.objects.filter(name=index_name).first()
            
            if not scheme:
                # If not found, create a new entry
                scheme = AMCFundScheme.objects.create(
                    name=index_name,
                    is_active=True,
                    fund_name=index_name,
                    is_scheme_benchmark=True,  # This is an index, so mark it as a benchmark
                    latest_nav_as_on_date=index_data.iloc[-1]["Date"]  # Set the latest date
                )
                stats['indices_created'] += 1
            else:
                stats['indices_updated'] += 1
            
            # Process each date's NAV for this index
            for _, row in index_data.iterrows():
                date = row["Date"]
                close_price = float(row["Close Price"])
                
                # Update or create a nav log entry
                nav_log, created = AMCFundSchemeNavLog.objects.update_or_create(
                    amcfundscheme=scheme,
                    as_on_date=date,
                    nav=close_price
                )
                
                if created:
                    stats['nav_entries_added'] += 1
            
            # Update the scheme with latest NAV
            latest_data = index_data.iloc[-1]
            latest_date = latest_data["Date"]
            latest_nav = float(latest_data["Close Price"])
            
            scheme.latest_nav = latest_nav
            scheme.latest_nav_as_on_date = latest_date
            
            # Calculate returns for various periods
            calculate_returns_for_index(scheme, index_data)
            stats['return_calculations_updated'] += 1
            
            # Save the scheme with updated returns
            scheme.save()
    
    return stats

def transform_scheme_name(original_name):
    """
    Transform TopSchemes format to RATIOS PE format
    Optimized for the specific patterns in the provided files
    """
    transformed = original_name
    
    # Fix case sensitivity for "One" vs "ONE"
    transformed = transformed.replace(" One ", " ONE ")
    
    # Handle Direct Plan variations
    if " (G) Direct" in transformed:
        transformed = transformed.replace(" (G) Direct", "(G)-Direct Plan")
    elif " Direct" in transformed and not transformed.endswith(" Plan"):
        transformed = transformed.replace(" Direct", "-Direct Plan")
    
    # Handle Regular Plan variations
    if " Reg (G)" in transformed:
        transformed = transformed.replace(" Reg (G)", "-Reg(G)")
    elif " Reg " in transformed:
        transformed = transformed.replace(" Reg ", "-Reg")
    
    # Remove spaces around parentheses
    transformed = transformed.replace(" (", "(").replace(") ", ")")
    
    # Handle "FlexiCap" vs "Flexicap" difference
    transformed = transformed.replace("Flexi Cap", "Flexicap")
    
    return transformed

def find_closest_match(db_scheme_name, ratios_pe_schemes, threshold=0.85):
    """
    Find closest matching scheme name in RATIOS PE file for a given database scheme name
    Optimized for faster matching with early returns for exact matches
    """
    # Try direct transformation first
    transformed_name = transform_scheme_name(db_scheme_name)
    
    # Check for exact match
    exact_matches = [name for name in ratios_pe_schemes if name == transformed_name]
    if exact_matches:
        return exact_matches[0], 1.0  # Return with confidence score of 1.0
    
    # Extract fund family and scheme type for matching
    db_parts = db_scheme_name.lower().split()
    
    # Create a key from first few words (usually the fund family and type)
    key_parts = ' '.join(db_parts[:min(4, len(db_parts))])
    
    # Check if scheme has "Direct" or "Reg" in the name
    is_direct = "direct" in db_scheme_name.lower()
    is_reg = "reg" in db_scheme_name.lower()
    is_growth = "(g)" in db_scheme_name.lower() or "growth" in db_scheme_name.lower()
    
    # Find potential matches
    potential_matches = []
    for ratios_name in ratios_pe_schemes:
        # Skip header or empty rows
        if not isinstance(ratios_name, str) or ratios_name == 'Scheme Name':
            continue
            
        # Match fund family and type
        ratios_lower = ratios_name.lower()
        
        # Check if the key parts are in the ratios scheme name
        if key_parts in ratios_lower:
            # Check for Direct/Regular match
            ratios_is_direct = "direct" in ratios_lower
            ratios_is_reg = "reg" in ratios_lower
            ratios_is_growth = "(g)" in ratios_lower or "growth" in ratios_lower
            
            # Calculate base similarity ratio
            similarity = SequenceMatcher(None, transformed_name.lower(), ratios_lower).ratio()
            
            # Boost similarity if Direct/Regular and Growth/Income type matches
            if (is_direct and ratios_is_direct) or (is_reg and ratios_is_reg):
                similarity += 0.1
            if is_growth and ratios_is_growth:
                similarity += 0.1
                
            potential_matches.append((ratios_name, similarity))
    
    # Sort by similarity score
    potential_matches.sort(key=lambda x: x[1], reverse=True)
    
    # Return the highest scoring match if above threshold
    if potential_matches and potential_matches[0][1] >= threshold:
        return potential_matches[0]
    
    # If no good match found
    return None, 0.0

def process_ratios_pe_file(excel_file):
    """
    Process RATIOS PE file and update the database with TS_Total Count, RR3_Up Capture Ratio, and RR3_Down Capture Ratio
    This version assumes consistent file structure with headers at row 4 and specific column names
    """
    # Read the Excel file, skipping the first 3 rows as header starts at row 4
    ratios_pe_df = pd.read_excel(excel_file, skiprows=3)
    
    # Now the columns should have proper names:
    # - "Scheme Code"
    # - "Scheme Name"
    # - "TS_Total Count"
    # - "RR3_Up Capture Ratio"
    # - "RR3_Down Capture Ratio"
    
    # Verify the expected columns exist
    expected_columns = ["Scheme Code", "Scheme Name", "TS_Total Count", 
                        "RR3_Up Capture Ratio", "RR3_Down Capture Ratio"]
    
    for col in expected_columns:
        if col not in ratios_pe_df.columns:
            raise ValueError(f"Expected column '{col}' not found in the RATIOS PE file")
    
    # Drop any rows where Scheme Name is NaN
    ratios_pe_df = ratios_pe_df.dropna(subset=["Scheme Name"])
    
    # Get scheme names from database (AMCFundScheme table)
    db_schemes = AMCFundScheme.objects.all().values_list('name', 'amcfundscheme_id')
    
    # Get unique scheme names from the RATIOS PE file
    ratios_pe_schemes = ratios_pe_df["Scheme Name"].unique()
    
    # Statistics for tracking the update process
    stats = {
        'total_schemes': len(db_schemes),
        'updated': 0,
        'not_found': 0,
        'low_confidence': 0
    }
    
    # Process each scheme in the database
    with transaction.atomic():  # Use transaction to ensure data integrity
        for db_scheme_name, scheme_id in db_schemes:
            # Find matching scheme in RATIOS PE
            matched_scheme, confidence = find_closest_match(db_scheme_name, ratios_pe_schemes)
            
            if matched_scheme is not None and confidence >= 0.85:  # Only update with high confidence matches
                # Get the data for this scheme from RATIOS PE
                ratios_data = ratios_pe_df[ratios_pe_df["Scheme Name"] == matched_scheme]
                
                if not ratios_data.empty:
                    # Get the first matching row
                    data_row = ratios_data.iloc[0]
                    
                    # Update the AMCFundScheme object
                    scheme_obj = AMCFundScheme.objects.get(amcfundscheme_id=scheme_id)
                    print(scheme_obj.name, data_row)

                    # Update the fields
                    if pd.notna(data_row["TS_Total Count"]):
                        scheme_obj.number_of_underlying_stocks = float(data_row["TS_Total Count"])
                    
                    if pd.notna(data_row["RR3_Up Capture Ratio"]):
                        scheme_obj.up_capture = float(data_row["RR3_Up Capture Ratio"])
                    
                    if pd.notna(data_row["RR3_Down Capture Ratio"]):
                        scheme_obj.down_capture = float(data_row["RR3_Down Capture Ratio"])
                    
                    # Save the updated object
                    scheme_obj.save()
                    stats['updated'] += 1
            elif matched_scheme is not None:
                # Match found but confidence too low
                stats['low_confidence'] += 1
            else:
                # No match found
                stats['not_found'] += 1
    
    return stats

def process_top_schemes_data_uploaded(excel_file):
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel(excel_file)
    current_date = datetime.date.today()
    amcfs_list = AMCFundScheme.objects.filter()
    for index, row in df.iterrows():
        amcfs = amcfs_list.filter(name=row["SCHEMES"]).first()
        if not amcfs:
            amcfs = AMCFundScheme.objects.create(name=row["SCHEMES"],
                                                is_active=True,
                                                fund_name = row["FUND NAME"] if "FUND NAME" in row and not pd.isna(row["FUND NAME"]) else None,
                                                is_direct_fund = True if "SCHEMES" in row and not pd.isna(row["SCHEMES"]) and "direct" in row["SCHEMES"].lower() else False,
                                                assets_under_management = row[" A U M"] if " A U M" in row and not pd.isna(row[" A U M"]) else None,
                                                launch_date = datetime.datetime.strptime(str(row["LAUNCH DATE"]), "%d-%m-%Y") if "LAUNCH DATE" in row and not pd.isna(row["LAUNCH DATE"]) else None,
                                                returns_1_day = row["1 DAY"] if "1 DAY" in row and not pd.isna(row["1 DAY"]) else None,
                                                returns_7_day = row["7 DAY"] if "7 DAY" in row and not pd.isna(row["7 DAY"]) else None,
                                                returns_15_day = row["15 DAY"] if "15 DAY" in row and not pd.isna(row["15 DAY"]) else None,
                                                returns_1_mth = row["30 DAY"] if "30 DAY" in row and not pd.isna(row["30 DAY"]) else None,
                                                returns_3_mth = row["3 MONTH"] if "3 MONTH" in row and not pd.isna(row["3 MONTH"]) else None,
                                                returns_6_mth = row["6 MONTH"] if "6 MONTH" in row and not pd.isna(row["6 MONTH"]) else None,
                                                returns_1_yr = row["1 YEAR"] if "1 YEAR" in row and not pd.isna(row["1 YEAR"]) else None,
                                                returns_2_yr = row["2 YEAR"] if "2 YEAR" in row and not pd.isna(row["2 YEAR"]) else None,
                                                returns_3_yr = row["3 YEAR"] if "3 YEAR" in row and not pd.isna(row["3 YEAR"]) else None,
                                                returns_5_yr = row["5 YEAR"] if "5 YEAR" in row and not pd.isna(row["5 YEAR"]) else None,
                                                returns_7_yr = row["7 YEAR"] if "7 YEAR" in row and not pd.isna(row["7 YEAR"]) else None,
                                                returns_10_yr = row["10 YEAR"] if "10 YEAR" in row and not pd.isna(row["10 YEAR"]) else None,
                                                returns_15_yr = row["15 YEAR"] if "15 YEAR" in row and not pd.isna(row["15 YEAR"]) else None,
                                                returns_20_yr = row["20 YEAR"] if "20 YEAR" in row and not pd.isna(row["20 YEAR"]) else None,
                                                returns_25_yr = row["25 YEAR"] if "25 YEAR" in row and not pd.isna(row["25 YEAR"]) else None,
                                                returns_from_launch = row["SINCE INCEPTION RETURN"] if "SINCE INCEPTION RETURN" in row and not pd.isna(row["SINCE INCEPTION RETURN"]) else None,
                                                fund_rating = row["FUND RAT"] if "FUND RAT" in row and not pd.isna(row["FUND RAT"]) else None,
                                                fund_class = row["CATEGORY"] if "CATEGORY" in row and not pd.isna(row["CATEGORY"]) else None,
                                                latest_nav = row["CURRENT NAV"] if "CURRENT NAV" in row and not pd.isna(row["CURRENT NAV"]) else None,
                                                latest_nav_as_on_date = datetime.date.today(),
                                                alpha = row["ALPHA"] if "ALPHA" in row and not pd.isna(row["ALPHA"]) else None,
                                                beta = row["BETA"] if "BETA" in row and not pd.isna(row["BETA"]) else None,
                                                mean = row["MEAN"] if "MEAN" in row and not pd.isna(row["MEAN"]) else None,
                                                standard_dev = row["STANDARD DEV"] if "STANDARD DEV" in row and not pd.isna(row["STANDARD DEV"]) else None,
                                                sharpe_ratio = row["SHARPE RATIO"] if "SHARPE RATIO" in row and not pd.isna(row["SHARPE RATIO"]) else None,
                                                sorti_i_no = row["SORTI I NO"] if "SORTI I NO" in row and not pd.isna(row["SORTI I NO"]) else None,
                                                fund_manager = row["FUND MANAGER"] if "FUND MANAGER" in row and not pd.isna(row["FUND MANAGER"]) else None,
                                                avg_mat = row["AVG MAT"] if "AVG MAT" in row and not pd.isna(row["AVG MAT"]) else None,
                                                modified_duration = row["MODIFIED DURATION"] if "MODIFIED DURATION" in row and not pd.isna(row["MODIFIED DURATION"]) else None,
                                                ytm = row["YTM"] if "YTM" in row and not pd.isna(row["YTM"]) else None,
                                                purchase_minimum_amount = row["PURCHASE MIN AMOUNT"] if "PURCHASE MIN AMOUNT" in row and not pd.isna(row["PURCHASE MIN AMOUNT"]) else None,
                                                sip_minimum_amount = row["SIP MIN AMOUNT"] if "SIP MIN AMOUNT" in row and not pd.isna(row["SIP MIN AMOUNT"]) else None,
                                                large_cap = row["LARGE CAP"] if "LARGE CAP" in row and not pd.isna(row["LARGE CAP"]) else None,
                                                mid_cap = row["MID CAP"] if "MID CAP" in row and not pd.isna(row["MID CAP"]) else None,
                                                small_cap = row["SMALL CAP"] if "SMALL CAP" in row and not pd.isna(row["SMALL CAP"]) else None,
                                                pb_ratio = row["PB RATIO"] if "PB RATIO" in row and not pd.isna(row["PB RATIO"]) else None,
                                                pe_ratio = row["PE RATIO"] if "PE RATIO" in row and not pd.isna(row["PE RATIO"]) else None,
                                                exit_load = row["EXIT LOAD"] if "EXIT LOAD" in row and not pd.isna(row["EXIT LOAD"]) else None,
                                                equity_percentage = row[" EQUITY(%)"] if " EQUITY(%)" in row and not pd.isna(row[" EQUITY(%)"]) else None,
                                                debt_percentage = row[" DEBT(%)"] if " DEBT(%)" in row and not pd.isna(row[" DEBT(%)"]) else None,
                                                gold_percentage = row[" GOLD(%)"] if " GOLD(%)" in row and not pd.isna(row[" GOLD(%)"]) else None,
                                                global_equity_percentage = row[" GLOBAL EQUITY(%)"] if " GLOBAL EQUITY(%)" in row and not pd.isna(row[" GLOBAL EQUITY(%)"]) else None,
                                                other_percentage = row[" OTHER(%)"] if " OTHER(%)" in row and not pd.isna(row[" OTHER(%)"]) else None,
                                                rs_quard = row["RS QUARD"] if "RS QUARD" in row and not pd.isna(row["RS QUARD"]) else None,
                                                expense_ratio = row["EXPENSE RATIO"] if "EXPENSE RATIO" in row and not pd.isna(row["EXPENSE RATIO"]) else None,
                                                SOV = row["SOV"] if "SOV" in row and not pd.isna(row["SOV"]) else None,
                                                A = row["A"] if "A" in row and not pd.isna(row["A"]) else None,
                                                AA = row["AA"] if "AA" in row and not pd.isna(row["AA"]) else None,
                                                AAA = row["AAA"] if "AAA" in row and not pd.isna(row["AAA"]) else None,
                                                BIG = row["BIG"] if "BIG" in row and not pd.isna(row["BIG"]) else None,
                                                cash = row["CASH"] if "CASH" in row and not pd.isna(row["CASH"]) else None,
                                                downside_deviation = row["DOWNSIDE DEVIATION"] if "DOWNSIDE DEVIATION" in row and not pd.isna(row["DOWNSIDE DEVIATION"]) else None,
                                                downside_probability = row["DOWNSIDE PROBABILITY"] if "DOWNSIDE PROBABILITY" in row and not pd.isna(row["DOWNSIDE PROBABILITY"]) else None,
                                                scheme_benchmark = row["SCHEME BENCHMARK"] if "SCHEME BENCHMARK" in row and not pd.isna(row["SCHEME BENCHMARK"]) else None,
                                                is_scheme_benchmark = False if " A U M" in row and not pd.isna(row[" A U M"]) else True)
        
        if amcfs and amcfs.latest_nav_as_on_date==current_date:
            continue
        else:
            amcfs.is_active=True
            amcfs.fund_name = row["FUND NAME"] if "FUND NAME" in row and not pd.isna(row["FUND NAME"]) else None
            amcfs.is_direct_fund = True if "SCHEMES" in row and not pd.isna(row["SCHEMES"]) and "direct" in row["SCHEMES"].lower() else False
            amcfs.assets_under_management = row[" A U M"] if " A U M" in row and not pd.isna(row[" A U M"]) else None
            amcfs.launch_date = datetime.datetime.strptime(str(row["LAUNCH DATE"]), "%d-%m-%Y") if "LAUNCH DATE" in row and not pd.isna(row["LAUNCH DATE"]) else None
            amcfs.returns_1_day = row["1 DAY"] if "1 DAY" in row and not pd.isna(row["1 DAY"]) else None
            amcfs.returns_7_day = row["7 DAY"] if "7 DAY" in row and not pd.isna(row["7 DAY"]) else None
            amcfs.returns_15_day = row["15 DAY"] if "15 DAY" in row and not pd.isna(row["15 DAY"]) else None
            amcfs.returns_1_mth = row["30 DAY"] if "30 DAY" in row and not pd.isna(row["30 DAY"]) else None
            amcfs.returns_3_mth = row["3 MONTH"] if "3 MONTH" in row and not pd.isna(row["3 MONTH"]) else None
            amcfs.returns_6_mth = row["6 MONTH"] if "6 MONTH" in row and not pd.isna(row["6 MONTH"]) else None
            amcfs.returns_1_yr = row["1 YEAR"] if "1 YEAR" in row and not pd.isna(row["1 YEAR"]) else None
            amcfs.returns_2_yr = row["2 YEAR"] if "2 YEAR" in row and not pd.isna(row["2 YEAR"]) else None
            amcfs.returns_3_yr = row["3 YEAR"] if "3 YEAR" in row and not pd.isna(row["3 YEAR"]) else None
            amcfs.returns_5_yr = row["5 YEAR"] if "5 YEAR" in row and not pd.isna(row["5 YEAR"]) else None
            amcfs.returns_7_yr = row["7 YEAR"] if "7 YEAR" in row and not pd.isna(row["7 YEAR"]) else None
            amcfs.returns_10_yr = row["10 YEAR"] if "10 YEAR" in row and not pd.isna(row["10 YEAR"]) else None
            amcfs.returns_15_yr = row["15 YEAR"] if "15 YEAR" in row and not pd.isna(row["15 YEAR"]) else None
            amcfs.returns_20_yr = row["20 YEAR"] if "20 YEAR" in row and not pd.isna(row["20 YEAR"]) else None
            amcfs.returns_25_yr = row["25 YEAR"] if "25 YEAR" in row and not pd.isna(row["25 YEAR"]) else None
            amcfs.returns_from_launch = row["SINCE INCEPTION RETURN"] if "SINCE INCEPTION RETURN" in row and not pd.isna(row["SINCE INCEPTION RETURN"]) else None
            amcfs.fund_rating = row["FUND RAT"] if "FUND RAT" in row and not pd.isna(row["FUND RAT"]) else None
            amcfs.fund_class = row["CATEGORY"] if "CATEGORY" in row and not pd.isna(row["CATEGORY"]) else None
            amcfs.latest_nav = row["CURRENT NAV"] if "CURRENT NAV" in row and not pd.isna(row["CURRENT NAV"]) else None
            amcfs.latest_nav_as_on_date = datetime.date.today()
            amcfs.alpha = row["ALPHA"] if "ALPHA" in row and not pd.isna(row["ALPHA"]) else None
            amcfs.beta = row["BETA"] if "BETA" in row and not pd.isna(row["BETA"]) else None
            amcfs.mean = row["MEAN"] if "MEAN" in row and not pd.isna(row["MEAN"]) else None
            amcfs.standard_dev = row["STANDARD DEV"] if "STANDARD DEV" in row and not pd.isna(row["STANDARD DEV"]) else None
            amcfs.sharpe_ratio = row["SHARPE RATIO"] if "SHARPE RATIO" in row and not pd.isna(row["SHARPE RATIO"]) else None
            amcfs.sorti_i_no = row["SORTI I NO"] if "SORTI I NO" in row and not pd.isna(row["SORTI I NO"]) else None
            amcfs.fund_manager = row["FUND MANAGER"] if "FUND MANAGER" in row and not pd.isna(row["FUND MANAGER"]) else None
            amcfs.avg_mat = row["AVG MAT"] if "AVG MAT" in row and not pd.isna(row["AVG MAT"]) else None
            amcfs.modified_duration = row["MODIFIED DURATION"] if "MODIFIED DURATION" in row and not pd.isna(row["MODIFIED DURATION"]) else None
            amcfs.ytm = row["YTM"] if "YTM" in row and not pd.isna(row["YTM"]) else None
            amcfs.purchase_minimum_amount = row["PURCHASE MIN AMOUNT"] if "PURCHASE MIN AMOUNT" in row and not pd.isna(row["PURCHASE MIN AMOUNT"]) else None
            amcfs.sip_minimum_amount = row["SIP MIN AMOUNT"] if "SIP MIN AMOUNT" in row and not pd.isna(row["SIP MIN AMOUNT"]) else None
            amcfs.large_cap = row["LARGE CAP"] if "LARGE CAP" in row and not pd.isna(row["LARGE CAP"]) else None
            amcfs.mid_cap = row["MID CAP"] if "MID CAP" in row and not pd.isna(row["MID CAP"]) else None
            amcfs.small_cap = row["SMALL CAP"] if "SMALL CAP" in row and not pd.isna(row["SMALL CAP"]) else None
            amcfs.pb_ratio = row["PB RATIO"] if "PB RATIO" in row and not pd.isna(row["PB RATIO"]) else None
            amcfs.pe_ratio = row["PE RATIO"] if "PE RATIO" in row and not pd.isna(row["PE RATIO"]) else None
            amcfs.exit_load = row["EXIT LOAD"] if "EXIT LOAD" in row and not pd.isna(row["EXIT LOAD"]) else None
            amcfs.equity_percentage = row[" EQUITY(%)"] if " EQUITY(%)" in row and not pd.isna(row[" EQUITY(%)"]) else None
            amcfs.debt_percentage = row[" DEBT(%)"] if " DEBT(%)" in row and not pd.isna(row[" DEBT(%)"]) else None
            amcfs.gold_percentage = row[" GOLD(%)"] if " GOLD(%)" in row and not pd.isna(row[" GOLD(%)"]) else None
            amcfs.global_equity_percentage = row[" GLOBAL EQUITY(%)"] if " GLOBAL EQUITY(%)" in row and not pd.isna(row[" GLOBAL EQUITY(%)"]) else None
            amcfs.other_percentage = row[" OTHER(%)"] if " OTHER(%)" in row and not pd.isna(row[" OTHER(%)"]) else None
            amcfs.rs_quard = row["RS QUARD"] if "RS QUARD" in row and not pd.isna(row["RS QUARD"]) else None
            amcfs.expense_ratio = row["EXPENSE RATIO"] if "EXPENSE RATIO" in row and not pd.isna(row["EXPENSE RATIO"]) else None
            amcfs.SOV = row["SOV"] if "SOV" in row and not pd.isna(row["SOV"]) else None
            amcfs.A = row["A"] if "A" in row and not pd.isna(row["A"]) else None
            amcfs.AA = row["AA"] if "AA" in row and not pd.isna(row["AA"]) else None
            amcfs.AAA = row["AAA"] if "AAA" in row and not pd.isna(row["AAA"]) else None
            amcfs.BIG = row["BIG"] if "BIG" in row and not pd.isna(row["BIG"]) else None
            amcfs.cash = row["CASH"] if "CASH" in row and not pd.isna(row["CASH"]) else None
            amcfs.downside_deviation = row["DOWNSIDE DEVIATION"] if "DOWNSIDE DEVIATION" in row and not pd.isna(row["DOWNSIDE DEVIATION"]) else None
            amcfs.downside_probability = row["DOWNSIDE PROBABILITY"] if "DOWNSIDE PROBABILITY" in row and not pd.isna(row["DOWNSIDE PROBABILITY"]) else None
            amcfs.scheme_benchmark = row["SCHEME BENCHMARK"] if "SCHEME BENCHMARK" in row and not pd.isna(row["SCHEME BENCHMARK"]) else None
            amcfs.is_scheme_benchmark = False if " A U M" in row and not pd.isna(row[" A U M"]) else True
            amcfs.save()

    update_avg_category_returns()

    return "Done"

def parse_date_value(value):
    """
    Parse various date formats from Excel into Python date object
    """
    if pd.isna(value) or value is None:
        return None
    
    # If it's already a datetime object
    if isinstance(value, datetime.datetime):
        return value.date()
    elif isinstance(value, datetime.date):
        return value
    
    # If it's a string, try to parse it
    if isinstance(value, str):
        try:
            # Try different date formats
            for fmt in ['%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                try:
                    return datetime.datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        except:
            pass
    
    return None

def parse_numeric_value(value):
    """
    Parse numeric values, handling Excel errors and formulas
    """
    if pd.isna(value) or value is None:
        return None

    # Handle Excel errors
    if isinstance(value, str) and value in ['#NAME?', '#DIV/0!', '#VALUE!', '#REF!']:
        return None

    try:
        return float(value)
    except (ValueError, TypeError):
        return None

def parse_date_from_period(period):
    """
    Parse period value into a date object.
    Handles both datetime objects and string formats.
    """
    import datetime

    if isinstance(period, datetime.datetime):
        return period.date()
    elif isinstance(period, datetime.date):
        return period
    elif isinstance(period, str):
        if '-' in period:
            # Date format: YYYY-MM-DD
            period_clean = period.split(' ')[0]  # Remove time if present
            return datetime.datetime.strptime(period_clean, '%Y-%m-%d').date()
        else:
            # YYYYMM format - convert to end of month
            year = int(period[:4])
            month = int(period[4:6])
            return datetime.date(year, month, 28)
    else:
        raise ValueError(f"Cannot parse period as date: {period}")

def parse_period_to_string(period):
    """
    Convert period value to string format (YYYYMM).
    Handles formulas, floats, and strings.
    """
    if pd.isna(period):
        return None

    # Handle string formulas like "=O8"
    if isinstance(period, str) and period.startswith('='):
        return None  # Skip formulas

    # Handle float/int period codes
    if isinstance(period, (int, float)):
        return str(int(period))

    # Handle string period codes
    if isinstance(period, str):
        try:
            return str(int(float(period)))
        except:
            return None

    return None

def validate_excel_structure(excel_file):
    """
    Validate the Excel file structure and return diagnostic information
    """
    try:
        # Read the full Excel file including headers to validate structure
        df_full = pd.read_excel(excel_file, header=None)
        
        # Read just the data portion
        df_data = pd.read_excel(excel_file, skiprows=8)
        
        diagnostics = {
            'total_rows': len(df_full),
            'total_columns': len(df_full.columns),
            'data_rows': len(df_data),
            'data_columns': len(df_data.columns),
            'header_row_8': df_full.iloc[7].tolist() if len(df_full) > 7 else [],
            'sample_data_row': df_data.iloc[0].tolist() if len(df_data) > 0 else [],
            'columns_with_data': sum(1 for col in df_data.iloc[0] if not pd.isna(col) and col != ''),
            'null_columns': sum(1 for col in df_data.iloc[0] if pd.isna(col) or col == ''),
        }
        
        return diagnostics
    except Exception as e:
        return {'error': str(e)}

def process_underlying_stocks_file(excel_file):
    """
    Process underlying stocks holdings file with robust error handling and batch processing

    Args:
        excel_file: The uploaded Excel file containing holdings data

    Returns:
        dict: Comprehensive statistics about the processing
    """
    try:
        # Auto-detect header row by looking for expected column patterns
        holdings_df = None
        header_row = None

        for skip in range(10):
            try:
                temp_df = pd.read_excel(excel_file, skiprows=skip)
                if not temp_df.empty and len(temp_df.columns) > 10:
                    cols_str = ' '.join([str(col) for col in temp_df.columns])
                    if 'Scheme Code' in cols_str and 'Scheme Name' in cols_str and 'SD_Scheme ISIN' in cols_str:
                        holdings_df = temp_df
                        header_row = skip
                        break
            except Exception:
                continue

        if holdings_df is None:
            holdings_df = pd.read_excel(excel_file)

        if holdings_df.empty:
            raise ValueError("Holdings file contains no data")

        # Define column mappings
        expected_columns = {
            'scheme_code': ['Scheme Code', 'SCHEME CODE', 'scheme_code'],
            'scheme_name': ['Scheme Name', 'SCHEME NAME', 'scheme_name'],
            'sd_scheme_amfi_code': ['SD_Scheme AMFI Code', 'SD SCHEME AMFI CODE', 'sd_scheme_amfi_code'],
            'sd_scheme_isin': ['SD_Scheme ISIN', 'SD SCHEME ISIN', 'sd_scheme_isin'],
            'pd_month_end': ['PD_Month End', 'PD MONTH END', 'pd_month_end'],
            'pd_date': ['PD_Date', 'PD DATE', 'pd_date'],
            'pd_instrument_name': ['PD_Instrument Name', 'PD INSTRUMENT NAME', 'pd_instrument_name'],
            'pd_holding_percent': ['PD_Holding (%)', 'PD HOLDING (%)', 'pd_holding_percent'],
            'pd_market_value': ['PD_Market Value', 'PD MARKET VALUE', 'pd_market_value'],
            'pd_no_of_shares': ['PD_No of Shares', 'PD NO OF SHARES', 'pd_no_of_shares'],
            'pd_bse_code': ['PD_BSE Code', 'PD BSE CODE', 'pd_bse_code'],
            'pd_nse_symbol': ['PD_NSE Symbol', 'PD NSE SYMBOL', 'pd_nse_symbol'],
            'pd_company_isin': ['PD_Company ISIN no', 'PD COMPANY ISIN NO', 'pd_company_isin']
        }

        # Map columns
        column_mapping = {}
        available_columns = holdings_df.columns.tolist()

        for key, possible_names in expected_columns.items():
            found = False
            for possible_name in possible_names:
                if possible_name in available_columns:
                    column_mapping[key] = possible_name
                    found = True
                    break
            if not found:
                for col in available_columns:
                    if any(name.lower() in str(col).lower() for name in possible_names):
                        column_mapping[key] = col
                        found = True
                        break

        # Validate required columns
        required_columns = ['scheme_name', 'sd_scheme_isin', 'pd_instrument_name']
        missing_columns = [col for col in required_columns if col not in column_mapping]
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")

        # **PHASE 1: DATA CLEANING AND VALIDATION**
        print(f"Initial data: {len(holdings_df)} rows")

        # Remove rows with NaN in critical fields
        holdings_df = holdings_df.dropna(subset=[
            column_mapping['scheme_name'],
            column_mapping['sd_scheme_isin'],
            column_mapping['pd_instrument_name']
        ]).copy()
        print(f"After removing NaN critical fields: {len(holdings_df)} rows")

        # Convert ISIN column to string and filter out 'nan' strings
        holdings_df['clean_isin'] = holdings_df[column_mapping['sd_scheme_isin']].astype(str).str.strip()
        holdings_df = holdings_df[~holdings_df['clean_isin'].isin(['nan', 'None', ''])].copy()
        print(f"After removing invalid ISINs: {len(holdings_df)} rows")

        # **PHASE 2: DEDUPLICATION**
        original_count = len(holdings_df)

        # Create composite key for deduplication
        holdings_df['composite_key'] = (
            holdings_df['clean_isin'] + '|' +
            holdings_df[column_mapping['pd_instrument_name']].astype(str) + '|' +
            holdings_df[column_mapping['pd_date']].astype(str)
        )

        # Remove exact duplicates
        holdings_df = holdings_df.drop_duplicates(subset=['composite_key']).copy()
        duplicates_removed = original_count - len(holdings_df)
        print(f"Removed {duplicates_removed} duplicate holdings, {len(holdings_df)} unique records remain")

        # **PHASE 3: PREPARE UNIQUE SCHEMES DATA**
        unique_schemes = holdings_df.groupby('clean_isin').agg({
            column_mapping['scheme_name']: 'first',
            column_mapping.get('sd_scheme_amfi_code', column_mapping['scheme_name']): 'first'
        }).reset_index()
        print(f"Found {len(unique_schemes)} unique schemes to process")

        # Initialize comprehensive statistics
        stats = {
            'total_rows_in_file': original_count,
            'duplicates_removed': duplicates_removed,
            'unique_records_processed': len(holdings_df),
            'schemes_created': 0,
            'schemes_updated': 0,
            'schemes_failed': 0,
            'stocks_matched': 0,
            'stocks_created': 0,
            'stocks_failed': 0,
            'holdings_created': 0,
            'holdings_failed': 0,
            'batch_errors': 0
        }

        # **PHASE 4: BULK PROCESS SCHEMES FIRST**
        print("Processing schemes...")
        existing_schemes = {s.isin_number: s for s in AMCFundScheme.objects.all()}

        schemes_to_create = []
        schemes_to_update = []

        for _, scheme_row in unique_schemes.iterrows():
            scheme_isin = scheme_row['clean_isin']
            scheme_name = str(scheme_row[column_mapping['scheme_name']]).strip()

            if scheme_isin in existing_schemes:
                # Update existing scheme
                existing_scheme = existing_schemes[scheme_isin]
                existing_scheme.accord_mf_name = scheme_name
                existing_scheme.is_active = True
                schemes_to_update.append(existing_scheme)
                stats['schemes_updated'] += 1
            else:
                # Prepare new scheme
                unique_scheme_name = f"{scheme_name} ({scheme_isin})"
                amfi_code = None
                if column_mapping.get('sd_scheme_amfi_code'):
                    try:
                        amfi_val = scheme_row[column_mapping['sd_scheme_amfi_code']]
                        if pd.notna(amfi_val):
                            amfi_code = int(float(str(amfi_val)))
                    except (ValueError, TypeError):
                        pass

                schemes_to_create.append(AMCFundScheme(
                    name=unique_scheme_name,
                    accord_mf_name=scheme_name,
                    isin_number=scheme_isin,
                    amfi_scheme_code=amfi_code,
                    is_active=True
                ))
                stats['schemes_created'] += 1

        # Bulk create/update schemes
        try:
            if schemes_to_create:
                AMCFundScheme.objects.bulk_create(schemes_to_create, ignore_conflicts=True)
            if schemes_to_update:
                AMCFundScheme.objects.bulk_update(schemes_to_update, ['accord_mf_name', 'is_active'])
            print(f"Schemes processed: {stats['schemes_created']} created, {stats['schemes_updated']} updated")
        except Exception as e:
            print(f"Error in bulk scheme operations: {str(e)}")
            stats['schemes_failed'] = len(schemes_to_create) + len(schemes_to_update)

        # Refresh scheme lookup after bulk operations
        existing_schemes = {s.isin_number: s for s in AMCFundScheme.objects.all()}

        # **PHASE 5: BATCH PROCESS HOLDINGS**
        print("Processing holdings...")
        batch_size = 500
        total_batches = (len(holdings_df) + batch_size - 1) // batch_size

        for batch_num in range(total_batches):
            start_idx = batch_num * batch_size
            end_idx = min(start_idx + batch_size, len(holdings_df))
            batch_df = holdings_df.iloc[start_idx:end_idx]

            # Process each row in the batch individually to prevent rollback issues
            holdings_to_create = []

            for _, row in batch_df.iterrows():
                try:
                    # Get scheme
                    scheme_isin = row['clean_isin']
                    scheme = existing_schemes.get(scheme_isin)
                    if not scheme:
                        stats['holdings_failed'] += 1
                        continue

                    # Find or create stock (with individual transaction to prevent rollback)
                    try:
                        with transaction.atomic():
                            stock = find_or_create_stock(row, column_mapping, stats)
                    except Exception as e:
                        print(f"Error with stock operation: {str(e)}")
                        stats['holdings_failed'] += 1
                        continue

                    if not stock:
                        stats['holdings_failed'] += 1
                        continue

                    # Parse holding data
                    holding_date = parse_holding_date(row, column_mapping)
                    holding_percentage = safe_float_conversion(row.get(column_mapping.get('pd_holding_percent')))
                    market_value = safe_float_conversion(row.get(column_mapping.get('pd_market_value')))
                    number_of_shares = safe_float_conversion(row.get(column_mapping.get('pd_no_of_shares')))

                    holdings_to_create.append(FundHolding(
                        scheme=scheme,
                        stock=stock,
                        holding_date=holding_date,
                        holding_percentage=holding_percentage,
                        market_value=market_value,
                        number_of_shares=number_of_shares
                    ))

                except Exception as e:
                    print(f"Error preparing holding record: {str(e)}")
                    stats['holdings_failed'] += 1

            # Bulk create holdings for this batch
            if holdings_to_create:
                try:
                    with transaction.atomic():
                        created_holdings = FundHolding.objects.bulk_create(holdings_to_create, ignore_conflicts=True)
                        stats['holdings_created'] += len(created_holdings)
                        print(f"Batch {batch_num + 1}/{total_batches} completed: {len(created_holdings)} holdings created")
                except Exception as e:
                    print(f"Batch {batch_num + 1} bulk create failed: {str(e)}")
                    stats['batch_errors'] += 1
                    stats['holdings_failed'] += len(holdings_to_create)
            else:
                print(f"Batch {batch_num + 1}/{total_batches} completed: 0 valid holdings to create")

        print(f"Processing complete. Holdings created: {stats['holdings_created']}, failed: {stats['holdings_failed']}")
        return stats

    except Exception as e:
        raise ValueError(f"Error processing holdings file: {str(e)}")


def safe_float_conversion(value, default=None):
    """Safely convert value to float with fallback"""
    try:
        if pd.isna(value) or str(value).strip() in ['', 'nan', 'NaN', 'None']:
            return default
        return float(str(value).strip())
    except (ValueError, TypeError):
        return default


def safe_int_conversion(value, default=None):
    """Safely convert value to int with fallback"""
    try:
        if pd.isna(value) or str(value).strip() in ['', 'nan', 'NaN', 'None']:
            return default
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def parse_holding_date(row, column_mapping):
    """Parse holding date with fallback to today"""
    holding_date = datetime.date.today()
    if column_mapping.get('pd_date'):
        try:
            date_val = row[column_mapping.get('pd_date')]
            if pd.notna(date_val):
                if isinstance(date_val, datetime.datetime):
                    holding_date = date_val.date()
                elif isinstance(date_val, datetime.date):
                    holding_date = date_val
                else:
                    holding_date = pd.to_datetime(str(date_val)).date()
        except Exception:
            pass
    return holding_date


def find_or_create_stock(row, column_mapping, stats):
    """Find existing stock or create new one"""
    try:
        instrument_name = str(row[column_mapping['pd_instrument_name']]).strip()
        company_isin = str(row.get(column_mapping.get('pd_company_isin', ''), '')).strip()
        bse_code = str(row.get(column_mapping.get('pd_bse_code', ''), '')).strip()
        nse_symbol = str(row.get(column_mapping.get('pd_nse_symbol', ''), '')).strip()

        # Clean values
        company_isin = company_isin if company_isin not in ['nan', 'None', ''] else None
        bse_code = safe_int_conversion(bse_code)
        nse_symbol = nse_symbol if nse_symbol not in ['nan', 'None', ''] else None

        # Try to match existing stock
        stock = None
        if company_isin:
            stock = Stock.objects.filter(isin=company_isin).first()
        if not stock and bse_code:
            stock = Stock.objects.filter(bse_code=bse_code).first()
        if not stock and nse_symbol:
            stock = Stock.objects.filter(nse_symbol=nse_symbol).first()

        if stock:
            stats['stocks_matched'] += 1
            return stock
        else:
            # Create new stock with unique accord_code
            import uuid
            unique_code = f"AUTO_{uuid.uuid4().hex[:8]}"
            stock = Stock.objects.create(
                company_name=instrument_name,
                accord_code=unique_code,
                isin=company_isin,
                bse_code=bse_code,
                nse_symbol=nse_symbol,
                sector='Unknown'
            )
            stats['stocks_created'] += 1
            return stock

    except Exception as e:
        print(f"Error finding/creating stock: {str(e)}")
        stats['stocks_failed'] += 1
        return None

def validate_import_completeness(analysis_results, stats, filename):
    """
    Validates that the import process captured all detected periods without data loss.
    Creates FileStructureMetadata record and validates imported data against detected structure.

    Args:
        analysis_results: Results from DynamicStockSheetAnalyzer
        stats: Import statistics
        filename: Original filename

    Returns:
        dict: Validation results with errors, warnings, and metadata
    """
    import uuid
    from collections import defaultdict

    validation_results = {
        'has_errors': False,
        'has_warnings': False,
        'errors': [],
        'warnings': [],
        'metadata_id': None,
        'detected_vs_imported': {}
    }

    try:
        # Generate unique session ID for this upload
        upload_session_id = f"upload_{uuid.uuid4().hex[:12]}_{int(timezone.now().timestamp())}"

        # Create FileStructureMetadata record
        metadata = FileStructureMetadata.objects.create(
            upload_session_id=upload_session_id,
            original_filename=filename,
            total_columns=analysis_results['total_columns'],
            market_cap_column_mapping=analysis_results['column_mapping']['time_series_ranges'].get('market_cap', {}),
            ttm_column_mapping=analysis_results['column_mapping']['time_series_ranges'].get('ttm', {}),
            quarterly_column_mapping=analysis_results['column_mapping']['time_series_ranges'].get('quarterly', {}),
            annual_column_mapping=analysis_results['column_mapping']['time_series_ranges'].get('annual', {}),
            price_column_mapping=analysis_results['column_mapping']['time_series_ranges'].get('price', {}),
            detected_periods=analysis_results['column_mapping']['periods'],
            basic_columns_end=12,
            import_status='completed',
            records_imported=stats['total_stocks']
        )
        validation_results['metadata_id'] = metadata.file_structure_id

        # Get actual imported periods from database
        actual_periods = {
            'market_cap': list(StockMarketCap.objects.values_list('date', flat=True).distinct()),
            'ttm': list(StockTTMData.objects.values_list('period', flat=True).distinct()),
            'quarterly': list(StockQuarterlyData.objects.values_list('period', flat=True).distinct()),
            'annual': list(StockAnnualRatios.objects.values_list('financial_year', flat=True).distinct()),
            'price': list(StockPrice.objects.values_list('price_date', flat=True).distinct())
        }

        # Convert dates to strings for comparison
        actual_periods['market_cap'] = [d.strftime('%Y-%m-%d') for d in actual_periods['market_cap']]
        actual_periods['price'] = [d.strftime('%Y-%m-%d') for d in actual_periods['price']]

        # Compare detected vs imported periods
        detected_periods = analysis_results['column_mapping']['periods']

        for category in ['market_cap', 'ttm', 'quarterly', 'annual', 'price']:
            detected = set(detected_periods.get(category, []))
            imported = set(actual_periods.get(category, []))

            validation_results['detected_vs_imported'][category] = {
                'detected_count': len(detected),
                'imported_count': len(imported),
                'detected_periods': sorted(list(detected)),
                'imported_periods': sorted(list(imported)),
                'missing_periods': sorted(list(detected - imported)),
                'extra_periods': sorted(list(imported - detected))
            }

            # Check for missing periods (data loss)
            missing = detected - imported
            if missing:
                validation_results['has_errors'] = True
                validation_results['errors'].append(
                    f"{category.upper()}: {len(missing)} periods detected but not imported: {sorted(list(missing))}"
                )

            # Check for extra periods (unexpected data)
            extra = imported - detected
            if extra:
                validation_results['has_warnings'] = True
                validation_results['warnings'].append(
                    f"{category.upper()}: {len(extra)} extra periods imported not detected in file: {sorted(list(extra))}"
                )

        # Validate record counts
        expected_max_records = stats['total_stocks']
        for category, count in [
            ('market_cap', stats['market_cap_records']),
            ('ttm', stats['ttm_records']),
            ('quarterly', stats['quarterly_records']),
            ('annual', stats['annual_ratios_records']),
            ('price', stats['price_records'])
        ]:
            detected_periods_count = len(detected_periods.get(category, []))
            expected_records = expected_max_records * detected_periods_count

            if count < (expected_records * 0.8):  # Allow 20% tolerance for missing data
                validation_results['has_warnings'] = True
                validation_results['warnings'].append(
                    f"{category.upper()}: Low record count - expected ~{expected_records}, got {count}"
                )

        # Update metadata with validation results
        metadata.validation_errors = {
            'errors': validation_results['errors'],
            'warnings': validation_results['warnings'],
            'detected_vs_imported': validation_results['detected_vs_imported']
        }

        if validation_results['has_errors']:
            metadata.import_status = 'completed_with_errors'
        elif validation_results['has_warnings']:
            metadata.import_status = 'completed_with_warnings'
        else:
            metadata.import_status = 'completed_successfully'

        metadata.save()

    except Exception as e:
        validation_results['has_errors'] = True
        validation_results['errors'].append(f"Validation process failed: {str(e)}")

    return validation_results

def process_stocks_base_sheet(excel_file, session_id=None, user=None):
    """
    HEADER-DRIVEN Process the Stocks Base Sheet Excel file.
    Reads Rows 6, 7, and 8 to identify column types and periods.
    Handles files with any number of periods automatically.

    Args:
        excel_file: The uploaded Excel file
        session_id: Optional session ID for progress tracking
        user: Optional user instance for progress tracking
    """
    from .models import UploadProgressSession
    from django.utils import timezone

    try:
        print("=== HEADER-DRIVEN STOCK SHEET PROCESSING ===")

        # Initialize the dynamic analyzer
        analyzer = DynamicStockSheetAnalyzer()

        # Use NEW header-driven analysis
        print("Step 1: Analyzing Excel structure (reading Rows 6, 7, 8)...")
        analysis_results = analyzer.analyze_excel_structure_header_driven(excel_file)

        # Log analysis results
        print(f"Step 2: Structure analysis completed:")
        print(f"   - Total columns detected: {analysis_results['total_columns']}")
        print(f"   - Analysis method: {analysis_results['method']}")

        # Get complete column mapping
        complete_column_mapping = analysis_results['complete_column_mapping']

        # Count data types
        data_type_counts = {}
        for col_info in complete_column_mapping.values():
            data_type = col_info['data_type']
            data_type_counts[data_type] = data_type_counts.get(data_type, 0) + 1

        print(f"   - Column types detected: {data_type_counts}")

        # Read the data portion (skip headers)
        print("Step 3: Reading data rows...")
        df = pd.read_excel(excel_file, skiprows=8)
        total_rows = len(df)
        print(f"   - Data rows to process: {total_rows}")

        # Create progress session if session_id provided
        progress_session = None
        if session_id:
            filename = getattr(excel_file, 'name', 'unknown')
            progress_session = UploadProgressSession.objects.create(
                session_id=session_id,
                user=user,
                filename=filename,
                file_type='stocks_base_sheet',
                total_rows=total_rows,
                status='processing'
            )

        # Statistics for tracking
        stats = {
            'total_stocks': 0,
            'stocks_created': 0,
            'stocks_updated': 0,
            'market_cap_records': 0,
            'ttm_records': 0,
            'quarterly_records': 0,
            'annual_ratios_records': 0,
            'price_records': 0,
            'skipped_rows': 0,
            'analysis_results': analysis_results
        }

        # Process each row using header-driven mapping
        print("Step 4: Processing stock data...")

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # First pass: Extract basic info
                    basic_info = {}
                    for col_idx, col_info in complete_column_mapping.items():
                        if col_info['data_type'] == 'basic_info':
                            field_name = str(col_info['period']).strip().lower().replace(' ', '_').replace('.', '')
                            value = row.iloc[col_idx]
                            basic_info[field_name] = value

                    # Get required fields
                    company_name = basic_info.get('company_name')
                    accord_code = basic_info.get('accord_code')

                    if pd.isna(company_name) or pd.isna(accord_code):
                        stats['skipped_rows'] += 1
                        continue

                    company_name = str(company_name).strip()
                    accord_code = str(accord_code).strip()

                    if not company_name or not accord_code:
                        stats['skipped_rows'] += 1
                        continue

                    stats['total_stocks'] += 1

                    # Prepare stock data
                    stock_data = {
                        'company_name': company_name,
                        'sector': str(basic_info.get('sector', '')) if not pd.isna(basic_info.get('sector')) else '',
                        'cap': str(basic_info.get('cap', '')) if not pd.isna(basic_info.get('cap')) else '',
                    }

                    # Add optional basic fields from basic_info dictionary
                    optional_fields_map = {
                        'free_float': 'free_float',
                        'revenue_6yr_cagr': '6_year_cagr',  # May need adjustment based on actual header
                        'revenue_ttm': 'ttm',
                        'pat_6yr_cagr': '6_year_cagr',
                        'pat_ttm': 'ttm',
                        'current_value': 'current',
                        'two_yr_avg': '2_yr_avg',
                        'reval_deval': 'reval/deval'
                    }

                    for stock_field, header_key in optional_fields_map.items():
                        if header_key in basic_info:
                            value = basic_info[header_key]
                            if not pd.isna(value):
                                stock_data[stock_field] = parse_numeric_value(value)

                    # Extract identifiers from complete column mapping
                    for col_idx, col_info in complete_column_mapping.items():
                        data_type = col_info['data_type']
                        value = row.iloc[col_idx]

                        if data_type == 'bse_code' and not pd.isna(value):
                            stock_data['bse_code'] = str(value).strip()
                        elif data_type == 'nse_symbol' and not pd.isna(value):
                            stock_data['nse_symbol'] = str(value).strip()
                        elif data_type == 'isin' and not pd.isna(value):
                            stock_data['isin'] = str(value).strip()

                    # Create or update Stock record
                    stock, created = Stock.objects.update_or_create(
                        accord_code=accord_code,
                        defaults=stock_data
                    )

                    if created:
                        stats['stocks_created'] += 1
                    else:
                        stats['stocks_updated'] += 1

                    # Update progress session every 5 rows or on first 3 rows
                    if progress_session and (stats['total_stocks'] % 5 == 0 or stats['total_stocks'] <= 3):
                        progress_percentage = (stats['total_stocks'] / total_rows) * 100

                        # Use autonomous transaction to commit progress immediately
                        # This makes progress visible to API endpoint while main transaction continues
                        from django.db import transaction as progress_transaction
                        with progress_transaction.atomic():
                            # Refresh to avoid stale data
                            progress_session.refresh_from_db()
                            progress_session.processed_rows = stats['total_stocks']
                            progress_session.current_stock_name = company_name
                            progress_session.progress_percentage = progress_percentage
                            progress_session.stocks_created = stats['stocks_created']
                            progress_session.stocks_updated = stats['stocks_updated']
                            progress_session.save(update_fields=[
                                'processed_rows', 'current_stock_name', 'progress_percentage',
                                'stocks_created', 'stocks_updated'
                            ])
                        # Progress update committed here, visible to other queries immediately

                    # Debug output for first few rows
                    if stats['total_stocks'] <= 3:
                        print(f"=== Processing Stock {stats['total_stocks']}: {company_name} ({accord_code}) ===")

                    # Second pass: Process time-series data using column mapping
                    for col_idx, col_info in complete_column_mapping.items():
                        if col_info['is_separator'] or col_info['data_type'] in ['basic_info', 'bse_code', 'nse_symbol', 'isin', 'unknown']:
                            continue

                        data_type = col_info['data_type']
                        period = col_info['period']
                        value = row.iloc[col_idx]

                        # Skip Excel formula errors
                        if isinstance(value, str) and value in ['#NAME?', '#REF!', '#VALUE!', '#DIV/0!', '#N/A', '#NULL!']:
                            continue

                        if pd.isna(value):
                            continue

                        value = parse_numeric_value(value)
                        if value is None:
                            continue

                        # Store data based on data type identified from headers
                        if data_type == 'market_cap':
                            try:
                                date_obj = parse_date_from_period(period)
                                StockMarketCap.objects.update_or_create(
                                    stock=stock, date=date_obj,
                                    defaults={'market_cap': value}
                                )
                                stats['market_cap_records'] += 1
                            except:
                                continue

                        elif data_type == 'market_cap_free_float':
                            try:
                                date_obj = parse_date_from_period(period)
                                StockMarketCap.objects.update_or_create(
                                    stock=stock, date=date_obj,
                                    defaults={'market_cap_free_float': value}
                                )
                                stats['market_cap_records'] += 1
                            except:
                                continue

                        elif data_type == 'ttm_revenue':
                            period_str = parse_period_to_string(period)
                            if period_str:
                                StockTTMData.objects.update_or_create(
                                    stock=stock, period=period_str,
                                    defaults={'ttm_revenue': value}
                                )
                                stats['ttm_records'] += 1

                        elif data_type == 'ttm_revenue_free_float':
                            period_str = parse_period_to_string(period)
                            if period_str:
                                StockTTMData.objects.update_or_create(
                                    stock=stock, period=period_str,
                                    defaults={'ttm_revenue_free_float': value}
                                )
                                stats['ttm_records'] += 1

                        elif data_type == 'ttm_pat':
                            period_str = parse_period_to_string(period)
                            if period_str:
                                StockTTMData.objects.update_or_create(
                                    stock=stock, period=period_str,
                                    defaults={'ttm_pat': value}
                                )
                                stats['ttm_records'] += 1

                        elif data_type == 'ttm_pat_free_float':
                            period_str = parse_period_to_string(period)
                            if period_str:
                                StockTTMData.objects.update_or_create(
                                    stock=stock, period=period_str,
                                    defaults={'ttm_pat_free_float': value}
                                )
                                stats['ttm_records'] += 1

                        elif data_type == 'quarterly_revenue':
                            period_str = parse_period_to_string(period)
                            if period_str:
                                StockQuarterlyData.objects.update_or_create(
                                    stock=stock, period=period_str,
                                    defaults={'quarterly_revenue': value}
                                )
                                stats['quarterly_records'] += 1

                        elif data_type == 'quarterly_revenue_free_float':
                            period_str = parse_period_to_string(period)
                            if period_str:
                                StockQuarterlyData.objects.update_or_create(
                                    stock=stock, period=period_str,
                                    defaults={'quarterly_revenue_free_float': value}
                                )
                                stats['quarterly_records'] += 1

                        elif data_type == 'quarterly_pat':
                            period_str = parse_period_to_string(period)
                            if period_str:
                                StockQuarterlyData.objects.update_or_create(
                                    stock=stock, period=period_str,
                                    defaults={'quarterly_pat': value}
                                )
                                stats['quarterly_records'] += 1

                        elif data_type == 'quarterly_pat_free_float':
                            period_str = parse_period_to_string(period)
                            if period_str:
                                StockQuarterlyData.objects.update_or_create(
                                    stock=stock, period=period_str,
                                    defaults={'quarterly_pat_free_float': value}
                                )
                                stats['quarterly_records'] += 1

                        elif data_type == 'roce':
                            financial_year = str(period).strip()
                            StockAnnualRatios.objects.update_or_create(
                                stock=stock, financial_year=financial_year,
                                defaults={'roce_percentage': value}
                            )
                            stats['annual_ratios_records'] += 1

                        elif data_type == 'roe':
                            financial_year = str(period).strip()
                            StockAnnualRatios.objects.update_or_create(
                                stock=stock, financial_year=financial_year,
                                defaults={'roe_percentage': value}
                            )
                            stats['annual_ratios_records'] += 1

                        elif data_type == 'retention':
                            financial_year = str(period).strip()
                            StockAnnualRatios.objects.update_or_create(
                                stock=stock, financial_year=financial_year,
                                defaults={'retention_percentage': value}
                            )
                            stats['annual_ratios_records'] += 1

                        elif data_type == 'share_price':
                            try:
                                date_obj = parse_date_from_period(period)
                                StockPrice.objects.update_or_create(
                                    stock=stock, price_date=date_obj,
                                    defaults={'share_price': value}
                                )
                                stats['price_records'] += 1
                            except:
                                continue

                        elif data_type == 'pr_ratio':
                            try:
                                date_obj = parse_date_from_period(period)
                                StockPrice.objects.update_or_create(
                                    stock=stock, price_date=date_obj,
                                    defaults={'pr_ratio': value}
                                )
                                stats['price_records'] += 1
                            except:
                                continue

                        elif data_type == 'pe_ratio':
                            try:
                                date_obj = parse_date_from_period(period)
                                StockPrice.objects.update_or_create(
                                    stock=stock, price_date=date_obj,
                                    defaults={'pe_ratio': value}
                                )
                                stats['price_records'] += 1
                            except:
                                continue


                except Exception as e:
                    print(f"Error processing row {index}: {str(e)}")
                    stats['skipped_rows'] += 1
                    continue

        print("=== PROCESSING COMPLETED ===")
        print(f"Summary:")
        print(f"   - Total stocks: {stats['total_stocks']}")
        print(f"   - Created: {stats['stocks_created']}, Updated: {stats['stocks_updated']}")
        print(f"   - Market cap records: {stats['market_cap_records']}")
        print(f"   - TTM records: {stats['ttm_records']}")
        print(f"   - Quarterly records: {stats['quarterly_records']}")
        print(f"   - Annual ratios: {stats['annual_ratios_records']}")
        print(f"   - Price records: {stats['price_records']}")
        print(f"   - Skipped rows: {stats['skipped_rows']}")

        # Mark progress session as completed
        if progress_session:
            total_records = (stats['market_cap_records'] + stats['ttm_records'] +
                           stats['quarterly_records'] + stats['annual_ratios_records'] +
                           stats['price_records'])

            # Use separate transaction for final progress update
            with transaction.atomic():
                progress_session.refresh_from_db()
                progress_session.status = 'completed'
                progress_session.progress_percentage = 100.0
                progress_session.processed_rows = total_rows
                progress_session.records_created = total_records
                progress_session.completed_at = timezone.now()
                progress_session.save()

        print("✅ Import completed successfully")

        return stats

    except Exception as e:
        # Mark progress session as failed
        if progress_session:
            # Use separate transaction for error status update
            with transaction.atomic():
                progress_session.refresh_from_db()
                progress_session.status = 'failed'
                progress_session.error_message = str(e)
                progress_session.completed_at = timezone.now()
                progress_session.save()

        raise ValueError(f"Error in header-driven stock sheet processing: {str(e)}")

@login_required
def process_amcfs_nav_and_returns(request):
    """
    View to process either TopSchemes data or RATIOS PE data 
    based on the selected file type in the form
    """
    form = MasterDataExcelUploadForm()
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        form = MasterDataExcelUploadForm(request.POST, request.FILES)
        
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            file_type = form.cleaned_data['file_type']

            # Check if the file is an Excel file
            if not excel_file.name.endswith(('.xls', '.xlsx')):
                messages.error(request, "The uploaded file is not an Excel file. Please upload a valid Excel file.")
                return render(request, 'gcia_app/upload_amcfs_excel.html', {'form': form})

            try:
                if file_type == 'top_schemes':
                    # Process Top Schemes File
                    stats = process_top_schemes_data_uploaded(excel_file)
                elif file_type == 'ratios_pe':
                    # Process RATIOS PE file
                    stats = process_ratios_pe_file(excel_file)
                elif file_type == "index_nav":
                    # Process INDEX NAV file
                    stats = process_index_nav_file(excel_file)
                    try:
                        pe_ratio = get_bse500_pe_ratio()
                        if pe_ratio:
                            benchmark_index = AMCFundScheme.objects.filter(name="BSE 500 - TRI").first()
                            benchmark_index.pe_ratio = pe_ratio
                            benchmark_index.save()
                        messages.success(request, "BSE 500 TRI PE Updated Successfully")
                    except Exception as e:
                        print(str(e))
                        print(traceback.format_exc())
                elif file_type == 'stocks_base_sheet':
                    # Process Stocks Base Sheet file
                    # Get session_id from request if provided (for AJAX uploads)
                    session_id = request.POST.get('session_id')
                    stats = process_stocks_base_sheet(excel_file, session_id=session_id, user=request.user)

                    # Create detailed success message with validation info
                    total_records = stats['market_cap_records'] + stats['ttm_records'] + stats['quarterly_records'] + stats['annual_ratios_records'] + stats['price_records']
                    validation_info = ""

                    if 'validation_results' in stats:
                        validation = stats['validation_results']
                        if validation.get('has_errors'):
                            messages.error(request, f"❌ Stock processing completed with errors! Please check data integrity. Errors: {'; '.join(validation['errors'])}")
                        elif validation.get('has_warnings'):
                            messages.warning(request, f"⚠️ Stock processing completed with warnings! Data imported but please review. Warnings: {'; '.join(validation['warnings'])}")
                            validation_info = f" ✅ Validation: {len(validation.get('warnings', []))} warnings detected"
                        else:
                            validation_info = " ✅ Validation: No data loss detected"

                    if not stats.get('validation_results', {}).get('has_errors'):
                        success_msg = (
                            f"🎉 Stocks data processed successfully! "
                            f"📊 Stocks: {stats['stocks_created']} created, {stats['stocks_updated']} updated | "
                            f"📈 Total Records: {total_records} (Market Cap: {stats['market_cap_records']}, "
                            f"TTM: {stats['ttm_records']}, Quarterly: {stats['quarterly_records']}, "
                            f"Annual: {stats['annual_ratios_records']}, Price: {stats['price_records']})"
                            f"{validation_info}"
                        )
                        messages.success(request, success_msg)
                elif file_type == 'underlying_stocks':
                    # Process Underlying Stocks Holdings file
                    stats = process_underlying_stocks_file(excel_file)
                    success_msg = (
                        f"Holdings processed successfully! "
                        f"📊 Total: {stats['total_rows_in_file']} rows, "
                        f"✅ Processed: {stats['unique_records_processed']} unique, "
                        f"🔄 Duplicates removed: {stats['duplicates_removed']}, "
                        f"🏛️ Schemes: {stats['schemes_created']} created + {stats['schemes_updated']} updated, "
                        f"📈 Stocks: {stats['stocks_created']} created + {stats['stocks_matched']} matched, "
                        f"🎯 Holdings: {stats['holdings_created']} created"
                    )
                    if stats.get('holdings_failed', 0) > 0 or stats.get('batch_errors', 0) > 0:
                        success_msg += f", ⚠️ {stats.get('holdings_failed', 0)} failed"
                    messages.success(request, success_msg)

                messages.success(request, "File uploaded and data processed successfully!")

            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                return render(request, 'gcia_app/upload_amcfs_excel.html', {'form': form})

    return render(request, 'gcia_app/upload_amcfs_excel.html', {'form': form})

def get_concentration_of_scheme(scheme):
    if not scheme.number_of_underlying_stocks:
        return "-"
    elif scheme.number_of_underlying_stocks < 40:
        return "High"
    elif scheme.number_of_underlying_stocks >= 40 and scheme.number_of_underlying_stocks <= 60:
        return "Medium"
    elif scheme.number_of_underlying_stocks > 60:
        return "Low"
    
def evaluate_fund_performance(scheme, index):
    """
    Evaluates fund performance against index and category averages using Django querysets
    
    Parameters:
    scheme (QuerySet object): Django QuerySet object for the fund scheme
    index (QuerySet object): Django QuerySet object for the index
        
    Returns:
    tuple: (performance_rating, valuation_status)
        - performance_rating: "High", "Medium", or "Low"
        - valuation_status: "Undervalued", "Fairly valued", or "Overvalued"
    """
    # Replace None values with 0 for all required fields
    scheme_returns_1_yr = scheme.returns_1_yr if scheme.returns_1_yr else 0
    scheme_returns_3_yr = scheme.returns_3_yr if scheme.returns_3_yr else 0
    scheme_returns_5_yr = scheme.returns_5_yr if scheme.returns_5_yr else 0
    
    scheme_fund_class_avg_1_yr = scheme.fund_class_avg_1_yr_returns if scheme.fund_class_avg_1_yr_returns else 0
    scheme_fund_class_avg_3_yr = scheme.fund_class_avg_3_yr_returns if scheme.fund_class_avg_3_yr_returns else 0
    scheme_fund_class_avg_5_yr = scheme.fund_class_avg_5_yr_returns if scheme.fund_class_avg_5_yr_returns else 0
    
    index_returns_1_yr = index.returns_1_yr if index.returns_1_yr else 0
    index_returns_3_yr = index.returns_3_yr if index.returns_3_yr else 0
    index_returns_5_yr = index.returns_5_yr if index.returns_5_yr else 0
    
    scheme_pe_ratio = scheme.pe_ratio if scheme.pe_ratio else 0
    index_pe_ratio = index.pe_ratio if index.pe_ratio else 0
    
    # Count periods where fund beats both index and category
    periods_beating_both = 0
    
    # 1-year comparison
    if (scheme_returns_1_yr > index_returns_1_yr and 
        scheme_returns_1_yr > scheme_fund_class_avg_1_yr):
        periods_beating_both += 1
    
    # 3-year comparison
    if (scheme_returns_3_yr > index_returns_3_yr and 
        scheme_returns_3_yr > scheme_fund_class_avg_3_yr):
        periods_beating_both += 1
    
    # 5-year comparison
    if (scheme_returns_5_yr > index_returns_5_yr and 
        scheme_returns_5_yr > scheme_fund_class_avg_5_yr):
        periods_beating_both += 1
    
    # Determine performance rating based on periods beating both index and category
    if periods_beating_both >= 2:
        performance_rating = "High"
    elif periods_beating_both >= 1:
        performance_rating = "Medium"
    else:
        performance_rating = "Low"
    
    # Determine valuation status based on PE ratio comparison
    # Handle special case where index PE is 0 to avoid division by zero
    if index_pe_ratio == 0:
        if scheme_pe_ratio == 0:
            valuation_status = "Fairly valued"  # Both are 0, consider them equal
        else:
            valuation_status = "Overvalued"  # Scheme has PE but index doesn't
    else:
        if scheme_pe_ratio < index_pe_ratio:
            valuation_status = "Undervalued"
        else:
            # Calculate percentage difference
            pe_difference_percentage = ((scheme_pe_ratio - index_pe_ratio) / index_pe_ratio) * 100
            
            if pe_difference_percentage > 10:
                valuation_status = "Overvalued"
            else:
                valuation_status = "Fairly valued"
    
    return performance_rating, valuation_status

@login_required
def process_portfolio_valuation(request):
    form = ExcelUploadForm()
    
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']

        # Check if the file is an Excel file
        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, "The uploaded file is not an Excel file. Please upload a valid Excel file.")
            return render(request, 'gcia_app/portfolio_valuation.html', {'form': form})  # Return the form with error message

        try:
            # # Load the workbook and select the active sheet
            # wb = load_workbook(excel_file)
            # sheet = wb.active

            # # Initialize variables
            # headers = []
            # transaction_data = []
            # summary_data = []
            # current_scheme = None
            # scheme_folio = None
            # avg_nav = current_price = balance_units = None
            # client_details = {}
            # client_name = None

            # for row in sheet.iter_rows():
            #     values = [cell.value for cell in row]

            #     # Skip empty rows
            #     if all(v is None for v in values):
            #         continue

            #     # Detect headers for transactions
            #     if 'TRANSACTION TYPE' in values:
            #         headers = values
            #         continue

            #     if 'Client:' in values[0]:
            #         client_name, pan_number = values[0].replace("Client:","").replace(")", "").split("(")
            #         relationship_manager = values[2].split("Relationship Manager:")[1].strip()
            #         client_details["Client Name"] = client_name.strip()
            #         client_details["Client Pan"] = pan_number.strip()
            #         client_details["Relationship Manager"] = relationship_manager
                
            #     if 'Grand Total' in values[0]:
            #         client_details["Invested Amount"] = values[4]
            #         client_details["Current Value"] = values[6]
            #         client_details["Dividend"] = values[7]
            #         client_details["Gain"] = values[8]
            #         client_details["Holding Days"] = values[9]
            #         client_details["Absolute Return"] = values[10]
            #         client_details["CAGR(%)"] = values[11]

            #     # Detect scheme summaries (rows with 'Total' and not just category totals)
            #     if values[0] and 'Total' in values[0] and 'Fund' in values[0] and current_scheme:
            #         scheme_name = current_scheme
            #         summary_data.append({
            #             "Scheme": scheme_name,
            #             "Folio": scheme_folio,
            #             "Total Investment": values[4],
            #             "Current Value": values[6],
            #             "Gain": values[8],
            #             "Holding Days": values[9],
            #             "Absolute Return (%)": values[10],
            #             "CAGR (%)": values[11],
            #             "Avg. Purchase NAV": avg_nav,
            #             "Current Price": current_price,
            #             "Balance Units": balance_units,
            #         })
            #         continue

            #     # Capture transaction rows
            #     if headers and current_scheme:
            #         if any(values) and not 'Total' in values[0] and not 'Avg. Purchase' in values[0] and values[2]:  # Skip summary rows
            #             transaction_data.append({
            #                 "Scheme": current_scheme,
            #                 "Folio": scheme_folio,
            #                 **dict(zip(headers, values))
            #             })

            #     # Detect scheme names (rows containing the scheme details)
            #     if values[0] and "[" in values[0]:  # Assuming scheme names have this format
            #         current_scheme = values[0].split("[")[0].strip()
            #         scheme_folio = values[0].split("[")[-1].strip().split(" ")[0].strip()
                    

            #     # Capture Avg. Purchase NAV, Current Price, and Balance Units
            #     if values[0] and 'Avg. Purchase NAV' in values[0]:
            #         avg_nav = values[0].split(":")[1].split(",")[0].strip()
            #         current_price = values[0].split("Current Price:")[1].split(",")[0].strip()
            #         balance_units = values[0].split("Balance Units:")[1].strip()

            wb = load_workbook(excel_file)
            
            # Results containers
            all_clients = {}
            
            # Process each sheet (focusing on Mutual Fund sheet first)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Skip processing if not the Mutual Fund sheet for now
                if not sheet_name.endswith("Mutual Fund"):
                    continue
                    
                # Initialize tracking variables
                headers = []
                current_client = None
                current_category = None
                current_scheme = None
                scheme_folio = None
                avg_nav = current_price = balance_units = None
                in_transaction_section = False
                
                # Process each row
                for row_idx, row in enumerate(sheet.iter_rows(), 1):
                    values = [cell.value for cell in row]
                    
                    # Skip empty rows
                    if all(v is None or v == "" for v in values):
                        continue
                        
                    # Detect client rows
                    if values[0] and "Client:" in str(values[0]):
                        client_text = str(values[0]).strip()
                        # Extract client name and PAN
                        if "(" in client_text and ")" in client_text:
                            client_parts = client_text.replace("Client:", "").strip()
                            client_name = client_parts.split("(")[0].strip()
                            pan_number = client_parts.split("(")[1].split(")")[0].strip()
                            
                            # Set as current client
                            current_client = client_name
                            
                            # Initialize client data if not exists
                            if current_client not in all_clients:
                                all_clients[current_client] = {
                                    "client_details": {
                                        "Client Name": client_name,
                                        "Client Pan": pan_number,
                                    },
                                    "summary_data": [],
                                    "transaction_data": []
                                }
                            
                            # Extract relationship manager
                            for cell_val in values:
                                if cell_val and "Relationship Manager:" in str(cell_val):
                                    manager = str(cell_val).split("Relationship Manager:")[1].strip()
                                    all_clients[current_client]["client_details"]["Relationship Manager"] = manager
                                    break
                        
                    # When we see a client name in a row by itself, switch to that client
                    elif values[0] and "(" in str(values[0]) and ")" in str(values[0]) and all(v is None or v == "" for v in values[1:]):
                        client_text = str(values[0]).strip()
                        for client_name in all_clients.keys():
                            if client_name in client_text:
                                current_client = client_name
                                break
                    
                    # Detect headers for transactions
                    if values[0] and "TRANSACTION TYPE" in str(values[0]):
                        headers = values
                        in_transaction_section = True
                        continue
                        
                    # Process Grand Total for current client
                    if values[0] and "Grand Total" in str(values[0]):
                        if current_client:
                            client_data = all_clients[current_client]
                            client_data["client_details"]["Invested Amount"] = values[4]
                            client_data["client_details"]["Current Value"] = values[6]
                            client_data["client_details"]["Dividend"] = values[7]
                            client_data["client_details"]["Gain"] = values[8]
                            client_data["client_details"]["Holding Days"] = values[9]
                            client_data["client_details"]["Absolute Return"] = values[10]
                            client_data["client_details"]["CAGR(%)"] = values[11]
                    
                    # Process individual client totals
                    if current_client and values[0] and str(values[0]).endswith("Total") and current_client in str(values[0]):
                        client_data = all_clients[current_client]
                        if "Total Investment" not in client_data["client_details"]:
                            client_data["client_details"]["Total Investment"] = values[4]
                            client_data["client_details"]["Current Value"] = values[6]
                            client_data["client_details"]["Gain"] = values[8]
                            client_data["client_details"]["Holding Days"] = values[9]
                            client_data["client_details"]["Absolute Return"] = values[10]
                            client_data["client_details"]["CAGR(%)"] = values[11]
                    
                    # Detect category rows (e.g., "Equity: ELSS")
                    if values[0] and ":" in str(values[0]) and not "[" in str(values[0]) and not "NAV:" in str(values[0]) and not "Client:" in str(values[0]):
                        current_category = str(values[0]).strip()
                        continue
                        
                    # Detect scheme rows
                    if values[0] and "[" in str(values[0]):
                        scheme_text = str(values[0]).strip()
                        current_scheme = scheme_text.split("[")[0].strip()
                        
                        # Extract folio number (handle cases with 'Comments:')
                        folio_section = scheme_text.split("[")[1].strip()
                        if "Comments:" in folio_section:
                            scheme_folio = folio_section.split("Comments:")[0].strip()
                            comments = folio_section.split("Comments:")[1].strip()
                        else:
                            scheme_folio = folio_section.rstrip("] ")
                            comments = ""
                            
                        # Ensure we're not capturing "]" at the end
                        scheme_folio = scheme_folio.rstrip("]").strip()
                        continue
                        
                    # Capture Avg. Purchase NAV, Current Price, and Balance Units
                    if values[0] and "Avg. Purchase NAV" in str(values[0]):
                        nav_text = str(values[0]).strip()
                        
                        # Extract the three values using regex
                        avg_nav_match = re.search(r'Avg\. Purchase NAV: ([\d\.]+)', nav_text)
                        current_price_match = re.search(r'Current Price: ([\d\.]+)', nav_text)
                        balance_units_match = re.search(r'Balance Units: ([\d\.]+)', nav_text)
                        
                        if avg_nav_match:
                            avg_nav = avg_nav_match.group(1)
                        if current_price_match:
                            current_price = current_price_match.group(1)
                        if balance_units_match:
                            balance_units = balance_units_match.group(1)
                        continue
                        
                    # Detect scheme totals (e.g., "Axis ELSS Tax Saver Fund (G) Total")
                    if values[0] and "Total" in str(values[0]) and current_scheme and current_scheme in str(values[0]):
                        if current_client and current_scheme:
                            all_clients[current_client]["summary_data"].append({
                                "Category": current_category,
                                "Scheme": current_scheme,
                                "Folio": scheme_folio,
                                "Total Investment": values[4],
                                "Current Value": values[6],
                                "Gain": values[8],
                                "Holding Days": values[9],
                                "Absolute Return (%)": values[10],
                                "CAGR (%)": values[11],
                                "Avg. Purchase NAV": avg_nav,
                                "Current Price": current_price,
                                "Balance Units": balance_units,
                            })
                        continue
                        
                    # Process transaction rows
                    if (headers and current_client and current_scheme and in_transaction_section and 
                        values[0] and any(x in str(values[0]) for x in ["Purchase", "SIP", "Switch In", "Redemption", "Switch Out"])):
                        
                        # Create a dictionary with header names as keys
                        transaction = {}
                        for i, header in enumerate(headers):
                            if i < len(values):
                                transaction[header] = values[i]
                            else:
                                transaction[header] = None
                                
                        # Add scheme info to transaction
                        transaction["Scheme"] = current_scheme
                        transaction["Folio"] = scheme_folio
                        transaction["Category"] = current_category
                        
                        # Add to the client's transactions
                        all_clients[current_client]["transaction_data"].append(transaction)
            
            for client_name, client_data in all_clients.items():
                print(f"Client: {client_name}")
                print(f"Details: {client_data['client_details']}")
                print(f"Number of schemes: {len(client_data['summary_data'])}")
                print(f"Number of transactions: {len(client_data['transaction_data'])}")

            client_details = client_data['client_details']
            summary_data = client_data['summary_data']
            transaction_data = client_data['transaction_data']

            ppt_data = []

            benchmark_index = AMCFundScheme.objects.filter(name="BSE 500 - TRI").first()

            for fund_data in summary_data:
                amcfs = AMCFundScheme.objects.filter(name=fund_data["Scheme"]).first()
                if amcfs:
                    concentration = get_concentration_of_scheme(amcfs)
                    quality, valuation = evaluate_fund_performance(amcfs, benchmark_index)

                    fund_data["Fund Category"] = amcfs.fund_class
                    fund_data["Is Direct Fund"] = amcfs.is_direct_fund
                    fund_data["Concentration"] = concentration
                    fund_data["Price"] = valuation
                    fund_data["Quality"] = quality
                    fund_data["Performance"] = quality

                    fund_data["No of Securities"] = int(amcfs.number_of_underlying_stocks) if amcfs.number_of_underlying_stocks else 0
                    fund_data["Age of the fund"] = calculate_scheme_age(amcfs.launch_date) if amcfs.launch_date else 0
                    fund_data["Equity Allocation"] = round(amcfs.equity_percentage, 1) if amcfs.equity_percentage else 0
                    fund_data["Debt Allocation"] = round(amcfs.debt_percentage, 1) if amcfs.debt_percentage else 0
                    fund_data["Cash Allocation"] = round(amcfs.cash, 1) if amcfs.cash else 0
                    fund_data["Expense ratio"] = round(amcfs.expense_ratio, 1) if amcfs.expense_ratio else 0
                    fund_data["Credit rating AAA"] = round(amcfs.AAA, 1) if amcfs.AAA else 0
                    fund_data["Credit rating AA"] = round(amcfs.AA, 1) if amcfs.AA else 0
                    fund_data["Credit rating A"] = round(amcfs.A, 1) if amcfs.A else 0
                    fund_data["Up Capture"] = round(amcfs.up_capture, 1) if amcfs.up_capture else 0
                    fund_data["Down Capture"] = round(amcfs.down_capture, 1) if amcfs.down_capture else 0

                    fund_data["Mod duration"] = round(amcfs.modified_duration, 1) if amcfs.modified_duration else 0
                    fund_data["YTM"] = round(amcfs.ytm, 1) if amcfs.ytm else 0

                    # benchmark_index = None
                    # if amcfs.scheme_benchmark:
                    #     benchmark_index = AMCFundScheme.objects.filter(name=amcfs.scheme_benchmark).first()
                    # if not benchmark_index:
                    #     benchmark_index = AMCFundScheme.objects.filter(name="BSE 500 TRI").first()
                    ppt_data.append({
                    "Scheme Name": [
                        amcfs.name,
                        benchmark_index.name,
                        amcfs.fund_class
                    ],
                    "Purchase Value": round(fund_data["Total Investment"], 1),
                    "Current Value": round(fund_data["Current Value"], 1),
                    "Gain": round(fund_data["Gain"], 1),
                    "Weight": round((fund_data["Current Value"]/client_details["Current Value"])*100,1),
                    "Stocks": int(amcfs.number_of_underlying_stocks) if amcfs.number_of_underlying_stocks else 0,
                    "SI(%)": [
                        round(amcfs.returns_from_launch, 1) if amcfs.returns_from_launch else 0,
                        round(benchmark_index.returns_from_launch, 1) if benchmark_index.returns_from_launch else 0,
                        round(amcfs.fund_class_avg_returns_from_launch, 1) if amcfs.fund_class_avg_returns_from_launch else 0
                    ],
                    "1YR(%)": [
                        round(amcfs.returns_1_yr, 1) if amcfs.returns_1_yr else 0,
                        round(benchmark_index.returns_1_yr, 1) if benchmark_index.returns_1_yr else 0,
                        round(amcfs.fund_class_avg_1_yr_returns, 1) if amcfs.fund_class_avg_1_yr_returns else 0
                    ],
                    "3YR(%)": [
                        round(amcfs.returns_3_yr, 1) if amcfs.returns_3_yr else 0,
                        round(benchmark_index.returns_3_yr, 1) if benchmark_index.returns_3_yr else 0,
                        round(amcfs.fund_class_avg_3_yr_returns, 1) if amcfs.fund_class_avg_3_yr_returns else 0
                    ],
                    "5YR(%)": [
                        round(amcfs.returns_5_yr, 1) if amcfs.returns_5_yr else 0,
                        round(benchmark_index.returns_5_yr, 1) if benchmark_index.returns_5_yr else 0,
                        round(amcfs.fund_class_avg_5_yr_returns, 1) if amcfs.fund_class_avg_5_yr_returns else 0
                    ],
                    "PE": round(amcfs.pe_ratio, 1) if amcfs.pe_ratio else 0,
                    "Index PE": round(benchmark_index.pe_ratio, 1) if benchmark_index.pe_ratio else 0,
                    "Comments": [
                        "Concentration: " + concentration,
                        "Performance: " + quality,
                        "Valuation: " + valuation
                    ]
                })
                    
            file = create_fund_presentation(ppt_data, client_details, summary_data, transaction_data)
            
            # Create response for file download
            if os.path.exists(file):
                with open(file, 'rb') as f:
                    response = HttpResponse(
                        f.read(),
                        content_type='application/vnd.openxmlformats-officedocument.presentationml.presentation'
                    )
                    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file)}"'
                
                # Clean up the file after sending
                try:
                    os.remove(file)
                except Exception as e:
                    print(f"Error removing file: {e}")
                
                return response
            
            messages.success(request, "PPT File Generated Successfully")
            return render(request, 'gcia_app/portfolio_valuation.html', {'form': form})
        except Exception as e:
            print(traceback.format_exc())
            messages.error(request, f"Error processing file: {e}")
            return render(request, 'gcia_app/portfolio_valuation.html')

    return render(request, 'gcia_app/portfolio_valuation.html', {'form': form})

@login_required
def process_financial_planning(request):
    return render(request, 'gcia_app/financial_planning.html')

@login_required
def fund_analysis_metrics_view(request):
    """
    Display page with dropdown to select mutual funds for analysis
    """
    # Get only active mutual fund schemes that have underlying stock holdings
    funds = AMCFundScheme.objects.filter(
        holdings__isnull=False,  # Must have holdings mapped
        is_active=True
    ).distinct().order_by('name')  # Distinct to avoid duplicates from multiple holdings

    context = {
        'funds': funds,
        'page_title': 'Fund Analysis Metrics'
    }

    return render(request, 'gcia_app/fund_analysis_metrics.html', context)

@login_required
def get_fund_holdings(request, scheme_id):
    """
    AJAX endpoint to get holdings for a fund (for exclusion dropdown)
    Returns JSON with list of holdings and their percentages
    """
    from django.http import JsonResponse

    try:
        scheme = AMCFundScheme.objects.get(amcfundscheme_id=scheme_id, is_active=True)
        holdings = FundHolding.objects.filter(scheme=scheme).select_related('stock').order_by('-holding_percentage')

        holdings_data = [
            {
                'id': holding.fund_holding_id,
                'stock_id': holding.stock.stock_id,
                'company_name': holding.stock.company_name,
                'percentage': f"{holding.holding_percentage:.2f}" if holding.holding_percentage else "0.00"
            }
            for holding in holdings
        ]

        return JsonResponse({
            'success': True,
            'holdings': holdings_data
        })
    except AMCFundScheme.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Fund not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

def get_filtered_holdings(scheme, exclude_marketcap_year, exclude_holdings_ids):
    """
    Filter holdings based on exclusion criteria

    Args:
        scheme: AMCFundScheme instance
        exclude_marketcap_year: Year string (e.g., "2019") or None
        exclude_holdings_ids: List of holding IDs to exclude

    Returns:
        QuerySet of filtered FundHolding objects
    """
    from datetime import date
    from django.db.models import Exists, OuterRef
    import logging

    logger = logging.getLogger(__name__)

    # Start with all holdings
    holdings = FundHolding.objects.filter(scheme=scheme).select_related('stock').prefetch_related(
        'stock__market_cap_data', 'stock__ttm_data', 'stock__quarterly_data',
        'stock__annual_ratios', 'stock__price_data'
    ).order_by('-holding_percentage')

    # EXCLUSION 1: MarketCap year filter
    if exclude_marketcap_year:
        try:
            cutoff_year = int(exclude_marketcap_year)
            cutoff_date = date(cutoff_year, 1, 1)

            # Only include stocks with MarketCap data FROM cutoff year onwards
            from gcia_app.models import StockMarketCap

            marketcap_exists = StockMarketCap.objects.filter(
                stock=OuterRef('stock'),
                date__gte=cutoff_date  # >= (from year onwards)
            )

            holdings = holdings.annotate(
                has_recent_marketcap=Exists(marketcap_exists)
            ).filter(has_recent_marketcap=True)

            logger.info(f"MarketCap filter: >{cutoff_year} (stocks with data from {cutoff_year} onwards)")

        except (ValueError, TypeError) as e:
            logger.error(f"Invalid marketcap year: {exclude_marketcap_year} - {e}")

    # EXCLUSION 2: Specific holdings filter
    if exclude_holdings_ids:
        holdings = holdings.exclude(fund_holding_id__in=exclude_holdings_ids)
        logger.info(f"Excluded {len(exclude_holdings_ids)} specific holdings")

    return holdings

@login_required
def download_fund_metrics(request, scheme_id):
    """
    Download Excel file with fund holdings integrated into stock structure (Portfolio Analysis format)
    Enhanced with calculated portfolio analysis metrics and dynamic column structure
    Includes all 22 calculated metrics for each period and portfolio-level summary
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.utils import timezone
    from io import BytesIO
    from .metrics_calculator import DynamicHeaderGenerator
    from .models import FundMetricsLog
    from datetime import datetime
    import logging

    logger = logging.getLogger(__name__)

    try:
        # Get the selected fund scheme
        scheme = AMCFundScheme.objects.get(amcfundscheme_id=scheme_id, is_active=True)

        # Parse exclusion parameters from GET request
        exclude_marketcap_year = request.GET.get('exclude_marketcap', None)
        exclude_holdings_param = request.GET.get('exclude_holdings', None)

        # Parse comma-separated holding IDs
        exclude_holdings_ids = []
        if exclude_holdings_param:
            try:
                exclude_holdings_ids = [int(h_id.strip()) for h_id in exclude_holdings_param.split(',') if h_id.strip()]
            except ValueError:
                logger.warning(f"Invalid exclude_holdings parameter: {exclude_holdings_param}")

        # Check if any exclusions are present
        has_exclusions = bool(exclude_marketcap_year) or bool(exclude_holdings_ids)

        logger.info(f"Generating Portfolio Analysis Excel for {scheme.name}")
        logger.info(f"Exclusion parameters - MarketCap: {exclude_marketcap_year}, Holdings: {exclude_holdings_ids}")

        # Use enhanced Excel export functionality
        from .enhanced_excel_export import (
            generate_enhanced_portfolio_analysis_excel,
            generate_recalculated_analysis_excel,
            extract_summary_data_from_worksheet,
            create_summary_sheet
        )
        from openpyxl import load_workbook

        if not has_exclusions:
            # NO EXCLUSIONS: Return single-sheet Excel with Summary
            logger.info("No exclusions - generating single-sheet Excel with Summary")
            excel_content = generate_enhanced_portfolio_analysis_excel(scheme)

            # Load the workbook to add Summary sheet
            wb = load_workbook(excel_content)
            analysis_ws = wb.active

            # Extract summary data from the analysis sheet
            logger.info("Extracting summary data for Summary sheet...")
            summary_data = extract_summary_data_from_worksheet(analysis_ws, scheme)

            # Create Summary sheet (will be inserted as first sheet)
            logger.info("Creating Summary sheet...")
            create_summary_sheet(wb, scheme, summary_data)

            # Save the updated workbook
            from io import BytesIO
            final_excel = BytesIO()
            wb.save(final_excel)
            final_excel.seek(0)

            response = HttpResponse(
                final_excel.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"{scheme.name}_Portfolio_Analysis_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            logger.info(f"Single-sheet Excel file with Summary generated: {filename}")
            return response

        else:
            # EXCLUSIONS PRESENT: Generate dual-sheet Excel
            logger.info("Exclusions detected - generating dual-sheet Excel")

            # STEP 1: Generate Default Analysis sheet (from database)
            logger.info("Generating Default Analysis sheet...")
            default_excel = generate_enhanced_portfolio_analysis_excel(scheme)

            # STEP 2: Get filtered holdings
            logger.info("Filtering holdings based on exclusion criteria...")
            filtered_holdings = get_filtered_holdings(scheme, exclude_marketcap_year, exclude_holdings_ids)

            if not filtered_holdings.exists():
                messages.error(request, "All holdings were excluded by the filters. Cannot generate Excel.")
                return redirect('fund_analysis_metrics')

            logger.info(f"Filtered holdings: {filtered_holdings.count()} stocks (out of {FundHolding.objects.filter(scheme=scheme).count()})")

            # STEP 3: Generate Recalculated Analysis sheet (from filtered holdings)
            logger.info("Generating Recalculated Analysis sheet...")
            recalculated_excel = generate_recalculated_analysis_excel(scheme, filtered_holdings)

            # STEP 4: Combine both sheets into one workbook
            logger.info("Combining sheets into dual-sheet workbook...")

            # Load default workbook
            default_wb = load_workbook(default_excel)
            default_ws = default_wb.active
            default_ws.title = "Default Analysis"

            # Load recalculated workbook
            recalculated_wb = load_workbook(recalculated_excel)
            recalculated_ws = recalculated_wb.active

            # Copy recalculated sheet to default workbook
            new_ws = default_wb.create_sheet(title="Recalculated Analysis")

            # Copy all cells from recalculated sheet to new sheet
            for row in recalculated_ws.iter_rows():
                for cell in row:
                    new_cell = new_ws[cell.coordinate]
                    new_cell.value = cell.value

                    # Copy formatting
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy()
                        new_cell.alignment = cell.alignment.copy()

            # Copy column dimensions
            for col_letter, col_dim in recalculated_ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = col_dim.width

            # Copy row dimensions
            for row_num, row_dim in recalculated_ws.row_dimensions.items():
                new_ws.row_dimensions[row_num].height = row_dim.height

            logger.info("Dual-sheet workbook created successfully")

            # STEP 5: Add Summary sheet based on Recalculated Analysis
            logger.info("Creating Summary sheet from Recalculated Analysis...")
            summary_data = extract_summary_data_from_worksheet(new_ws, scheme)
            create_summary_sheet(default_wb, scheme, summary_data)
            logger.info("Summary sheet added successfully")

            # STEP 6: Save and return dual-sheet workbook with Summary
            dual_excel_content = BytesIO()
            default_wb.save(dual_excel_content)
            dual_excel_content.seek(0)

            response = HttpResponse(
                dual_excel_content.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"{scheme.name}_Portfolio_Analysis_Dual_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            logger.info(f"Dual-sheet Excel file with Summary generated: {filename}")
            return response

    except AMCFundScheme.DoesNotExist:
        messages.error(request, "Fund not found or inactive")
        return redirect('fund_analysis_metrics')
    except Exception as e:
        logger.error(f"Error in download_fund_metrics for scheme {scheme_id}: {e}", exc_info=True)
        traceback.print_exc()
        messages.error(request, f"Error generating Excel file: {str(e)}")
        return redirect('fund_analysis_metrics')
# Portfolio Metrics Calculation Views
import uuid
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .metrics_calculator import PortfolioMetricsCalculator
from .models import MetricsCalculationSession, Customer

@login_required
@require_http_methods(["POST"])
def trigger_metrics_calculation(request):
    """
    Trigger real-time metrics calculation with progress tracking
    """

    try:
        # Get parameters from request
        limit_periods = request.POST.get('limit_periods')

        # Parse limit_periods if provided
        limit_params = None
        if limit_periods:
            try:
                limit_params = json.loads(limit_periods)
            except (json.JSONDecodeError, ValueError):
                # If parsing fails, ignore limits
                limit_params = None

        # Generate session ID for tracking
        session_id = str(uuid.uuid4())

        # Create progress session immediately for tracking
        from gcia_app.models import AMCFundScheme, MetricsCalculationSession

        # Count funds first
        funds_with_holdings = AMCFundScheme.objects.filter(
            holdings__isnull=False, is_active=True
        ).distinct()
        total_funds = funds_with_holdings.count()

        # Create session for immediate progress tracking
        progress_session = MetricsCalculationSession.objects.create(
            session_id=session_id,
            user=request.user,
            total_funds=total_funds,
            status='started'
        )

        # Start calculation in background using threading
        import threading
        def run_calculation():
            try:
                calculator = PortfolioMetricsCalculator(
                    session_id=session_id,
                    user=request.user
                )
                # Set the pre-created session
                calculator.progress_session = progress_session
                calculator.calculate_metrics_for_all_funds(limit_params)
            except Exception as e:
                # Update session with error
                progress_session.status = 'failed'
                progress_session.error_message = str(e)
                progress_session.save()

        # Start calculation thread
        calc_thread = threading.Thread(target=run_calculation)
        calc_thread.daemon = True
        calc_thread.start()

        return JsonResponse({
            'success': True,
            'session_id': session_id,
            'message': 'Metrics calculation started in background',
            'total_funds': total_funds
        })

    except Exception as e:
        error_message = str(e)
        full_traceback = traceback.format_exc()

        print(f"Error in trigger_metrics_calculation: {error_message}")
        print(f"Full traceback: {full_traceback}")

        # Special handling for period_date field errors
        if "Cannot resolve keyword 'period_date'" in error_message:
            print("*** PERIOD_DATE ERROR DETECTED ***")
            print(f"Error details: {error_message}")
            print("This error occurs when Django ORM tries to use 'period_date' field on StockTTMData or StockQuarterlyData models")
            print("These models only have 'period' field, not 'period_date'")

        return JsonResponse({
            'success': False,
            'session_id': session_id if 'session_id' in locals() else None,
            'error': error_message,
            'traceback': full_traceback if 'DEBUG' in globals() else None
        })

@login_required
def get_calculation_progress(request, session_id):
    """
    Get real-time progress status for metrics calculation
    """

    try:
        progress = MetricsCalculationSession.objects.get(
            session_id=session_id,
            user=request.user
        )

        return JsonResponse({
            'total_funds': progress.total_funds,
            'processed_funds': progress.processed_funds,
            'current_fund_name': progress.current_fund_name or '',
            'current_stock_name': progress.current_stock_name or '',
            'status': progress.status,
            'progress_percentage': progress.progress_percentage,
            'error_message': progress.error_message or ''
        })

    except MetricsCalculationSession.DoesNotExist:
        return JsonResponse({
            'error': 'Calculation session not found'
        }, status=404)
    except Exception as e:
        print(f"Error in get_calculation_progress: {e}")
        return JsonResponse({
            'error': f'Error retrieving progress: {str(e)}'
        }, status=500)

@login_required
def get_upload_progress(request, session_id):
    """
    API endpoint to get real-time upload progress for stock data processing
    """
    from .models import UploadProgressSession

    try:
        progress = UploadProgressSession.objects.get(session_id=session_id)

        return JsonResponse({
            'session_id': progress.session_id,
            'filename': progress.filename,
            'total_rows': progress.total_rows,
            'processed_rows': progress.processed_rows,
            'current_stock_name': progress.current_stock_name or '',
            'status': progress.status,
            'progress_percentage': progress.progress_percentage,
            'stocks_created': progress.stocks_created,
            'stocks_updated': progress.stocks_updated,
            'records_created': progress.records_created,
            'error_message': progress.error_message or ''
        })

    except UploadProgressSession.DoesNotExist:
        return JsonResponse({
            'error': 'Upload session not found'
        }, status=404)
    except Exception as e:
        print(f"Error in get_upload_progress: {e}")
        return JsonResponse({
            'error': f'Error retrieving upload progress: {str(e)}'
        }, status=500)

@login_required
@require_http_methods(["POST"])
def cancel_metrics_calculation(request, session_id):
    """
    Cancel a running metrics calculation
    """

    try:
        progress = MetricsCalculationSession.objects.get(
            session_id=session_id,
            user=request.user
        )

        if progress.status in ['started', 'processing']:
            progress.status = 'cancelled'
            progress.error_message = 'Calculation cancelled by user'
            progress.completed_at = timezone.now()
            progress.save()

            return JsonResponse({
                'success': True,
                'message': 'Calculation cancelled successfully'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': f'Cannot cancel calculation with status: {progress.status}'
            })

    except MetricsCalculationSession.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Calculation session not found'
        }, status=404)
    except Exception as e:
        print(f"Error in cancel_metrics_calculation: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Error cancelling calculation: {str(e)}'
        }, status=500)


