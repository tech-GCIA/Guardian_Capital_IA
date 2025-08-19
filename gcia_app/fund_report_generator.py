import os
import tempfile
from datetime import datetime, date
from decimal import Decimal
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from django.conf import settings
from gcia_app.models import Stock, StockQuarterlyData
import logging
import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)

class FundReportGenerator:
    """
    Generate Excel report by building from scratch based on sample structure
    """
    
    def __init__(self):
        self.workbook = None
        self.sheets = {}
        
    def generate_report(self, fund_name, selected_stocks):
        """
        Generate the complete Excel report
        
        Args:
            fund_name (str): Name of the fund
            selected_stocks (list): List of stock data with weightage and shares
            
        Returns:
            str: Path to the generated Excel file
        """
        try:
            # Create a new workbook from scratch
            self.workbook = Workbook()
            
            # Remove default sheet
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])
            
            # Process selected stocks and get database data
            stock_data = self._process_selected_stocks(selected_stocks)
            
            # Create sheets in order (matching sample file)
            self._create_summary_sheet(fund_name, stock_data)
            self._create_stock_weights_sheet(stock_data)
            self._create_portfolio_analysis_sheet(stock_data)
            
            # Save the file
            output_path = self._save_report(fund_name)
            
            logger.info(f"Report generated successfully for {fund_name} with {len(stock_data)} stocks")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            raise
    
    def _process_selected_stocks(self, selected_stocks):
        """
        Process selected stocks and get all required data from database
        """
        enhanced_data = []
        
        logger.info(f"Processing {len(selected_stocks)} selected stocks")
        
        for stock_selection in selected_stocks:
            stock_id = stock_selection['id']
            weightage = float(stock_selection['weightage'])
            shares = int(stock_selection['shares'])
            
            try:
                # Get stock from database
                stock = Stock.objects.get(stock_id=stock_id)
                
                # Get quarterly data for this stock
                quarterly_data = StockQuarterlyData.objects.filter(
                    stock=stock
                ).order_by('-quarter_year', '-quarter_number')[:24]  # Get 6 years of data
                
                logger.info(f"Found {len(quarterly_data)} quarters for {stock.name}")
                
                # Calculate all required metrics
                metrics = self._calculate_comprehensive_metrics(stock, quarterly_data)
                
                # Extract accord code from symbol (STOCK_XXX -> XXX)
                accord_code = stock.symbol.split('_')[-1] if '_' in stock.symbol else stock.symbol
                
                enhanced_data.append({
                    'stock': stock,
                    'accord_code': accord_code,
                    'weightage': weightage,
                    'shares': shares,
                    'quarterly_data': list(quarterly_data),
                    'metrics': metrics
                })
                
            except Stock.DoesNotExist:
                logger.warning(f"Stock with ID {stock_id} not found in database")
                continue
        
        return enhanced_data
    
    def _calculate_comprehensive_metrics(self, stock, quarterly_data):
        """
        Calculate all metrics required for the report
        """
        metrics = {}
        quarters = list(quarterly_data)
        
        if not quarters:
            return self._get_empty_metrics()
        
        # Get latest quarter
        current_q = quarters[0] if quarters else None
        
        # Basic metrics
        metrics['mcap'] = float(current_q.mcap or 0) if current_q else 0
        metrics['free_float'] = 75.0  # Default value, update if available in your model
        
        # Quarterly market caps for all quarters
        metrics['quarterly_mcaps'] = {}
        for q in quarters:
            quarter_key = f"{q.quarter_year}Q{q.quarter_number}"
            metrics['quarterly_mcaps'][quarter_key] = float(q.mcap or 0)
        
        # TTM metrics
        metrics['ttm_revenue'] = float(current_q.revenue or 0) if current_q else 0
        metrics['ttm_pat'] = float(current_q.pat or 0) if current_q else 0
        
        # Growth calculations
        # QoQ Growth
        if len(quarters) >= 2:
            curr_pat = float(quarters[0].pat or 0)
            prev_pat = float(quarters[1].pat or 0)
            metrics['qoq_growth'] = ((curr_pat - prev_pat) / prev_pat * 100) if prev_pat > 0 else 0
        else:
            metrics['qoq_growth'] = 0
        
        # YoY Growth
        if len(quarters) >= 5:
            curr_pat = float(quarters[0].pat or 0)
            year_ago_pat = float(quarters[4].pat or 0)
            metrics['yoy_growth'] = ((curr_pat - year_ago_pat) / year_ago_pat * 100) if year_ago_pat > 0 else 0
        else:
            metrics['yoy_growth'] = 0
        
        # 6-Year CAGR
        if len(quarters) >= 24:
            current_revenue = float(quarters[0].revenue or 0)
            six_year_ago_revenue = float(quarters[23].revenue or 0)
            if six_year_ago_revenue > 0:
                metrics['cagr_6yr'] = (((current_revenue / six_year_ago_revenue) ** (1/6)) - 1) * 100
            else:
                metrics['cagr_6yr'] = 0
        else:
            metrics['cagr_6yr'] = 0
        
        # PE metrics
        pe_values = [float(q.pe_ratio or 0) for q in quarters if q.pe_ratio and float(q.pe_ratio) > 0]
        
        metrics['current_pe'] = float(current_q.pe_ratio or 0) if current_q else 0
        
        # 2-Year and 5-Year Average PE
        pe_2yr = pe_values[:8] if len(pe_values) >= 8 else pe_values
        metrics['pe_2yr_avg'] = np.mean(pe_2yr) if pe_2yr else 0
        
        pe_5yr = pe_values[:20] if len(pe_values) >= 20 else pe_values
        metrics['pe_5yr_avg'] = np.mean(pe_5yr) if pe_5yr else 0
        
        # PE Revaluation
        if metrics['pe_2yr_avg'] > 0:
            metrics['pe_reval_2yr'] = ((metrics['current_pe'] - metrics['pe_2yr_avg']) / metrics['pe_2yr_avg']) * 100
        else:
            metrics['pe_reval_2yr'] = 0
            
        if metrics['pe_5yr_avg'] > 0:
            metrics['pe_reval_5yr'] = ((metrics['current_pe'] - metrics['pe_5yr_avg']) / metrics['pe_5yr_avg']) * 100
        else:
            metrics['pe_reval_5yr'] = 0
        
        # PB metrics
        pb_values = [float(q.pb_ratio or 0) for q in quarters if q.pb_ratio and float(q.pb_ratio) > 0]
        
        metrics['current_pb'] = float(current_q.pb_ratio or 0) if current_q else 0
        
        pb_2yr = pb_values[:8] if len(pb_values) >= 8 else pb_values
        metrics['pb_2yr_avg'] = np.mean(pb_2yr) if pb_2yr else 0
        
        pb_5yr = pb_values[:20] if len(pb_values) >= 20 else pb_values
        metrics['pb_5yr_avg'] = np.mean(pb_5yr) if pb_5yr else 0
        
        # PB Revaluation
        if metrics['pb_2yr_avg'] > 0:
            metrics['pb_reval_2yr'] = ((metrics['current_pb'] - metrics['pb_2yr_avg']) / metrics['pb_2yr_avg']) * 100
        else:
            metrics['pb_reval_2yr'] = 0
            
        if metrics['pb_5yr_avg'] > 0:
            metrics['pb_reval_5yr'] = ((metrics['current_pb'] - metrics['pb_5yr_avg']) / metrics['pb_5yr_avg']) * 100
        else:
            metrics['pb_reval_5yr'] = 0
        
        # 10 Quarter PB Low/High
        pb_10q = pb_values[:10] if len(pb_values) >= 10 else pb_values
        if pb_10q:
            metrics['pb_10q_low'] = min(pb_10q)
            metrics['pb_10q_high'] = max(pb_10q)
        else:
            metrics['pb_10q_low'] = 0
            metrics['pb_10q_high'] = 0
        
        # Alpha calculations
        bond_rate = 7.0
        
        if len(quarters) >= 8:
            recent_pats = [float(q.pat or 0) for q in quarters[:4] if q.pat]
            older_pats = [float(q.pat or 0) for q in quarters[4:8] if q.pat]
            
            if recent_pats and older_pats:
                avg_recent = np.mean(recent_pats)
                avg_older = np.mean(older_pats)
                if avg_older > 0:
                    annual_growth = ((avg_recent / avg_older) ** 0.25 - 1) * 100
                    metrics['alpha_over_bond'] = annual_growth - bond_rate
                    metrics['absolute_alpha'] = annual_growth
                else:
                    metrics['alpha_over_bond'] = 0
                    metrics['absolute_alpha'] = 0
            else:
                metrics['alpha_over_bond'] = 0
                metrics['absolute_alpha'] = 0
        else:
            metrics['alpha_over_bond'] = 0
            metrics['absolute_alpha'] = 0
        
        # PE Yield
        if metrics['current_pe'] > 0:
            metrics['pe_yield'] = (1 / metrics['current_pe']) * 100
        else:
            metrics['pe_yield'] = 0
        
        metrics['growth_rate'] = metrics.get('yoy_growth', 0)
        metrics['bond_rate'] = bond_rate
        
        return metrics
    
    def _get_empty_metrics(self):
        """Return empty metrics structure"""
        return {
            'mcap': 0, 'free_float': 0, 'quarterly_mcaps': {},
            'ttm_revenue': 0, 'ttm_pat': 0,
            'qoq_growth': 0, 'yoy_growth': 0, 'cagr_6yr': 0,
            'current_pe': 0, 'pe_2yr_avg': 0, 'pe_5yr_avg': 0,
            'pe_reval_2yr': 0, 'pe_reval_5yr': 0,
            'current_pb': 0, 'pb_2yr_avg': 0, 'pb_5yr_avg': 0,
            'pb_reval_2yr': 0, 'pb_reval_5yr': 0,
            'pb_10q_low': 0, 'pb_10q_high': 0,
            'alpha_over_bond': 0, 'absolute_alpha': 0,
            'pe_yield': 0, 'growth_rate': 0, 'bond_rate': 7.0
        }
    
    def _create_portfolio_analysis_sheet(self, stock_data):
        """
        Create Portfolio Analysis sheet from scratch with proper structure
        """
        ws = self.workbook.create_sheet(title="3. Portfolio Analysis")
        
        # Define styles
        header_style = self._get_header_style()
        data_style = self._get_data_style()
        total_style = self._get_total_style()
        
        # Headers (based on Old Bridge sample structure)
        # Row 8: Category headers
        category_headers = [
            ('A8', 'Stock Details'),
            ('F8', 'Market Data'),
            ('K8', 'Growth Metrics'),
            ('O8', 'PE Analysis'),
            ('U8', 'PB Analysis'),
            ('AB8', 'Alpha Metrics'),
        ]
        
        for cell_ref, header in category_headers:
            ws[cell_ref] = header
            ws[cell_ref].font = Font(bold=True, size=11)
            ws.merge_cells(f'{cell_ref[0]}8:{chr(ord(cell_ref[0])+3)}8')
        
        # Row 9: Column headers
        headers = [
            'S.No', 'Company Name', 'Accord Code', 'Sector', 'Cap',  # A-E
            'Free Float', 'Weightage (%)', 'No. of Shares', 'Market Cap (Cr)', 'TTM PAT',  # F-J
            'QoQ (%)', 'YoY (%)', '6Y CAGR (%)', '',  # K-N
            'Current PE', '2Y Avg PE', '5Y Avg PE', 'PE Reval 2Y (%)', 'PE Reval 5Y (%)', '',  # O-T
            'Current PB', '2Y Avg PB', '5Y Avg PB', 'PB Reval 2Y (%)', 'PB Reval 5Y (%)',  # U-Y
            '10Q PB Low', '10Q PB High', '',  # Z-AB
            'Alpha/Bond', 'Abs Alpha', 'PE Yield', 'Growth', 'Bond Rate'  # AC-AG
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col_idx, value=header)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Fill stock data starting from row 10
        start_row = 10
        total_mcap = 0
        total_weightage = 0
        
        for idx, stock_info in enumerate(stock_data):
            row = start_row + idx
            stock = stock_info['stock']
            metrics = stock_info['metrics']
            
            # Fill row data
            row_data = [
                idx + 1,  # S.No
                stock.name,  # Company Name
                stock_info['accord_code'],  # Accord Code
                stock.sector or '-',  # Sector
                stock.market_cap_category or '-',  # Cap
                round(metrics['free_float'], 1),  # Free Float
                round(stock_info['weightage'], 2),  # Weightage
                stock_info['shares'],  # No. of Shares
                round(metrics['mcap'], 2),  # Market Cap
                round(metrics['ttm_pat'], 2),  # TTM PAT
                round(metrics['qoq_growth'], 2),  # QoQ
                round(metrics['yoy_growth'], 2),  # YoY
                round(metrics['cagr_6yr'], 2),  # 6Y CAGR
                '',  # Empty column
                round(metrics['current_pe'], 2),  # Current PE
                round(metrics['pe_2yr_avg'], 2),  # 2Y Avg PE
                round(metrics['pe_5yr_avg'], 2),  # 5Y Avg PE
                round(metrics['pe_reval_2yr'], 2),  # PE Reval 2Y
                round(metrics['pe_reval_5yr'], 2),  # PE Reval 5Y
                '',  # Empty column
                round(metrics['current_pb'], 2),  # Current PB
                round(metrics['pb_2yr_avg'], 2),  # 2Y Avg PB
                round(metrics['pb_5yr_avg'], 2),  # 5Y Avg PB
                round(metrics['pb_reval_2yr'], 2),  # PB Reval 2Y
                round(metrics['pb_reval_5yr'], 2),  # PB Reval 5Y
                round(metrics['pb_10q_low'], 2),  # 10Q PB Low
                round(metrics['pb_10q_high'], 2),  # 10Q PB High
                '',  # Empty column
                round(metrics['alpha_over_bond'], 2),  # Alpha/Bond
                round(metrics['absolute_alpha'], 2),  # Abs Alpha
                round(metrics['pe_yield'], 2),  # PE Yield
                round(metrics['growth_rate'], 2),  # Growth
                round(metrics['bond_rate'], 2),  # Bond Rate
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                if col_idx in [7, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33]:
                    # Numeric columns - align right
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
            
            # Accumulate totals
            total_mcap += metrics['mcap'] * stock_info['weightage'] / 100
            total_weightage += stock_info['weightage']
        
        # Add TOTAL row
        total_row = start_row + len(stock_data)
        ws.cell(row=total_row, column=2, value="TOTAL")
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        
        # Total weightage
        ws.cell(row=total_row, column=7, value=round(total_weightage, 2))
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        
        # Weighted average market cap
        ws.cell(row=total_row, column=9, value=round(total_mcap, 2))
        ws.cell(row=total_row, column=9).font = Font(bold=True)
        
        # Calculate weighted averages for other metrics
        if stock_data and total_weightage > 0:
            weighted_pe = sum(s['metrics']['current_pe'] * s['weightage'] / total_weightage for s in stock_data)
            weighted_pb = sum(s['metrics']['current_pb'] * s['weightage'] / total_weightage for s in stock_data)
            
            ws.cell(row=total_row, column=15, value=round(weighted_pe, 2))
            ws.cell(row=total_row, column=15).font = Font(bold=True)
            
            ws.cell(row=total_row, column=21, value=round(weighted_pb, 2))
            ws.cell(row=total_row, column=21).font = Font(bold=True)
        
        # Add borders to the data range
        self._add_borders(ws, 9, total_row, 1, 33)
        
        # Adjust column widths
        column_widths = {
            'A': 8, 'B': 30, 'C': 12, 'D': 15, 'E': 10,
            'F': 10, 'G': 12, 'H': 12, 'I': 15, 'J': 12,
            'K': 10, 'L': 10, 'M': 10, 'N': 5,
            'O': 12, 'P': 12, 'Q': 12, 'R': 15, 'S': 15, 'T': 5,
            'U': 12, 'V': 12, 'W': 12, 'X': 15, 'Y': 15,
            'Z': 12, 'AA': 12, 'AB': 5,
            'AC': 12, 'AD': 12, 'AE': 10, 'AF': 10, 'AG': 10
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        logger.info(f"Created Portfolio Analysis sheet with {len(stock_data)} stocks")
    
    def _create_stock_weights_sheet(self, stock_data):
        """
        Create Stock Weights sheet from scratch
        """
        ws = self.workbook.create_sheet(title="2. Stock Weights", index=0)
        
        # Title
        ws['A1'] = "Stock Weights Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:K1')
        
        # Get all unique quarters from stock data
        all_quarters = set()
        for stock_info in stock_data:
            for q in stock_info['quarterly_data']:
                q_key = f"Q{q.quarter_number} {q.quarter_year}"
                all_quarters.add((q.quarter_year, q.quarter_number, q_key))
        
        # Sort quarters (most recent first)
        sorted_quarters = sorted(all_quarters, key=lambda x: (x[0], x[1]), reverse=True)[:12]  # Last 12 quarters
        
        # Headers
        ws['A4'] = "S.No"
        ws['B4'] = "Stock Name"
        ws['C4'] = "Current Weight (%)"
        
        # Quarter headers
        for idx, (year, qtr, label) in enumerate(sorted_quarters):
            col = 4 + idx  # Start from column D
            ws.cell(row=4, column=col, value=label)
        
        # Apply header formatting
        for col in range(1, 4 + len(sorted_quarters)):
            cell = ws.cell(row=4, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Fill stock data
        start_row = 5
        for idx, stock_info in enumerate(stock_data):
            row = start_row + idx
            
            ws.cell(row=row, column=1, value=idx + 1)  # S.No
            ws.cell(row=row, column=2, value=stock_info['stock'].name)  # Stock Name
            ws.cell(row=row, column=3, value=round(stock_info['weightage'], 2))  # Current Weight
            
            # Fill quarter-wise weights based on market cap proportions
            for q_idx, (year, qtr, label) in enumerate(sorted_quarters):
                col = 4 + q_idx
                
                # Find the quarter data for this stock
                quarter_key = f"{year}Q{qtr}"
                if quarter_key in stock_info['metrics']['quarterly_mcaps']:
                    # Use the same weightage for historical quarters (simplified)
                    # In reality, you might calculate based on historical portfolio composition
                    ws.cell(row=row, column=col, value=round(stock_info['weightage'], 2))
                else:
                    ws.cell(row=row, column=col, value=0)
        
        # Add TOTAL row
        total_row = start_row + len(stock_data)
        ws.cell(row=total_row, column=2, value="TOTAL")
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        
        # Calculate totals
        total_weight = sum(s['weightage'] for s in stock_data)
        ws.cell(row=total_row, column=3, value=round(total_weight, 2))
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        
        # Calculate quarter totals
        for q_idx in range(len(sorted_quarters)):
            col = 4 + q_idx
            col_total = 0
            for i in range(len(stock_data)):
                cell_value = ws.cell(row=start_row + i, column=col).value
                if cell_value:
                    col_total += cell_value
            ws.cell(row=total_row, column=col, value=round(col_total, 2))
            ws.cell(row=total_row, column=col).font = Font(bold=True)
        
        # Add borders
        self._add_borders(ws, 4, total_row, 1, 3 + len(sorted_quarters))
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        for idx in range(len(sorted_quarters)):
            col_letter = get_column_letter(4 + idx)
            ws.column_dimensions[col_letter].width = 12
        
        logger.info(f"Created Stock Weights sheet with {len(stock_data)} stocks and {len(sorted_quarters)} quarters")
    
    def _create_summary_sheet(self, fund_name, stock_data):
        """
        Create Summary sheet from scratch
        """
        ws = self.workbook.create_sheet(title="1. Summary", index=0)
        
        # Title
        ws['A1'] = f"Portfolio Summary - {fund_name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Report Date: {datetime.now().strftime('%Y-%m-%d')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        # Calculate summary metrics
        total_stocks = len(stock_data)
        total_weightage = sum(s['weightage'] for s in stock_data)
        
        if stock_data:
            # Weighted metrics
            weighted_pe = sum(s['metrics']['current_pe'] * s['weightage'] / 100 for s in stock_data)
            weighted_pb = sum(s['metrics']['current_pb'] * s['weightage'] / 100 for s in stock_data)
            weighted_mcap = sum(s['metrics']['mcap'] * s['weightage'] / 100 for s in stock_data)
            
            # Simple averages
            avg_cagr = np.mean([s['metrics']['cagr_6yr'] for s in stock_data])
            avg_yoy = np.mean([s['metrics']['yoy_growth'] for s in stock_data])
            avg_qoq = np.mean([s['metrics']['qoq_growth'] for s in stock_data])
            avg_alpha = np.mean([s['metrics']['alpha_over_bond'] for s in stock_data])
            avg_pe_yield = np.mean([s['metrics']['pe_yield'] for s in stock_data])
            
            # Risk metrics
            pe_values = [s['metrics']['current_pe'] for s in stock_data if s['metrics']['current_pe'] > 0]
            pe_std = np.std(pe_values) if pe_values else 0
            
            # Sector allocation
            sector_allocation = {}
            for stock_info in stock_data:
                sector = stock_info['stock'].sector or 'Others'
                if sector not in sector_allocation:
                    sector_allocation[sector] = 0
                sector_allocation[sector] += stock_info['weightage']
        else:
            weighted_pe = weighted_pb = weighted_mcap = 0
            avg_cagr = avg_yoy = avg_qoq = avg_alpha = avg_pe_yield = 0
            pe_std = 0
            sector_allocation = {}
        
        # Portfolio Overview Section
        ws['A4'] = "PORTFOLIO OVERVIEW"
        ws['A4'].font = Font(bold=True, size=12)
        ws.merge_cells('A4:C4')
        
        overview_data = [
            ("Total Holdings", total_stocks, ""),
            ("Total Weightage (%)", round(total_weightage, 2), "%"),
            ("Weighted Market Cap (Cr)", round(weighted_mcap, 2), "Cr"),
        ]
        
        for idx, (label, value, suffix) in enumerate(overview_data):
            row = 5 + idx
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'C{row}'] = suffix
        
        # Valuation Metrics Section
        ws['A9'] = "VALUATION METRICS"
        ws['A9'].font = Font(bold=True, size=12)
        ws.merge_cells('A9:C9')
        
        valuation_data = [
            ("Portfolio PE", round(weighted_pe, 2), ""),
            ("Portfolio PB", round(weighted_pb, 2), ""),
            ("PE Yield (%)", round(avg_pe_yield, 2), "%"),
            ("PE Std Deviation", round(pe_std, 2), ""),
        ]
        
        for idx, (label, value, suffix) in enumerate(valuation_data):
            row = 10 + idx
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'C{row}'] = suffix
        
        # Growth Metrics Section
        ws['A15'] = "GROWTH METRICS"
        ws['A15'].font = Font(bold=True, size=12)
        ws.merge_cells('A15:C15')
        
        growth_data = [
            ("Average 6Y CAGR (%)", round(avg_cagr, 2), "%"),
            ("Average YoY Growth (%)", round(avg_yoy, 2), "%"),
            ("Average QoQ Growth (%)", round(avg_qoq, 2), "%"),
            ("Average Alpha over Bond", round(avg_alpha, 2), "%"),
        ]
        
        for idx, (label, value, suffix) in enumerate(growth_data):
            row = 16 + idx
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'C{row}'] = suffix
        
        # Top Holdings Section (right side)
        ws['E4'] = "TOP HOLDINGS"
        ws['E4'].font = Font(bold=True, size=12)
        ws.merge_cells('E4:H4')
        
        ws['E5'] = "Stock Name"
        ws['F5'] = "Weight (%)"
        ws['G5'] = "PE"
        ws['H5'] = "YoY (%)"
        
        for col in ['E5', 'F5', 'G5', 'H5']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Sort stocks by weightage
        sorted_stocks = sorted(stock_data, key=lambda x: x['weightage'], reverse=True)[:10]
        
        for idx, stock_info in enumerate(sorted_stocks):
            row = 6 + idx
            ws[f'E{row}'] = stock_info['stock'].name[:25]  # Truncate long names
            ws[f'F{row}'] = round(stock_info['weightage'], 2)
            ws[f'G{row}'] = round(stock_info['metrics']['current_pe'], 2)
            ws[f'H{row}'] = round(stock_info['metrics']['yoy_growth'], 2)
        
        # Sector Allocation Section
        ws['E17'] = "SECTOR ALLOCATION"
        ws['E17'].font = Font(bold=True, size=12)
        ws.merge_cells('E17:G17')
        
        ws['E18'] = "Sector"
        ws['F18'] = "Weight (%)"
        ws['G18'] = "# Stocks"
        
        for col in ['E18', 'F18', 'G18']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Calculate sector counts
        sector_counts = {}
        for stock_info in stock_data:
            sector = stock_info['stock'].sector or 'Others'
            if sector not in sector_counts:
                sector_counts[sector] = 0
            sector_counts[sector] += 1
        
        # Sort sectors by allocation
        sorted_sectors = sorted(sector_allocation.items(), key=lambda x: x[1], reverse=True)
        
        for idx, (sector, weight) in enumerate(sorted_sectors[:8]):  # Top 8 sectors
            row = 19 + idx
            ws[f'E{row}'] = sector[:20]  # Truncate long sector names
            ws[f'F{row}'] = round(weight, 2)
            ws[f'G{row}'] = sector_counts.get(sector, 0)
        
        # Add borders to sections
        self._add_borders(ws, 5, 7, 1, 3)   # Portfolio Overview
        self._add_borders(ws, 10, 13, 1, 3)  # Valuation Metrics
        self._add_borders(ws, 16, 19, 1, 3)  # Growth Metrics
        self._add_borders(ws, 5, 15, 5, 8)   # Top Holdings
        self._add_borders(ws, 18, 18 + len(sorted_sectors[:8]), 5, 7)  # Sector Allocation
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 3
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        
        # Add Key Insights Section at the bottom
        ws['A22'] = "KEY INSIGHTS"
        ws['A22'].font = Font(bold=True, size=12)
        ws.merge_cells('A22:H22')
        
        insights = []
        
        # Generate insights based on metrics
        if weighted_pe > 25:
            insights.append("• Portfolio is trading at premium valuations (PE > 25)")
        elif weighted_pe < 15:
            insights.append("• Portfolio appears to be attractively valued (PE < 15)")
        
        if avg_cagr > 15:
            insights.append("• Strong historical growth with 6Y CAGR > 15%")
        
        if total_weightage < 100:
            insights.append(f"• Portfolio has {round(100 - total_weightage, 2)}% cash/unallocated")
        
        if len(sorted_sectors) == 1:
            insights.append("• Portfolio is concentrated in a single sector")
        elif len(sorted_sectors) > 5:
            insights.append("• Well-diversified portfolio across multiple sectors")
        
        if avg_alpha > 0:
            insights.append(f"• Portfolio generating positive alpha of {round(avg_alpha, 2)}% over bond rate")
        
        for idx, insight in enumerate(insights[:5]):  # Limit to 5 insights
            ws[f'A{23 + idx}'] = insight
            ws[f'A{23 + idx}'].font = Font(size=10)
        
        logger.info(f"Created Summary sheet with {total_stocks} stocks and {len(sorted_sectors)} sectors")
    
    def _get_header_style(self):
        """Get header cell style"""
        style = NamedStyle(name="header")
        style.font = Font(bold=True, size=11)
        style.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        style.alignment = Alignment(horizontal='center', vertical='center')
        style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        return style
    
    def _get_data_style(self):
        """Get data cell style"""
        style = NamedStyle(name="data")
        style.font = Font(size=10)
        style.alignment = Alignment(horizontal='left', vertical='center')
        style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        return style
    
    def _get_total_style(self):
        """Get total row style"""
        style = NamedStyle(name="total")
        style.font = Font(bold=True, size=11)
        style.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        style.alignment = Alignment(horizontal='right', vertical='center')
        style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='double'),
            bottom=Side(style='double')
        )
        return style
    
    def _add_borders(self, ws, start_row, end_row, start_col, end_col):
        """Add borders to a range of cells"""
        thin = Side(style='thin')
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if not cell.border:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    def _save_report(self, fund_name):
        """Save the generated report"""
        # Create filename - IMPORTANT: Use .xlsx extension for regular Excel files
        safe_fund_name = "".join(c for c in fund_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_fund_name = safe_fund_name.replace(' ', '_')
        current_date = datetime.now().strftime('%Y%m%d')
        filename = f"{safe_fund_name}_MF_Analysis_{current_date}.xlsx"  # Changed to .xlsx
        
        # Create temporary file
        temp_dir = tempfile.gettempdir()
        output_path = os.path.join(temp_dir, filename)
        
        # Save the workbook as .xlsx format
        try:
            self.workbook.save(output_path)
            logger.info(f"Excel file saved successfully as .xlsx: {output_path}")
        except Exception as e:
            logger.error(f"Error saving Excel file: {str(e)}")
            raise
        
        return output_path