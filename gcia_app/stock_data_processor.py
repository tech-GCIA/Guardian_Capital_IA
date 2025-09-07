# Create a new file: gcia_app/stock_data_processor.py

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
import re
from django.db import transaction
from django.utils import timezone
from gcia_app.models import Stock, StockQuarterlyData, StockUploadLog
import logging

logger = logging.getLogger(__name__)

class StockDataProcessor:
    """
    Processes Excel files containing stock data and updates the database
    """
    
    def __init__(self, user):
        self.user = user
        self.stats = {
            'stocks_added': 0,
            'stocks_updated': 0,
            'quarterly_records_added': 0,
            'quarterly_records_updated': 0,
            'errors': []
        }
    
    def process_excel_file(self, excel_file, upload_log):
        """
        Main method to process the uploaded Excel file
        """
        start_time = timezone.now()
        
        try:
            upload_log.status = 'processing'
            upload_log.processing_started_at = start_time
            upload_log.save()
            
            # Load the workbook
            workbook = load_workbook(excel_file, data_only=True)
            
            # Look for the App-Base Sheet
            app_base_sheet = self._find_app_base_sheet(workbook)
            if not app_base_sheet:
                raise ValueError("App-Base Sheet not found in the uploaded file")
            
            # Process the data
            with transaction.atomic():
                self._process_app_base_sheet(app_base_sheet)
            
            # Update upload log with success
            upload_log.status = 'completed'
            upload_log.stocks_added = self.stats['stocks_added']
            upload_log.stocks_updated = self.stats['stocks_updated']
            upload_log.quarterly_records_added = self.stats['quarterly_records_added']
            upload_log.quarterly_records_updated = self.stats['quarterly_records_updated']
            
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            upload_log.status = 'failed'
            upload_log.error_message = str(e)
            self.stats['errors'].append(str(e))
            raise
        
        finally:
            end_time = timezone.now()
            upload_log.processing_completed_at = end_time
            upload_log.processing_time = end_time - start_time
            upload_log.save()
        
        return self.stats
    
    def _find_app_base_sheet(self, workbook):
        """
        Find the App-Base Sheet in the workbook
        """
        # Look for exact match first
        if "App-Base Sheet" in workbook.sheetnames:
            return workbook["App-Base Sheet"]
        
        # Common names for the app-base sheet
        possible_names = [
            'App-Base Sheet', 'App Base Sheet', 'AppBase Sheet', 
            'App-Base', 'AppBase', 'Base Sheet', 'Stock Data', 'Data'
        ]
        
        for sheet_name in workbook.sheetnames:
            if any(name.lower() in sheet_name.lower() for name in possible_names):
                return workbook[sheet_name]
        
        # If not found, use the first sheet
        if workbook.sheetnames:
            return workbook[workbook.sheetnames[0]]
        
        return None
    
    def _process_app_base_sheet(self, sheet):
        """
        Process the App-Base Sheet data based on the actual structure
        """
        # Convert sheet to list of lists for easier processing
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        
        if len(data) < 10:
            raise ValueError("Sheet appears to be empty or has insufficient data")
        
        # Based on analysis, data starts from row 8 (index 7) for headers
        # and row 9 (index 8) for actual data
        header_row_index = 7  # Row 8 in Excel
        data_start_index = 9   # Row 10 in Excel (skip the sample row)
        
        # Process each data row
        for row_index in range(data_start_index, len(data)):
            try:
                row_data = data[row_index]
                if not row_data or not any(cell for cell in row_data[:10] if cell):
                    continue  # Skip empty rows
                
                self._process_stock_row(row_data, row_index + 1)  # +1 for Excel row number
                
            except Exception as e:
                error_msg = f"Error processing row {row_index + 1}: {str(e)}"
                logger.warning(error_msg)
                self.stats['errors'].append(error_msg)
    
    def _process_stock_row(self, row_data, excel_row_num):
        """
        Process a single stock row from the Excel data
        Based on the Excel structure analysis:
        Column 1: Company Name
        Column 2: Accord Code (will use as symbol)
        Column 3: Sector
        Column 4: Cap (market cap category)
        Column 5: Free Float
        """
        # Extract basic stock information
        company_name = str(row_data[1]).strip() if row_data[1] else None
        accord_code = str(row_data[2]).strip() if row_data[2] else None
        sector = str(row_data[3]).strip() if row_data[3] else None
        cap_category = str(row_data[4]).strip() if row_data[4] else None
        free_float = self._convert_to_decimal(row_data[5]) if len(row_data) > 5 else None
        
        # Skip if essential data is missing
        if not company_name or not accord_code or company_name.startswith("Sample"):
            return
        
        # Create stock symbol from accord code
        symbol = f"STOCK_{accord_code}"
        
        # Prepare stock data
        stock_data = {
            'name': company_name,
            'symbol': symbol,
            'sector': sector if sector and sector != 'Sample' else None,
            'market_cap_category': cap_category if cap_category and cap_category != 'Sample' else None,
            'is_active': True
        }
        
        # Get or create stock
        stock, created = Stock.objects.get_or_create(
            symbol=symbol,
            defaults=stock_data
        )
        
        if created:
            self.stats['stocks_added'] += 1
            logger.info(f"Added new stock: {company_name} ({symbol})")
        else:
            # Update existing stock data
            updated = False
            for field, value in stock_data.items():
                if field != 'symbol' and hasattr(stock, field):
                    current_value = getattr(stock, field)
                    if current_value != value and value is not None:
                        setattr(stock, field, value)
                        updated = True
            
            if updated:
                stock.save()
                self.stats['stocks_updated'] += 1
                logger.info(f"Updated stock: {company_name} ({symbol})")
        
        # Process quarterly financial data
        self._process_quarterly_financial_data(stock, row_data)
    
    def _process_quarterly_financial_data(self, stock, row_data):
        """
        Enhanced processing of quarterly financial data for the stock
        Based on comprehensive Excel structure to populate ALL required fields for MF metrics
        
        Expected Excel structure (approximate column positions):
        - Market Cap data: columns 14-41
        - Free Float data: columns 42-69  
        - TTM Revenue: columns 74-108 
        - Quarterly Revenue: columns 108-142
        - TTM PAT: columns 142-176
        - Quarterly PAT: columns 176-210
        - PE ratios: columns 210-244
        - P/R ratios: columns 244-278  
        - ROCE data: columns 278-312
        - ROE data: columns 312-346
        """
        
        # Enhanced quarter mappings with all financial metrics
        quarter_mappings = [
            # Recent quarters first (Latest data) - estimated column positions
            {'year': 2025, 'quarter': 1, 'mcap_col': 14, 'free_float_col': 42, 'ttm_rev_col': 74, 'q_rev_col': 108, 'ttm_pat_col': 142, 'q_pat_col': 176, 'pe_col': 210, 'pr_col': 244, 'roce_col': 278, 'roe_col': 312},
            {'year': 2024, 'quarter': 4, 'mcap_col': 15, 'free_float_col': 43, 'ttm_rev_col': 75, 'q_rev_col': 109, 'ttm_pat_col': 143, 'q_pat_col': 177, 'pe_col': 211, 'pr_col': 245, 'roce_col': 279, 'roe_col': 313},
            {'year': 2024, 'quarter': 3, 'mcap_col': 16, 'free_float_col': 44, 'ttm_rev_col': 76, 'q_rev_col': 110, 'ttm_pat_col': 144, 'q_pat_col': 178, 'pe_col': 212, 'pr_col': 246, 'roce_col': 280, 'roe_col': 314},
            {'year': 2024, 'quarter': 2, 'mcap_col': 17, 'free_float_col': 45, 'ttm_rev_col': 77, 'q_rev_col': 111, 'ttm_pat_col': 145, 'q_pat_col': 179, 'pe_col': 213, 'pr_col': 247, 'roce_col': 281, 'roe_col': 315},
            {'year': 2024, 'quarter': 1, 'mcap_col': 18, 'free_float_col': 46, 'ttm_rev_col': 78, 'q_rev_col': 112, 'ttm_pat_col': 146, 'q_pat_col': 180, 'pe_col': 214, 'pr_col': 248, 'roce_col': 282, 'roe_col': 316},
            {'year': 2023, 'quarter': 4, 'mcap_col': 19, 'free_float_col': 47, 'ttm_rev_col': 79, 'q_rev_col': 113, 'ttm_pat_col': 147, 'q_pat_col': 181, 'pe_col': 215, 'pr_col': 249, 'roce_col': 283, 'roe_col': 317},
            {'year': 2023, 'quarter': 3, 'mcap_col': 20, 'free_float_col': 48, 'ttm_rev_col': 80, 'q_rev_col': 114, 'ttm_pat_col': 148, 'q_pat_col': 182, 'pe_col': 216, 'pr_col': 250, 'roce_col': 284, 'roe_col': 318},
            {'year': 2023, 'quarter': 2, 'mcap_col': 21, 'free_float_col': 49, 'ttm_rev_col': 81, 'q_rev_col': 115, 'ttm_pat_col': 149, 'q_pat_col': 183, 'pe_col': 217, 'pr_col': 251, 'roce_col': 285, 'roe_col': 319},
            {'year': 2023, 'quarter': 1, 'mcap_col': 22, 'free_float_col': 50, 'ttm_rev_col': 82, 'q_rev_col': 116, 'ttm_pat_col': 150, 'q_pat_col': 184, 'pe_col': 218, 'pr_col': 252, 'roce_col': 286, 'roe_col': 320},
            {'year': 2022, 'quarter': 4, 'mcap_col': 23, 'free_float_col': 51, 'ttm_rev_col': 83, 'q_rev_col': 117, 'ttm_pat_col': 151, 'q_pat_col': 185, 'pe_col': 219, 'pr_col': 253, 'roce_col': 287, 'roe_col': 321},
        ]
        
        # Process each quarter's data
        for quarter_info in quarter_mappings:
            try:
                year = quarter_info['year']
                quarter_num = quarter_info['quarter']
                
                # Extract ALL financial data for this quarter
                financial_data = {}
                
                # Market Cap and Free Float data
                financial_data['mcap'] = self._safely_extract_value(row_data, quarter_info['mcap_col'])
                financial_data['free_float_mcap'] = self._safely_extract_value(row_data, quarter_info['free_float_col'])
                
                # Revenue data (both TTM and quarterly)
                financial_data['ttm_revenue'] = self._safely_extract_value(row_data, quarter_info['ttm_rev_col'])
                financial_data['quarterly_revenue'] = self._safely_extract_value(row_data, quarter_info['q_rev_col'])
                
                # PAT data (both TTM and quarterly)
                financial_data['ttm_pat'] = self._safely_extract_value(row_data, quarter_info['ttm_pat_col'])
                financial_data['quarterly_pat'] = self._safely_extract_value(row_data, quarter_info['q_pat_col'])
                
                # Ratio data
                financial_data['pe_quarterly'] = self._safely_extract_value(row_data, quarter_info['pe_col'])
                financial_data['pr_quarterly'] = self._safely_extract_value(row_data, quarter_info['pr_col'])
                
                # Profitability ratios
                financial_data['roce'] = self._safely_extract_value(row_data, quarter_info['roce_col'])
                financial_data['roe'] = self._safely_extract_value(row_data, quarter_info['roe_col'])
                
                # Legacy fields (for backwards compatibility)
                financial_data['revenue'] = financial_data['ttm_revenue']  # Legacy field
                financial_data['pat'] = financial_data['quarterly_pat'] or financial_data['ttm_pat']  # Prefer quarterly
                financial_data['ttm'] = financial_data['ttm_revenue']  # Legacy field
                
                # Only create record if we have meaningful financial data
                has_data = any(v is not None for v in financial_data.values())
                
                if has_data:
                    quarter_date = self._get_quarter_end_date(year, quarter_num)
                    
                    # Prepare defaults with all financial data
                    defaults = {
                        'quarter_date': quarter_date,
                        **financial_data  # Unpack all financial data
                    }
                    
                    # Create or update quarterly data
                    quarterly_data, created = StockQuarterlyData.objects.get_or_create(
                        stock=stock,
                        quarter_year=year,
                        quarter_number=quarter_num,
                        defaults=defaults
                    )
                    
                    if created:
                        self.stats['quarterly_records_added'] += 1
                        logger.debug(f"Created Q{quarter_num} {year} data for {stock.name}")
                    else:
                        # Update existing record with comprehensive data
                        updated = False
                        
                        for field, value in financial_data.items():
                            if value is not None and hasattr(quarterly_data, field):
                                current_value = getattr(quarterly_data, field)
                                if current_value != value:
                                    setattr(quarterly_data, field, value)
                                    updated = True
                        
                        if updated:
                            quarterly_data.save()
                            self.stats['quarterly_records_updated'] += 1
                            logger.debug(f"Updated Q{quarter_num} {year} data for {stock.name}")
                            
            except Exception as e:
                error_msg = f"Error processing Q{quarter_num} {year} for {stock.name}: {str(e)}"
                logger.warning(error_msg)
                self.stats['errors'].append(error_msg)
    
    def _safely_extract_value(self, row_data, column_index):
        """
        Safely extract and convert a value from the specified column
        """
        try:
            if column_index < len(row_data):
                return self._convert_to_decimal(row_data[column_index])
            return None
        except Exception as e:
            logger.debug(f"Error extracting value from column {column_index}: {e}")
            return None
    
    def _convert_to_decimal(self, value):
        """
        Convert various number formats to Decimal
        """
        if value is None or value == '' or value == 0:
            return None
        
        try:
            # Handle string values that might contain commas, currency symbols, etc.
            if isinstance(value, str):
                # Remove common currency symbols and formatting
                cleaned = re.sub(r'[₹$,\s]', '', value.strip())
                # Handle parentheses for negative numbers
                if cleaned.startswith('(') and cleaned.endswith(')'):
                    cleaned = '-' + cleaned[1:-1]
                
                if not cleaned or cleaned == '-' or cleaned.lower() == 'sample':
                    return None
                
                return Decimal(cleaned)
            
            # Handle numeric values
            return Decimal(str(value))
            
        except (ValueError, InvalidOperation, TypeError):
            return None
    
    def _get_quarter_end_date(self, year, quarter):
        """
        Get the end date for a given quarter and year
        """
        quarter_end_months = {1: 3, 2: 6, 3: 9, 4: 12}  # March, June, Sep, Dec
        month = quarter_end_months[quarter]
        
        # Get last day of the month
        if month == 12:
            next_month = datetime(year + 1, 1, 1)
        else:
            next_month = datetime(year, month + 1, 1)
        
        last_day = next_month - timedelta(days=1)
        return last_day.date()


def process_stock_data_file(excel_file, user):
    """
    Main function to process stock data Excel file
    """
    # Create upload log entry
    upload_log = StockUploadLog.objects.create(
        uploaded_by=user,
        filename=excel_file.name,
        file_size=excel_file.size,
        status='pending'
    )
    
    try:
        processor = StockDataProcessor(user)
        stats = processor.process_excel_file(excel_file, upload_log)
        return stats, upload_log
    except Exception as e:
        logger.error(f"Failed to process stock data file: {str(e)}")
        raise